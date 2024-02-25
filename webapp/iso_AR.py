from webapp import db
from flask import render_template, flash, redirect, url_for, session, logging, request
from requests import get
from CCC_system_setup import apikeys
from CCC_system_setup import myoslist, addpath, tpath, companydata, usernames, passwords, scac, imap_url, accessorials, signoff
from webapp.viewfuncs import d2s, stat_update, hasinput, d1s
#from viewfuncs import d2s, d1s
import imaplib, email
import math
import re
from email.header import decode_header
import webbrowser
import os
from email.utils import parsedate_tz, mktime_tz
from email.utils import parsedate_to_datetime
from bs4 import BeautifulSoup
from sqlalchemy.sql import desc
import ast
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter

import datetime
from webapp.models import Quotes, Quoteinput, Orders, People, Ardata
from send_mimemail import send_mimemail
from pyzipcode import ZipCodeDatabase
from webapp.class8_utils_email import html_mimemail
import json

zcdb = ZipCodeDatabase()

API_KEY_GEO = apikeys['gkey']
API_KEY_DIS = apikeys['dkey']
cdata = companydata()

date_y4=re.compile(r'([1-9]|0[1-9]|[12][0-9]|3[01]) (Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) (\d{4})')

today_now = datetime.datetime.now()
today = today_now.date()
timenow = today_now.time()

def roundup(x):
    return int(math.ceil(x / 10.0)) * 10

def get_body(msg):
    if msg.is_multipart():
        return get_body(msg.get_payload(0))
    else:
        return msg.get_payload(None,True)

def search(key,value,con):
    result,data=con.search(None,key,'"{}"'.format(value))
    return data

def search_from_date(key,value,con,datefrom):
    result,data=con.search( None, '(SENTSINCE {0})'.format(datefrom) , key, '"{}"'.format(value) )
    return data

def get_emails(result_bytes,con):
    msgs=[]
    for num in result_bytes[0].split():
        typ,data=con.fetch(num,'(RFC822)')
        msgs.append(data)
    return msgs

def get_date(data):
    for response_part in data:
        if isinstance(response_part, tuple):
            try:
                part = response_part[1].decode('utf-8')
                msg = email.message_from_string(part)
                date=msg['Date']
            except:
                date=None
    return date

def get_subject(data):
    subject = 'none'
    mid = 'none'
    for response_part in data:
        if isinstance(response_part, tuple):
            part = response_part[1].decode('utf-8')
            msg = email.message_from_string(part)
            subject=msg['Subject']
            mid = msg['Message-ID']
    return subject, mid

def get_from(data):
    for response_part in data:
        if isinstance(response_part, tuple):
            part = response_part[1]
            try:
                part = part.decode('utf-8')
                msg = email.message_from_string(part)
                thisfrom=msg['From']
                return thisfrom
            except:
                return 'Nonefound'

def get_msgs():
    username = usernames['quot']
    password = passwords['quot']
    dayback = 100
    #datefrom = (datetime.date.today() - datetime.timedelta(dayback)).strftime("%d-%b-%Y")
    con = imaplib.IMAP4_SSL(imap_url)
    con.login(username, password)
    con.select('INBOX')
    result, data = con.search(None,'ALL')
    msgs = get_emails(data, con)
    return msgs


def compact(body):
    newbody = ''
    blines = body.splitlines()
    for line in blines:
        #Remove non-ascii characters
        line = re.sub(r'[^\x00-\x7F]+', ' ', line)
        line = re.sub(r'=[A-Z,0-9][A-Z,0-9]', '', line)
        if len(line.strip())>1:
            if not 'Forwarded Message' in line or 'Subject:' in line or 'Date:' in line or 'To:' in line or 'CC:' in line or 'From:' in line or 'Content-Type' in line or 'Content-Transfer' in line:
                newbody = newbody + line +'\n' + '<br>'
    return newbody

def hard_decode(raw):
    raw = str(raw)
    rawl = raw.splitlines()
    appendit = 0
    ebody=''
    efrom=''
    edate=''
    mid=''
    for line in rawl:
        test = line[0:5]
        line = re.sub(r'[^\x00-\x7F]+', ' ', line)
        line = re.sub(r'=[A-Z,0-9][A-Z,0-9]', '', line)
        line = line.replace('=09','')
        if 'Subj' in test:
            subject = line.split('Subject:')[1]
            subject = subject.replace('Fwd:','')
            subject = subject.strip()
            #print(f'Subject:{subject}')
        if 'Message-ID' in line:
            mid = line.split('Message-ID:')[1]
            mid = mid.replace('Fwd:', '')
            mid = mid.strip()
            #print(f'MID:{mid}')
        if 'From' in test and '@' in line and 'firsteagle' not in line and 'onestop' not in line:
            #print('efrom',line)
            efrom = line.split('From:')[1]
            efrom = efrom.strip()
            #print(f'From:{efrom}')
        if 'Date' in test:
            edate = line.split('Date:')[1]
            edate = edate.strip()
            #print(f'Date:{edate}')
        if 'Content-Type:' in line and 'plain' in line:
            #print(f'BodyStart:{line}')
            appendit = 1
        if 'Content-Type:' in line and 'html' in line:
            #print(f'BodyStop:{line}')
            appendit = 0
        if appendit == 1:
            line = line.strip()
            if len(line)>0:
                ebody=ebody+line+'\n'
    #print(f'ebody={ebody}')
    return subject,efrom,edate,ebody,mid

def clean(text):
    # clean text for creating a folder
    return "".join(c if c.isalnum() else "_" for c in text)

def extract_for_code(data):
    try:
        text, encoding = decode_header(data)[0]
    except:
        text = 'No Decode Available'
        encoding = None
    if isinstance(text, bytes):
        if encoding is not None: text = text.decode(encoding)
    return text

def get_body_text(qdat):

    mid = qdat.Mid
    #print(f'this mid is {mid}')
    username = usernames['quot']
    password = passwords['quot']
    imap = imaplib.IMAP4_SSL(imap_url)
    imap.login(username, password)
    status, messages = imap.select('INBOX')
    try:
        result, data = imap.search(None, f'HEADER Message-ID {mid}')
        msg_id_list = data[0].split()
        result, data = imap.fetch(msg_id_list[0], '(RFC822)')
        email_message = email.message_from_bytes(data[0][1])

        # extract the subject of the email
        subject = extract_for_code(email_message["Subject"])
        #print(f'****Getting the Body Text***** for Subject: {subject}')
    except:
        #print('Could not locate this email header')
        return 'Email ID not found', None

    # Set default text particulars
    plain_text_content = ''
    html_content = None

    # extract the email content as a string
    if email_message.is_multipart():
        for part in email_message.walk():
            if part.get_content_type() == 'text/plain':
                try:
                    plain_text_content = part.get_payload(decode=True).decode('utf-8')
                except UnicodeDecodeError as e:
                    plain_text_content = part.get_payload(decode=True).decode('utf-8', errors='replace')
            if part.get_content_type() == "text/html":
                try:
                    html_content = part.get_payload(decode=True).decode("utf-8")
                except UnicodeDecodeError as e:
                    html_content = part.get_payload(decode=True).decode("utf-8", errors='replace')

                soup = BeautifulSoup(html_content, "html.parser")
                plain_text_content = soup.get_text()
    else:
        try:
            plain_text_content = email_message.get_payload(decode=True).decode('utf-8')
        except:
            plain_text_content = 'Could not decode payload'

    #print('Returning from get_body_text', plain_text_content)
    return plain_text_content, html_content

def get_sorted_cust(arorders):
    ucust = []
    for aro in arorders:
        shipper = aro.Shipper
        if shipper not in ucust: ucust.append(shipper)
    ucust.sort()
    return ucust

def get_open_sort_totals(arlist):
    dat30 = today - datetime.timedelta(30)
    lb360 = today - datetime.timedelta(360)
    cdata = []
    for cust in arlist:
        odata = Orders.query.filter((Orders.Shipper == cust) & ((Orders.Istat>1) & ((Orders.Istat<5) | (Orders.Istat==7))) & (Orders.InvoDate>lb360)).order_by(Orders.InvoDate).all()
        iall, iu30, io30 = 0, 0, 0
        dolall, dolu30, dolo30 = 0.00, 0.00, 0.00
        for odat in odata:
            invodate = odat.InvoDate
            #iif has an invoice date then it has been invoiced
            #print(f'{dat30} and {invodate}')
            invototal = odat.InvoTotal
            if invodate is not None and invototal is not None:
                invototal = float(invototal)
                iall += 1
                dolall += invototal
                if invodate < dat30:
                    io30 += 1
                    dolo30 += invototal
                else:
                    iu30 += 1
                    dolu30 += invototal

        cdata.append([cust,io30, d2s(dolo30), iu30, d2s(dolu30), iall, d2s(dolall)])
    return cdata

def get_open_for_cust(this_shipper):
    dat30 = today - datetime.timedelta(30)
    lb360 = today - datetime.timedelta(360)
    odata = Orders.query.filter((Orders.Shipper == this_shipper) & ((Orders.Istat>1) & ((Orders.Istat<5) | (Orders.Istat==7))) & (Orders.Date3>lb360) & (Orders.InvoTotal != None)).order_by(Orders.Date3).all()
    adata = Ardata.query.filter((Ardata.Customer == this_shipper) & (Ardata.Date1 > lb360) & (Ardata.Emailtype == 'Direct')).all()
    tdata = []
    for adat in adata:
        jol= adat.Jolist
        if jol is not None and isinstance(jol, str):
            try:
                jolist = ast.literal_eval(jol)
            except:
                #print(f'Could not form ast for adat {adat.id}')
                jolist = []
        else:
            jolist = []
        for jo in jolist:
            jo = str(jo)
            odat = Orders.query.filter(Orders.Jo == jo).first()
            if odat is not None:
                tdata.append([f'{adat.Date1}', adat.Emailtype, adat.Mid, adat.Emailto, adat.Etitle])
                break

    return odata, tdata

def read_tboxes():
    tboxes = [0]*30
    for ix in range(30):
        tboxes[ix] = request.values.get(f'tbox{ix}')
    #print(f'the tboxes here are {tboxes}')
    return tboxes

def attach_rename_inv(odat, name):
    newname = odat.Invoice
    if name == 'Invoice_Container':  newname = f'Invoice_{odat.Container}.pdf'
    elif name == 'Invoice_Order_xxx': newname = f'Invoice_Order_{odat.Order}.pdf'
    elif name == 'Invoice_Release_xxx': newname = f'Invoice_Release_{odat.Booking}.pdf'
    elif name == 'Inv_Booking_Container': newname = f'Inv_{odat.Booking}_{odat.Container}.pdf'
    elif name == 'Inv_Order_Container': newname = f'Inv_{odat.Order}_{odat.Container}.pdf'
    return newname

def attach_rename_pack(odat, name):
    newname = odat.Package
    if newname is not None:
        if 'SI' in newname:
            pass
        else:
            if name == 'Inv_Package_Container':  newname = f'Inv_Package_{odat.Container}.pdf'
            elif name == 'Inv_Package_Order_xxx': newname = f'Inv_Package_Order_{odat.Order}.pdf'
            elif name == 'Inv_Package_Release_xxx': newname = f'Inv_Package_Release_{odat.Booking}.pdf'
            elif name == 'Inv_Package_Booking_Container': newname = f'Inv_Package_{odat.Booking}_{odat.Container}.pdf'
            elif name == 'Inv_Package_Order_Container': newname = f'Inv_Package_{odat.Order}_{odat.Container}.pdf'
    return newname

def column_wide(headers,ydata):
    column_widths = []
    for cell in enumerate(headers):
        cell = str(cell)
        column_widths.append(len(cell))
    for row in ydata:
        for i, cell in enumerate(row):
            test = len(str(cell))
            column_widths[i] = max(test,column_widths[i])
    return column_widths

def make_workbook(customer, data, tboxes, ftotal):
    # if this flag on we will also create a new workbook to attach
    wb = openpyxl.Workbook()
    shtlist = wb.sheetnames
    for sht in shtlist:
        std = wb.get_sheet_by_name(sht)
        wb.remove_sheet(std)
    newsheet = f'{customer}_Open_{today}'
    dfc = wb.create_sheet(newsheet)
    # formats for writing to excel
    money = '$#,##0.00'
    dec2 = '#,##0.00'
    dec0 = '#,##0'
    hdrs = ['JO', 'Order/Summary', 'Booking In', 'Container', 'Date Invoiced', 'Amount', 'Date Due']
    keephdrs = []
    for jx in range(7):
        if tboxes[jx] == 'on':
            keephdrs.append(hdrs[jx])
    for col, hdr in enumerate(keephdrs):
        d = dfc.cell(row=1, column=col + 1, value=hdr)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)

    for ix, dat in enumerate(data):
        #print(f'dat is {dat} keephdrs is {keephdrs}')
        for jx, each in enumerate(dat):
            d = dfc.cell(row=ix+2, column=jx + 1, value=each)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            if keephdrs[jx] == 'Amount':
                d.number_format = money
                amtcol = jx+1
                rowtot = ix+4

    d = dfc.cell(row=rowtot, column=amtcol-1, value='Total:')
    d.alignment = Alignment(horizontal='right')
    d.font = Font(name='Calibri', size=10, bold=True)

    d = dfc.cell(row=rowtot, column=amtcol, value=ftotal)
    d.alignment = Alignment(horizontal='center')
    d.font = Font(name='Calibri', size=10, bold=True)
    d.number_format = money

    column_widths = column_wide(keephdrs, [])
    for i, column_width in enumerate(column_widths):
        dfc.column_dimensions[get_column_letter(i + 1)].width = column_width + 4


    wbpath = addpath(f'static/{scac}/data/temp/{customer}_open_{today}')
    wbpath = wbpath.replace(' ','_')
    wbpath = wbpath.replace('.','')
    wbpath = wbpath.replace('-','')
    wbpath = wbpath + '.xlsx'
    wbfile = os.path.basename(wbpath)


    #print(f'the wbfile is {wbfile} and the wbpath is: {wbpath}')
    wb.save(wbpath)

    return wbfile


def get_table_formatted(odata, etype, tboxes, boxes, make_wb, customer):
    intable='<table><tr>'
    labels = ['JO', 'Order', 'Release', 'Container', 'Invoice Date', 'Amount', 'Due Date']
    align = ["center", "center", "center", "center", "center", "right", "center"]
    ftotal = 0.00
    invoices = []
    packages = []
    new_invoices = []
    new_packages = []
    invoname = request.values.get('invoname')
    packname = request.values.get('packname')
    ydata = []
    wbfile = None

    for jx in range(7):
        if tboxes[jx]=='on':
            intable = f'{intable}<td align={align[jx]}><b>{labels[jx]}</b></td>'
    intable = f'{intable}</tr><tr>'
    for ix, odat in enumerate(odata):
        if boxes[ix]=='on':
            datei = odat.InvoDate
            if datei is None: datei = odat.Date3

            ftotal = ftotal + float(odat.InvoTotal)
            duedate = datei + datetime.timedelta(30)
            odr = odat.Label
            if odr is None: odr = odat.Order
            data = [odat.Jo, odr, odat.Booking, odat.Container, f'{datei}', f'${odat.InvoTotal}', f'{duedate}']
            datline=[]
            intable = f'{intable}<tr>'
            for jx in range(7):
                if tboxes[jx]=='on':
                    datline.append(data[jx])
                    intable = f'{intable}<td align={align[jx]}>{data[jx]}</td>'
            intable = f'{intable}</tr>'
            ydata.append(datline)
            if tboxes[7] == 'on':
                newinvo = odat.Invoice
                if newinvo is not None:
                    if newinvo not in invoices:
                        invoices.append(odat.Invoice)
                        new_invoices.append(attach_rename_inv(odat,invoname))
            if tboxes[8] == 'on':
                newpack = odat.Package
                if newpack is not None:
                    if newpack not in packages:
                        packages.append(newpack)
                        new_packages.append(attach_rename_pack(odat, packname))
    intable = f'{intable}</table>'

    if make_wb: wbfile = make_workbook(customer, ydata, tboxes, ftotal)

    num_invoices = len(ydata)

    return intable, ftotal, invoices, packages, new_invoices, new_packages, wbfile, num_invoices

def final_update_email(this_shipper, odata, tboxes, boxes, emailsend, email_update):
    cdata = companydata()
    dat30 = today - datetime.timedelta(30)
    etitle = request.values.get('etitle')
    ebody = request.values.get('ebody')
    salutation = request.values.get('salutation')
    efrom = usernames['invo']
    epass = passwords['invo']
    eto = request.values.get('etolist')
    ecc = request.values.get('ecclist')
    eto = ast.literal_eval(eto)
    ecc = ast.literal_eval(ecc)
    invoices = []
    packages = []
    new_invoices = []
    new_packages = []
    intnexti = 0
    intnextp = 0
    tone = tboxes[27]

    for ix, odat in enumerate(odata):
        if boxes[ix] == 'on':
            if tboxes[7] == 'on':
                next_invoice = odat.Invoice
                if next_invoice is not None:
                    if next_invoice not in invoices:
                        invoices.append(odat.Invoice)
                        nextinvo = request.values.get(f'edati{intnexti}')
                        intnexti += 1
                        new_invoices.append(nextinvo)
            if tboxes[8] == 'on':
                next_package = odat.Package
                if next_package is not None:
                    if next_package not in packages:
                        packages.append(next_package)
                        nextpack = request.values.get(f'edatp{intnextp}')
                        intnextp += 1
                        new_packages.append(nextpack)

    wba = request.values.get('wbattach')
    wbf = request.values.get('wbcreated')

    emaildata = [etitle, ebody, eto, ecc, efrom, epass, f'/static/{scac}/data/vInvoice/', dat30, invoices, packages, new_invoices, new_packages, salutation, wbf, wba, tone, f'/static/{scac}/data/vPackage/']

    return emaildata

def update_email(this_shipper, odata, tboxes, boxes, emailsend, email_update):
    cdata = companydata()
    dat30 = today - datetime.timedelta(30)

    #Items same regarless of email tone
    salutation = request.values.get('salutation')
    if not hasinput(salutation) or salutation == f'{this_shipper} Accounting':
        email_to_selected = emailsend[1]
        #print(email_to_selected)
        if email_to_selected is not None and email_to_selected != []:
            salutation = emailsend[4]
            if not hasinput(salutation):
                ets = email_to_selected[0]
                etslist = ets.split('@')
                salutation = etslist[0].title()
        else:
            salutation = f'{this_shipper} Accounting'
    company_info = f'{cdata[2]}<br>{cdata[8]}<br>{cdata[16]}<br><br>{cdata[13]}<br><br>{cdata[14]}<br><br>{cdata[15]}'



    etype = 'o30'
    efrom = usernames['invo']
    epass = passwords['invo']
    eto = emailsend[1]
    ecc = emailsend[3]
    invoices = []
    packages = []
    #ecc1 = usernames['expo']
    #ecc2 = usernames['info']
    if tboxes[9] == 'on': make_wb = 1
    else: make_wb = 0
    newwb = None

    table_in, ftotal, invoices, packages, new_invoices, new_packages, newwb, ndue = get_table_formatted(odata, etype, tboxes, boxes, make_wb, this_shipper)

    # Tone specific items:
    tone = tboxes[27]
    if tone == 'Light Reminder':
        etitle = 'Friendly Reminder:  Invoice Payments Due'
        closing = f'  Thank you for your cooperation.<br><br>Kind Regards,<br>Accounts Payable Team<br>{company_info}'
        ebody = f'Hi {salutation},<br><br>'
        if ndue == 1:
            ebody = f'{ebody} I hope this message finds you well. We appreciate your business and would like to remind you the invoice shown below has become due:'
        else:
            ebody = f'{ebody} I hope this message finds you well. We appreciate your business and would like to remind you the invoices shown below have become due:'
        ebody = f'{ebody}<br><br>{table_in}'
        ebody = f'{ebody}<br>Please make the necessary arrangements to ensure timely payment. If you have already processed the payment, we sincerely thank you for your prompt attention to this matter.'
        ebody = f'{ebody}<br><br>If there are any concerns or if you require any additional information, feel free to reach out to us. We value our partnership and are here to assist you.. {closing}'

        #For readability on preview put line breaks whereever there is a <br> then take them out before sending email...
        ebody = ebody.replace('<br><br>', '<br><br>\n\n')

    elif tone == 'Standard Request':
        etitle = f'Open Balance Report for {this_shipper} as of {today}'
        closing = f'  Thank you for your prompt attention to this matter.<br><br>Sincerely,<br>Accounts Payable Team<br>{company_info}'
        ebody = f'Hello {salutation},<br><br>'
        ebody = f'{ebody} We are showing a total balance due of ${d2s(ftotal)} as listed in the following table:'
        ebody = f'{ebody}<br><br>{table_in}'
        ebody = f'{ebody}<br>Please review this information and let us know when we can expect payment, or if you have sent payment already please indicate what has been paid so that we may review on our side. {closing}'
        #For readability on preview put line breaks whereever there is a <br> then take them out before sending email...
        ebody = ebody.replace('<br><br>', '<br><br>\n\n')

    elif tone == 'Strong Request':
        etitle = f'Open Balance Report - Overdue Notice - Urgent Attention Required'
        closing = f'  Thank you for your prompt attention to this matter.<br><br>Sincerely,<br>Accounts Payable Team<br>{company_info}'
        ebody = f'Dear {salutation},<br><br>'
        ebody = f'{ebody} We are showing a total balance due of ${d2s(ftotal)} as listed in the following table:'
        ebody = f'{ebody}<br><br>{table_in}'
        ebody = f'{ebody}<br>We have communicated about this matter previously, but have not yet received a payment.  Please review this information and let us know when we may expect resoluton of this matter.'
        ebody = f'{ebody}<br><br>If there are reasons for the delay please contact us immediately. {closing}'
        #For readability on preview put line breaks whereever there is a <br> then take them out before sending email...
        ebody = ebody.replace('<br><br>', '<br><br>\n\n')

    elif tone == 'Strongest Request':
        if ndue == 1:
            etitle = f'Outstanding Invoice - Urgent Attention Required'
            invtense = 'invoice remains'
            phrase2 = 'This is'
            phrase3 = 'date'
        else:
            etitle = f'Outstanding Invoices - Urgent Attention Required'
            invtense = 'invoices remain'
            phrase2 = 'These are all'
            phrase3 = 'dates'

        closing = f'  Thank you for your prompt attention to this matter.<br><br>Sincerely,<br>Accounts Payable Team<br>{company_info}'

        ebody = f'Dear {salutation},<br><br>'
        ebody = f'{ebody} I trust this note finds you well.  We appreciate your business and the opportunity to serve your needs.  \
                 However, it has come to our attention that the following {invtense} unpaid, despite our previous communications.'
        ebody = f'{ebody}<br><br>{table_in}'
        ebody = f'{ebody}<br>{phrase2} well past the due {phrase3}, and the outstanding amount of ${d2s(ftotal)} is causing a significant impact on our cash flow.  We understand\
        that unforeseen circumstances may arise, but we kindly request your immediate attention to this matter.'
        ebody = f'{ebody}<br><br>Previous attempts to collect this balance have failed.  We urge you to address this matter promply and sumit the payment by {today + datetime.timedelta(12)}.'
        ebody = f'{ebody}<br><br>If payment is not received by this specified deadline, we may have no alternative but to escalate this matter for collection.  This action could result in\
                 additional fees, damage to your credit rating, and potential legal proceedings.'
        ebody = f'{ebody}<br><br>We value our business relationship and would prefer to resolve this matter amicably.  Should you encounter any challenges or require assistance, please contact our accounts team as soon as possible. {closing}'

        #For readability on preview put line breaks whereever there is a <br> then take them out before sending email...
        ebody = ebody.replace('<br><br>', '<br><br>\n\n')

    else:
        etitle = 'No Specifice Title'
        ebody = f'Dear {salutation},<br><br>'


    emaildata = [etitle, ebody, eto, ecc, efrom, epass, f'/static/{scac}/data/vInvoice/', dat30, invoices, packages, new_invoices, new_packages, salutation, newwb, newwb, tone, f'/static/{scac}/data/vPackage/']
    #etitle, ebody, emailin1, emailin2, emailcc1, emailcc2, efrom, folder, dat30date, sourcenamelist, sendnamellist = emaildata
    return emaildata

def get_email_customer(pdat, ar_emails_cust):
    sal_default = None
    emailto_selected = request.values.getlist('emailtolist')
    emailcc_selected = request.values.getlist('emailcclist')
    #print(f'{emailto_selected}')
    if emailto_selected == []:
        #set a default value...
        emailto_selected = [pdat.Associate2]
        sal_default = pdat.Salap
    #print(f'{emailcc_selected}')
    emailtos, emailccs = [], []
    if hasinput(pdat.Email):
        emailtos.append(pdat.Email)
        emailccs.append(pdat.Email)
    if hasinput(pdat.Associate1):
        emailtos.append(pdat.Associate1)
        emailccs.append(pdat.Associate1)
    if hasinput(pdat.Associate2):
        emailtos.append(pdat.Associate2)
        emailccs.append(pdat.Associate2)

    # Add in email addresses from related emails
    for ar in ar_emails_cust:
        eto = ar.Emailto
        ecc = ar.Emailcc
        efrom = ar.From
        eto = ast.literal_eval(eto)
        ecc = ast.literal_eval(ecc)
        for et in eto:
            emailtos.append(et)
        for ec in ecc:
            emailccs.append(ec)
        if efrom is not None:
            emailtos.append(efrom)
            emailccs.append(efrom)

    #Add these emails for information purposes
    emailtos.append(usernames['info'])
    emailccs.append(usernames['info'])

    unique_emailtolist = set(emailtos)
    unique_emailcclist = set(emailccs)
    emailtos = list(unique_emailtolist)
    emailccs = list(unique_emailcclist)
    emailsend = [emailtos, emailto_selected, emailccs, emailcc_selected, sal_default]
    return emailsend

def get_ardata(containerlist, tboxes, customer):
    ardata_all = []
    for container in containerlist:
        if tboxes[22]=='on':
            ardata = Ardata.query.filter((Ardata.Container == container) & (Ardata.Emailtype == 'Invoice')).all()
            if ardata: ardata_all = ardata_all + ardata
        if tboxes[23]=='on':
            ardata = Ardata.query.filter((Ardata.Container == container) & (Ardata.Emailtype == 'Invoice Response')).all()
            if ardata: ardata_all = ardata_all + ardata

    if tboxes[24]=='on':
        ardata = Ardata.query.filter((Ardata.Customer == customer) & ((Ardata.Emailtype == 'Report') | (Ardata.Emailtype == 'Direct'))).all()
        if ardata: ardata_all = ardata_all + ardata
    if tboxes[25]=='on':
        ardata = Ardata.query.filter((Ardata.Customer == customer) & (Ardata.Emailtype == 'Report Response')).all()
        if ardata: ardata_all = ardata_all + ardata
    return ardata_all

def rselect(nemail):
    rview = [0]*nemail
    rtest = request.values.get('email_radio')
    if rtest is not None:
        try:
            rtest = int(rtest)
            rview[rtest] = 'on'
        except:
            rtest = 0
    #print(f'rtest = {rtest}')
    return rview

def shortbody(jolist, containerlist):
    text = 'The original email body is too long to store here, but the following jo and containers were referenced:'
    jolist = ast.literal_eval(jolist)
    for ix, jo in enumerate(jolist):
        #print(jo, containerlist[ix])
        text = f'{text}<br>{jo}  {containerlist[ix]}'
    return text

def ardata_email_update(emaildata, shipper, jolist, containerlist):

    etitle, ebody, emailin, emailcc, username, password, folder1, dat30date, invoices, packages, ni, np, salutation, wbfile, wbattach, tone, folder2 = emaildata
    sentfiles = []
    sentasfiles = []
    for inv in invoices:
        sentfiles.append(inv)
    for pac in packages:
        sentfiles.append(pac)
    for inv in ni:
        sentasfiles.append(inv)
    for pac in np:
        sentasfiles.append(pac)

    itype = 'Direct'
    #print(jolist)
    #print(containerlist)
    #print(today)
    #print(emailin)
    #print(emailcc)
    ncl = f'{containerlist}'
    jl = f'{jolist}'
    sf = f'{sentfiles}'
    saf = f'{sentasfiles}'
    if len(ebody) > 5998:
        try:
            ebody = shortbody(jl, containerlist)
        except:
            ebody = 'Email too long to store'
    if len(jl) > 499: jl = None
    if len(ncl) > 499: ncl = None
    if len(sf) > 999: sf = None
    if len(saf) > 999: saf = None


    input = Ardata(Etitle=etitle, Ebody=ebody, Emailto=f'{emailin}', Emailcc=f'{emailcc}', Sendfiles=sf,
                   Sendasfiles=saf, Jolist=jl, Emailtype=itype, Mid=tone,
                   Customer=shipper, Container=ncl, Date1=today, Datelist=None, From=username, Box='SENT')
    db.session.add(input)
    db.session.commit()


def isoAR():
    username = session['username'].capitalize()
    #define User variables
    # Qote being worked
    uquot = f'{username}_quot'
    uiter = f'{username}_iter'
    umid= f'{username}_mid'
    utext= f'{username}_text'
    uhtml= f'{username}_html'
    quot=0
    tbox = [0]*28
    bidthis = [0]*5
    expdata=[]
    costdata=[]
    multibid=['off', 1, 0, 0]
    locs = []
    ebodytxt=''
    qdat = None
    mid = ''
    htmltext = None
    plaintext = None
    boxes = []
    tboxes = [0]*30
    analysis = request.values.get('analysis')
    emailgo = request.values.get('sendemail')
    exitnow = request.values.get('exitAR')
    active_task = request.values.get('active_task')
    redirect = request.values.get('exitAR2')
    update = request.values.get('updateall')
    update_e = request.values.get('updateemail')
    invoname = request.values.get('invoname')
    packname = request.values.get('packname')


    if request.method == 'POST':
        try:
            iter = int(os.environ[uiter])
        except:
            iter = 1
        try:
            oldmid = os.environ[umid]
        except:
            oldmid = 'Not Defined'
        try:
            plaintext = os.environ[utext]
        except:
            plaintext = ''
        try:
            htmltext = os.environ[uhtml]
        except:
            htmltext = ''
        #print(f'This is a POST with iter {iter} and last mid {oldmid}')



        #print(f'Values are {exitnow} {analysis} {emailgo} {exitnow}')

        if exitnow is not None:
            #print('Exiting quotes')
            return 'exitnow', None, None, None, None, None, None, None, None, None, None, None, None, None, None

        elif redirect is not None:
            this_shipper = request.values.get('this_shipper')
            task = 'artable review'


        elif analysis is not None or active_task == 'analysis':
            if analysis is not None:
                this_shipper = request.values.get('optradio')
                tboxes = ['on'] * 7 + ['off'] * 23
            else:
                this_shipper = request.values.get('this_shipper')
                tboxes = read_tboxes()

            #print(f'The shipper of interest is {this_shipper}')
            task = 'analysis'



        # Not exiting after a Post
        else:
            this_shipper = request.values.get('optradio')
            #print(f'The shipper of interest is {this_shipper}')
            task = 'artable review'

    #If not a POST
    else:
        iter = 1
        os.environ['MID'] = 'None Selected'
        #print('This is NOT a Post')
        ebodytxt = ''
        #print('Entering Quotes1',flush=True)
        username = session['username'].capitalize()
        this_shipper = None
        task = 'artable review'

    lb360 = today - datetime.timedelta(360)
    arorders = Orders.query.filter( ((Orders.Istat>1) & ((Orders.Istat<5) | (Orders.Istat==7))) & (Orders.Date3>lb360)).order_by(Orders.Date3).all()
    arlist = get_sorted_cust(arorders)
    arbycust = get_open_sort_totals(arlist)
    if this_shipper is None:
        this_shipper = arlist[0]
    odata, arsent = get_open_for_cust(this_shipper)

    boxes = [0] * len(odata)

    if task == 'analysis':
        dat30 = today - datetime.timedelta(30)
        if tboxes[20] == 'on':
            for ix, odat in enumerate(odata):
                if odat.Date3 < dat30: boxes[ix] = 'on'
        else:
            for ix, odat in enumerate(odata):
                boxes[ix] = request.values.get(f'box{ix}')

    #for ar in arbycust: print(f'{ar[0]} {ar[1]} {ar[2]} {ar[3]} {ar[4]} {ar[5]} {ar[6]}')

    if htmltext is None:
        showtext = plaintext
    else:
        showtext = htmltext

    containerlist = []
    jolist = []
    for ix, odat in enumerate(odata):
        if boxes[ix] == 'on':
            containerlist.append(odat.Container)
            jolist.append(odat.Jo)
    ar_emails_cust = get_ardata(containerlist, tboxes, this_shipper)
    nemail = len(ar_emails_cust)
    rview = rselect(nemail)

    pdat = People.query.filter(People.Company == this_shipper).first()
    emailsend = get_email_customer(pdat, ar_emails_cust)
    #print(f'Emailsend = {emailsend}')

    if emailgo is not None:
        #Need to get some emaildata from the website in case and changes made....
        emaildata = final_update_email(this_shipper, odata, tboxes, boxes, emailsend, update_e)
        #print(f'The final emaildata for send is {emaildata}')
        err = html_mimemail(emaildata)
        #print(f'err[0] is {err[0]}')
        if 'fail' not in err[0].lower():
            ardata_email_update(emaildata, this_shipper, jolist, containerlist)
    elif update_e is not None:
        #Get the emaildata setup from the website in case changes made....
        emaildata = final_update_email(this_shipper, odata, tboxes, boxes, emailsend, update_e)
        #print(f'The final emaildata for send is {emaildata}')
    else:
        emaildata = update_email(this_shipper, odata, tboxes, boxes, emailsend, update_e)





    #Save all the session variables that may have been updated...
    iter = iter + 1
    os.environ[uiter] = str(iter)
    if plaintext is None: plaintext = ''
    if htmltext is None: htmltext = ''
    os.environ[utext] = plaintext
    os.environ[uhtml] = htmltext
    os.environ[umid] = mid
    #print(f'Exiting with iter = {iter} and mid: {mid} for umid: {umid} and osenv for uiter: {os.environ[uiter]}')
    #print(f'email create tboxes={tboxes}')
    #print(f'data selection boxes={boxes}')
    #print(f'invoname:{invoname}, packname"{packname}')
    #print(f'emaildata is {emaildata}')
    #print(f'rview is {rview}')
    return 'keepgoing', arbycust, arsent, this_shipper, odata, task, emaildata, boxes, tboxes, invoname, packname, pdat, emailsend, ar_emails_cust, rview