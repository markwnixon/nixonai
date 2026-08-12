from flask import abort, render_template, redirect, url_for, jsonify, request, send_file, flash, session, Response
from flask import Blueprint
import json

from webapp.extensions import db, bcrypt
from webapp.models import People, Gledger, Accounts, Orders, Invoices, Deposits, PaymentsRec, Bills, Vehicles, Autos, DepreciationAsset, PlaidAccount, PlaidItem, PlaidTransaction, PortClosed, Drivers, DriverAssign, Driverlog, DispatchDriverMessage, DispatchDriverAssignment, Interchange, Reconciliations, OverSeas, users
#from webapp.forms import TruckingFormNew
from webapp.class8_tasks import Table_maker, get_dispatch
from webapp.revenues import get_revenues
from webapp.business_reports import ensure_interchange_port_trip_column, rebaseline_interchange_port_trips_for_year
from flask_login import login_required, current_user
from sqlalchemy import func, inspect, or_, text
from webapp.financial_mfa import FINANCIAL_GENRES, financial_mfa_redirect, financial_mfa_required
from webapp.plaid_integration import (
    create_bill_from_plaid_transaction,
    create_transfer_from_plaid_transaction,
    create_link_token,
    exchange_public_token,
    ignore_plaid_transaction,
    match_plaid_transaction_to_bill,
    plaid_accounts_for_current_scac,
    plaid_bill_match_options,
    plaid_dashboard_data,
    plaid_expense_account_options,
    plaid_processed_transactions,
    plaid_review_transactions,
    plaid_transfer_account_options,
    plaid_vendor_rule_lookup,
    plaid_ready,
    sync_item_transactions,
    upsert_item_from_exchange,
    update_item_status,
)
from webapp.bot_skills import bot_skill_rows, set_bot_skill_enabled, SKILL_DEFINITIONS
from webapp.bot_usage import usage_dashboard

from decimal import Decimal

from datetime import date, timedelta
import datetime

from twilio.twiml.messaging_response import MessagingResponse
from webapp.messager import msg_analysis
import requests
import mimetypes
from urllib.parse import urlparse
from webapp.viewfuncs import nonone, monvals, getmonths, newjo
from webapp.class8_tasks_gledger import post_balanced_journal
from webapp.class8_utils_email import email_template, info_mimemail, check_person, add_person
from webapp.dispatch_calendar import (
    calendar_events,
    calendar_notes,
    capacity_summary,
    delete_calendar_note,
    ensure_dispatch_calendar_tables,
    filter_options,
    move_event,
    parse_date as dispatch_parse_date,
    save_calendar_note,
    update_event,
)
from webapp.dispatch_kanban import (
    activate_pin_pairing,
    delete_pin_pairing,
    ensure_dispatch_kanban_tables,
    create_kanban_pin,
    kanban_jobs,
    kanban_options,
    log_review as kanban_log_review,
    make_pin_candidates,
    move_job as kanban_move_job,
    review_logs_for_order as kanban_review_logs,
    update_job as kanban_update_job,
    upload_proof as kanban_upload_proof,
)
from webapp.collection_kanban import (
    collection_email_logs,
    collection_jobs,
    collection_options,
    log_collection_call,
    send_collection_email,
    update_collection_job,
    upload_collection_rate_con,
)
from webapp.compliance_check import (
    add_compliance_item_type,
    compliance_asset_summary,
    compliance_due_alert_payload,
    compliance_items,
    compliance_item_type_options,
    compliance_profile,
    create_next_compliance_item,
    delete_compliance_item,
    delete_compliance_item_file,
    delete_compliance_item_type,
    delete_vehicle_repair,
    display_date as compliance_display_date,
    due_compliance_items,
    ensure_compliance_check_tables,
    file_link as compliance_file_link,
    save_vehicle_repair,
    save_compliance_item_file,
    send_compliance_alert_digest,
    update_compliance_profile,
    upsert_compliance_item,
)
from webapp.general_information import (
    add_general_information_category,
    combined_general_information_items,
    delete_general_information_category,
    delete_general_information_item,
    delete_secure_payment_method,
    ensure_general_information_tables,
    general_company_profile,
    general_information_categories,
    reveal_secure_payment_method,
    update_general_company_profile,
    upsert_general_information_item,
    upsert_secure_payment_method_from_general_form,
)

from webapp.class8_utils import *

TASKS = {}

today = datetime.datetime.today()
year = str(today.year)
day = str(today.day)
month = str(today.month)
now = datetime.datetime.now()
today_str = today.strftime('%Y-%m-%d')

from webapp.CCC_system_setup import apikeys, companydata, statpath, addpath, scac, tpath
from webapp.class8_dicts import PublicSite_setup
cmpdata = companydata()

main = Blueprint('main',__name__)

@main.route('/FileUpload', methods=['GET', 'POST'])
def FileUpload():
    err=[]
    uptype = request.values['uptype']
    print(uptype)
    oder = request.values['oid']
    oder = nonone(oder)
    print(oder)
    user = request.values['user']
    print(user)
    odat = Orders.query.get(oder)
    jo = odat.Jo
    pcache = odat.Pcache
    scache = odat.Scache
    fileob = request.files["file2upload"]
    name, ext = os.path.splitext(fileob.filename)
    ext = ext.lower()
    if uptype == 'proof':
        filename1 = f'Proof_{jo}_c{str(pcache)}{ext}'
        filename2 = f'Proof_{jo}_c{str(pcache)}.pdf'
        output1 = addpath(tpath('poof', filename1))
        output2 = addpath(tpath('poof', filename2))
        odat.Pcache = pcache + 1
        db.session.commit()
    elif uptype == 'source':
        filename1 = f'Source_{jo}_c{str(scache)}{ext}'
        filename2 = f'Source_{jo}_c{str(scache)}.pdf'
        output1 = addpath(tpath('oder', filename1))
        output2 = addpath(tpath('oder', filename2))
        odat.Scache = scache + 1
        db.session.commit()
    if ext != '.pdf':
        try:
            fileob.save(output1)
            with open(output2, "wb") as f:
                f.write(img2pdf.convert(output1))
            os.remove(output1)
        except:
            filename2 = filename1
    else:
        fileob.save(output2)
    if uptype == 'proof':
        odat.Proof = filename2
    elif uptype == 'source':
        odat.Original = filename2
    db.session.commit()

    print(f'File {fileob.filename} uploaded as {filename2}')

    return "successful_upload"


@main.route('/dispatch/planning-calendar', methods=['GET'])
@login_required
def DispatchPlanningCalendar():
    ensure_dispatch_calendar_tables()
    return render_template(
        'dispatch_planning_calendar.html',
        cmpdata=cmpdata,
        scac=scac,
        options=filter_options(),
    )


@main.route('/dispatch/port-closures', methods=['GET', 'POST'])
@login_required
def DispatchPortClosures():
    selected_year = request.values.get('year') or str(date.today().year)
    try:
        selected_year_int = int(selected_year)
    except:
        selected_year_int = date.today().year
    err = []
    if request.method == 'POST':
        action = request.values.get('action')
        closure_id = request.values.get('closure_id')
        closure_date = dispatch_parse_date(request.values.get('closure_date'))
        reason = (request.values.get('reason') or '').strip()
        if action in ['add', 'update'] and closure_date is None:
            err.append('Closure date is required.')
        elif action == 'add':
            existing = PortClosed.query.filter(PortClosed.Date == closure_date).first()
            if existing is None:
                db.session.add(PortClosed(Date=closure_date, Reason=reason or 'Port closed'))
            else:
                existing.Reason = reason or existing.Reason or 'Port closed'
            db.session.commit()
            return redirect(url_for('main.DispatchPortClosures', year=closure_date.year))
        elif action == 'update':
            closure = PortClosed.query.get(closure_id)
            if closure is None:
                err.append('Port closure row was not found.')
            else:
                closure.Date = closure_date
                closure.Reason = reason or 'Port closed'
                db.session.commit()
                return redirect(url_for('main.DispatchPortClosures', year=closure_date.year))
        elif action == 'delete':
            closure = PortClosed.query.get(closure_id)
            if closure is not None:
                row_year = closure.Date.year if closure.Date else selected_year_int
                db.session.delete(closure)
                db.session.commit()
                return redirect(url_for('main.DispatchPortClosures', year=row_year))
        else:
            err.append('Choose a valid port closure action.')

    year_start = date(selected_year_int, 1, 1)
    year_end = date(selected_year_int, 12, 31)
    closures = PortClosed.query.filter(
        (PortClosed.Date >= year_start) &
        (PortClosed.Date <= year_end)
    ).order_by(PortClosed.Date.asc(), PortClosed.id.asc()).all()
    years = list(range(selected_year_int - 2, selected_year_int + 4))
    return render_template(
        'dispatch_port_closures.html',
        cmpdata=cmpdata,
        scac=scac,
        closures=closures,
        selected_year=selected_year_int,
        years=years,
        err=err,
    )


def ensure_admin_payroll_calendar_table():
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS admin_payroll_calendar (
            id INT AUTO_INCREMENT PRIMARY KEY,
            RunDate DATE,
            PayDate DATE NOT NULL,
            Driver VARCHAR(100) NOT NULL,
            GrossCents INT NOT NULL DEFAULT 0,
            NetCents INT NOT NULL DEFAULT 0,
            Notes TEXT,
            CreatedBy VARCHAR(45),
            CreatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UpdatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_admin_payroll_calendar_date (PayDate),
            INDEX idx_admin_payroll_calendar_driver (Driver)
        )
    """))
    db.session.commit()
    columns = [column['name'] for column in inspect(db.engine).get_columns('admin_payroll_calendar')]
    if 'RunDate' not in columns:
        db.session.execute(text('ALTER TABLE admin_payroll_calendar ADD COLUMN RunDate DATE NULL'))
        db.session.commit()
    rows_missing_run_date = db.session.execute(text("""
        SELECT id, PayDate
        FROM admin_payroll_calendar
        WHERE RunDate IS NULL AND PayDate IS NOT NULL
    """)).mappings().all()
    for row in rows_missing_run_date:
        pay_date = row.get('PayDate')
        db.session.execute(text("""
            UPDATE admin_payroll_calendar
            SET RunDate = :run_date
            WHERE id = :row_id
        """), {
            'run_date': pay_date - timedelta(days=2),
            'row_id': row.get('id'),
        })
    if rows_missing_run_date:
        db.session.commit()


def parse_money_cents(value):
    clean = str(value or '').replace('$', '').replace(',', '').strip()
    if not clean:
        return 0
    return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))


def format_cents(cents):
    cents = int(cents or 0)
    return f'${Decimal(cents) / Decimal("100"):,.2f}'


def nth_weekday(year_value, month_value, weekday_value, occurrence):
    first_day = date(year_value, month_value, 1)
    offset = (weekday_value - first_day.weekday()) % 7
    return first_day + timedelta(days=offset + (occurrence - 1) * 7)


def last_weekday(year_value, month_value, weekday_value):
    if month_value == 12:
        last_day = date(year_value + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year_value, month_value + 1, 1) - timedelta(days=1)
    offset = (last_day.weekday() - weekday_value) % 7
    return last_day - timedelta(days=offset)


def observed_holiday(actual_date):
    if actual_date.weekday() == 5:
        return actual_date - timedelta(days=1)
    if actual_date.weekday() == 6:
        return actual_date + timedelta(days=1)
    return actual_date


def federal_holidays(year_value):
    fixed_holidays = [
        ("New Year's Day", date(year_value, 1, 1)),
        ('Juneteenth', date(year_value, 6, 19)),
        ('Independence Day', date(year_value, 7, 4)),
        ('Veterans Day', date(year_value, 11, 11)),
        ('Christmas Day', date(year_value, 12, 25)),
    ]
    holiday_rows = [
        ('Martin Luther King Jr. Day', nth_weekday(year_value, 1, 0, 3)),
        ("Washington's Birthday", nth_weekday(year_value, 2, 0, 3)),
        ('Memorial Day', last_weekday(year_value, 5, 0)),
        ('Labor Day', nth_weekday(year_value, 9, 0, 1)),
        ('Columbus Day', nth_weekday(year_value, 10, 0, 2)),
        ('Thanksgiving Day', nth_weekday(year_value, 11, 3, 4)),
    ]
    for name, actual in fixed_holidays:
        holiday_rows.append((name, observed_holiday(actual)))
    return sorted({
        holiday_date: holiday_name
        for holiday_name, holiday_date in holiday_rows
        if holiday_date.year == year_value
    }.items())


def holiday_in_pay_week(pay_date, holiday_dates):
    if not pay_date:
        return None
    week_start = pay_date - timedelta(days=pay_date.weekday())
    week_end = week_start + timedelta(days=6)
    for holiday_date, holiday_name in holiday_dates:
        if week_start <= holiday_date <= week_end:
            return {'date': holiday_date, 'name': holiday_name}
    return None


def recommended_run_date(pay_date, has_holiday):
    if not pay_date:
        return None
    week_start = pay_date - timedelta(days=pay_date.weekday())
    return week_start + timedelta(days=1 if has_holiday else 2)


@main.route('/AdminCalendar', methods=['GET', 'POST'])
@login_required
def AdminCalendar():
    ensure_admin_payroll_calendar_table()
    today_local = date.today()
    selected_year = request.values.get('year') or str(today_local.year)
    selected_month = request.values.get('month') or str(today_local.month)
    err = []
    try:
        selected_year_int = int(selected_year)
    except:
        selected_year_int = today_local.year
    try:
        selected_month_int = int(selected_month)
        if selected_month_int < 1 or selected_month_int > 12:
            selected_month_int = today_local.month
    except:
        selected_month_int = today_local.month

    if request.method == 'POST':
        action = request.values.get('action')
        row_id = request.values.get('row_id')
        run_date = dispatch_parse_date(request.values.get('run_date'))
        pay_date = dispatch_parse_date(request.values.get('pay_date'))
        driver = (request.values.get('driver') or '').strip()
        gross_cents = 0
        net_cents = 0
        notes = (request.values.get('notes') or '').strip()
        if action in ['add', 'update']:
            try:
                gross_cents = parse_money_cents(request.values.get('gross_amount'))
                net_cents = parse_money_cents(request.values.get('net_amount'))
            except:
                err.append('Payroll amounts must be valid numbers.')
            if run_date is None:
                err.append('Payroll run date is required.')
            if pay_date is None:
                err.append('Driver pay date is required.')
            if not driver:
                err.append('Driver is required.')
            if gross_cents < 0 or net_cents < 0:
                err.append('Payroll amounts cannot be negative.')
            if run_date and pay_date and run_date > pay_date:
                err.append('Payroll run date cannot be after the driver pay date.')

        if not err and action == 'add':
            db.session.execute(text("""
                INSERT INTO admin_payroll_calendar
                    (RunDate, PayDate, Driver, GrossCents, NetCents, Notes, CreatedBy, CreatedAt, UpdatedAt)
                VALUES
                    (:run_date, :pay_date, :driver, :gross_cents, :net_cents, :notes, :created_by, :created_at, :updated_at)
            """), {
                'run_date': run_date,
                'pay_date': pay_date,
                'driver': driver,
                'gross_cents': gross_cents,
                'net_cents': net_cents,
                'notes': notes,
                'created_by': getattr(current_user, 'username', None),
                'created_at': datetime.datetime.utcnow(),
                'updated_at': datetime.datetime.utcnow(),
            })
            db.session.commit()
            return redirect(url_for('main.AdminCalendar', year=pay_date.year, month=pay_date.month))
        if not err and action == 'update':
            existing = db.session.execute(
                text('SELECT id FROM admin_payroll_calendar WHERE id = :row_id'),
                {'row_id': row_id},
            ).first()
            if existing is None:
                err.append('Payroll calendar row was not found.')
            else:
                db.session.execute(text("""
                    UPDATE admin_payroll_calendar
                    SET RunDate = :run_date,
                        PayDate = :pay_date,
                        Driver = :driver,
                        GrossCents = :gross_cents,
                        NetCents = :net_cents,
                        Notes = :notes,
                        UpdatedAt = :updated_at
                    WHERE id = :row_id
                """), {
                    'row_id': row_id,
                    'run_date': run_date,
                    'pay_date': pay_date,
                    'driver': driver,
                    'gross_cents': gross_cents,
                    'net_cents': net_cents,
                    'notes': notes,
                    'updated_at': datetime.datetime.utcnow(),
                })
                db.session.commit()
                return redirect(url_for('main.AdminCalendar', year=pay_date.year, month=pay_date.month))
        if action == 'delete':
            row = db.session.execute(
                text('SELECT PayDate FROM admin_payroll_calendar WHERE id = :row_id'),
                {'row_id': row_id},
            ).mappings().first()
            if row is not None:
                row_date = row.get('PayDate') or date(selected_year_int, selected_month_int, 1)
                db.session.execute(
                    text('DELETE FROM admin_payroll_calendar WHERE id = :row_id'),
                    {'row_id': row_id},
                )
                db.session.commit()
                return redirect(url_for('main.AdminCalendar', year=row_date.year, month=row_date.month))
        elif action not in ['add', 'update']:
            err.append('Choose a valid payroll calendar action.')

    month_start = date(selected_year_int, selected_month_int, 1)
    if selected_month_int == 12:
        month_end = date(selected_year_int + 1, 1, 1)
    else:
        month_end = date(selected_year_int, selected_month_int + 1, 1)
    year_start = date(selected_year_int, 1, 1)
    year_end = date(selected_year_int + 1, 1, 1)

    rows = db.session.execute(text("""
        SELECT id, RunDate, PayDate, Driver, GrossCents, NetCents, Notes
        FROM admin_payroll_calendar
        WHERE (PayDate >= :month_start AND PayDate < :month_end)
           OR (RunDate >= :month_start AND RunDate < :month_end)
        ORDER BY PayDate, RunDate, Driver, id
    """), {
        'month_start': month_start,
        'month_end': month_end,
    }).mappings().all()
    year_rows = db.session.execute(text("""
        SELECT id, RunDate, PayDate, Driver, GrossCents, NetCents, Notes
        FROM admin_payroll_calendar
        WHERE (PayDate >= :year_start AND PayDate < :year_end)
           OR (RunDate >= :year_start AND RunDate < :year_end)
        ORDER BY PayDate, RunDate, Driver, id
    """), {
        'year_start': year_start,
        'year_end': year_end,
    }).mappings().all()
    driver_totals = {}
    for row in year_rows:
        bucket = driver_totals.setdefault(row.get('Driver') or 'Unassigned', {'gross': 0, 'net': 0, 'count': 0})
        bucket['gross'] += int(row.get('GrossCents') or 0)
        bucket['net'] += int(row.get('NetCents') or 0)
        bucket['count'] += 1

    holidays = federal_holidays(selected_year_int)
    holiday_events = [{
        'id': f'holiday-{holiday_date.strftime("%Y%m%d")}',
        'title': holiday_name,
        'start': holiday_date.strftime('%Y-%m-%d'),
        'allDay': True,
        'classNames': ['admin-calendar-holiday'],
        'extendedProps': {
            'event_type': 'holiday',
            'holiday_name': holiday_name,
        },
    } for holiday_date, holiday_name in holidays]
    events = []
    table_rows = []
    for row in year_rows:
        warning = holiday_in_pay_week(row.get('PayDate'), holidays)
        suggested = recommended_run_date(row.get('PayDate'), bool(warning))
        base_props = {
            'driver': row.get('Driver') or '',
            'gross': format_cents(row.get('GrossCents')),
            'net': format_cents(row.get('NetCents')),
            'notes': row.get('Notes') or '',
            'holiday_warning': bool(warning),
            'holiday_name': warning['name'] if warning else '',
            'holiday_date': warning['date'].strftime('%Y-%m-%d') if warning else '',
            'recommended_run_date': suggested.strftime('%Y-%m-%d') if suggested else '',
        }
        if row.get('RunDate'):
            events.append({
                'id': f"run-{row.get('id')}",
                'title': f"Run payroll: {row.get('Driver')}",
                'start': row.get('RunDate').strftime('%Y-%m-%d'),
                'allDay': True,
                'classNames': ['admin-calendar-payroll-run', 'admin-calendar-payroll-warning'] if warning else ['admin-calendar-payroll-run'],
                'extendedProps': dict(base_props, event_type='run', row_id=row.get('id')),
            })
        if row.get('PayDate'):
            events.append({
                'id': f"pay-{row.get('id')}",
                'title': f"Pay driver: {row.get('Driver')}",
                'start': row.get('PayDate').strftime('%Y-%m-%d'),
                'allDay': True,
                'classNames': ['admin-calendar-payroll-pay', 'admin-calendar-payroll-warning'] if warning else ['admin-calendar-payroll-pay'],
                'extendedProps': dict(base_props, event_type='pay', row_id=row.get('id')),
            })
    events.extend(holiday_events)
    for row in rows:
        warning = holiday_in_pay_week(row.get('PayDate'), holidays)
        suggested = recommended_run_date(row.get('PayDate'), bool(warning))
        row_dict = dict(row)
        row_dict['HolidayWarning'] = warning
        row_dict['RecommendedRunDate'] = suggested
        table_rows.append(row_dict)
    drivers = Drivers.query.filter((Drivers.Active == 1) | (Drivers.Active == None)).order_by(Drivers.Name).all()
    driver_names = [driver.Name for driver in drivers]
    years = list(range(selected_year_int - 2, selected_year_int + 4))
    months = [(index, date(2000, index, 1).strftime('%B')) for index in range(1, 13)]
    return render_template(
        'admin_calendar.html',
        cmpdata=cmpdata,
        scac=scac,
        err=err,
        rows=table_rows,
        events=events,
        holidays=holidays,
        drivers=drivers,
        driver_names=driver_names,
        selected_year=selected_year_int,
        selected_month=selected_month_int,
        years=years,
        months=months,
        driver_totals=driver_totals,
        format_cents=format_cents,
        initial_date=month_start.strftime('%Y-%m-%d'),
        initial_run_date=recommended_run_date(month_start, False).strftime('%Y-%m-%d'),
    )


@main.route('/ComplianceCheck', methods=['GET', 'POST'])
@login_required
def ComplianceCheck():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_compliance_check_tables()
    company_code = cmpdata[10]
    err = []
    msg = []
    profile = compliance_profile(company_code, cmpdata[2], cmpdata[8])

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update_profile':
            update_compliance_profile(company_code, request.form)
            msg.append('Company compliance profile updated.')
        elif action == 'save_item':
            upsert_compliance_item(company_code, request.form)
            msg.append('Compliance item saved.')
        elif action == 'delete_item':
            delete_compliance_item(company_code, request.form.get('item_id'))
            msg.append('Compliance item deleted.')
        elif action == 'create_next_item':
            next_err = create_next_compliance_item(company_code, request.form.get('item_id'))
            if next_err:
                err.append(next_err)
            else:
                msg.append('Next compliance item created.')
        elif action == 'add_item_type':
            type_err = add_compliance_item_type(company_code, request.form.get('type_label'))
            if type_err:
                err.append(type_err)
            else:
                msg.append('Compliance item type added.')
        elif action == 'delete_item_type':
            delete_compliance_item_type(company_code, request.form.get('type_value'))
            msg.append('Compliance item type removed from selection list.')
        elif action == 'upload_item_file':
            upload_err = save_compliance_item_file(
                company_code,
                request.form.get('item_id'),
                request.files.get('item_file'),
                request.form.get('file_label'),
            )
            if upload_err:
                err.append(upload_err)
            else:
                msg.append('Compliance document uploaded.')
        elif action == 'delete_item_file':
            delete_compliance_item_file(company_code, request.form.get('file_id'))
            msg.append('Compliance document deleted.')
        elif action == 'send_alerts':
            due_rows = due_compliance_items(company_code, profile.get('AlertLeadDays') or 30)
            err.extend(send_compliance_alert_digest(company_code, profile, due_rows))

    profile = compliance_profile(company_code, cmpdata[2], cmpdata[8])
    items = compliance_items(company_code)
    due_items = due_compliance_items(company_code, profile.get('AlertLeadDays') or 30)
    drivers, vehicles = compliance_asset_summary(company_code)
    return render_template(
        'compliance_check.html',
        cmpdata=cmpdata,
        scac=scac,
        profile=profile,
        item_types=compliance_item_type_options(company_code),
        items=items,
        due_items=due_items,
        drivers=drivers,
        vehicles=vehicles,
        file_link=compliance_file_link,
        display_date=compliance_display_date,
        err=err,
        msg=msg,
    )


def _driver_log_filter_date(value, default_value):
    if not value:
        return default_value
    try:
        return datetime.date.fromisoformat(value)
    except (TypeError, ValueError):
        return default_value


def _driver_log_duration(clockin, clockout):
    if not clockin or not clockout:
        return ''
    delta = clockout - clockin
    total_minutes = max(0, int(delta.total_seconds() // 60))
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f'{hours}:{minutes:02d}'


def _driver_log_aliases(driver_name):
    aliases = [driver_name]
    user_rows = users.query.filter(
        (users.username == driver_name) | (users.name == driver_name)
    ).all()
    for user_row in user_rows:
        for value in [user_row.username, user_row.name]:
            if value and value not in aliases:
                aliases.append(value)
    return aliases


def ensure_dispatch_driver_message_table():
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS dispatch_driver_messages (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Co VARCHAR(12) NULL,
            OrderId INT NULL,
            Jo VARCHAR(25) NULL,
            Container VARCHAR(50) NULL,
            Driver VARCHAR(100) NULL,
            SenderName VARCHAR(100) NULL,
            SenderType VARCHAR(30) NULL,
            MessageText TEXT NULL,
            RouteJson TEXT NULL,
            CreatedAt DATETIME NOT NULL,
            ReadAt DATETIME NULL,
            Active INT NOT NULL DEFAULT 1,
            INDEX idx_dispatch_driver_msg_order (OrderId),
            INDEX idx_dispatch_driver_msg_jo (Jo),
            INDEX idx_dispatch_driver_msg_driver (Driver),
            INDEX idx_dispatch_driver_msg_created (CreatedAt)
        )
    """))
    inspector = inspect(db.engine)
    existing_columns = [column['name'] for column in inspector.get_columns('dispatch_driver_messages')]
    if 'RouteJson' not in existing_columns:
        db.session.execute(text("ALTER TABLE dispatch_driver_messages ADD COLUMN RouteJson TEXT NULL"))
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS dispatch_driver_assignments (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Co VARCHAR(12) NULL,
            Driver VARCHAR(100) NULL,
            AssignmentType VARCHAR(45) NULL,
            PrimaryOrderId INT NULL,
            SecondaryOrderId INT NULL,
            PrimaryJo VARCHAR(25) NULL,
            SecondaryJo VARCHAR(25) NULL,
            PrimaryContainer VARCHAR(50) NULL,
            SecondaryContainer VARCHAR(50) NULL,
            RouteOrderId INT NULL,
            DestinationName VARCHAR(100) NULL,
            DestinationAddress VARCHAR(500) NULL,
            MessageText TEXT NULL,
            RouteJson TEXT NULL,
            Status VARCHAR(45) NULL,
            CreatedBy VARCHAR(100) NULL,
            CreatedAt DATETIME NOT NULL,
            Active INT NOT NULL DEFAULT 1,
            INDEX idx_dispatch_driver_assign_driver (Driver),
            INDEX idx_dispatch_driver_assign_primary (PrimaryOrderId),
            INDEX idx_dispatch_driver_assign_secondary (SecondaryOrderId),
            INDEX idx_dispatch_driver_assign_created (CreatedAt)
        )
    """))
    existing_assignment_columns = [column['name'] for column in inspector.get_columns('dispatch_driver_assignments')]
    if 'RouteJson' not in existing_assignment_columns:
        db.session.execute(text("ALTER TABLE dispatch_driver_assignments ADD COLUMN RouteJson TEXT NULL"))
    db.session.commit()


def _dispatch_panel_clean_text(value):
    return (value or '').replace('\r', ' ').replace('\n', ' ').strip()


def _dispatch_panel_order_payload(order):
    return {
        'id': order.id,
        'jo': order.Jo,
        'container': order.Container,
        'customer': order.Shipper,
        'destination_name': order.Company2,
        'destination_address': _dispatch_panel_clean_text(order.Dropblock2),
        'delivery_date': order.Date3.strftime('%Y-%m-%d') if order.Date3 else None,
        'delivery_time': order.Time3,
        'truck': order.Truck,
    }


def _dispatch_panel_message_payload(row, dispatcher_name):
    sender_type = (row.SenderType or '').lower()
    direction = 'outbound' if sender_type == 'dispatch' else 'inbound'
    if row.SenderName == dispatcher_name:
        direction = 'outbound'
    route_payload = None
    if getattr(row, 'RouteJson', None):
        try:
            route_payload = json.loads(row.RouteJson)
        except (TypeError, ValueError):
            route_payload = None
    return {
        'id': row.id,
        'direction': direction,
        'message_flow': 'dispatch_to_driver' if sender_type == 'dispatch' else 'driver_to_dispatch',
        'sender_name': row.SenderName,
        'sender_type': row.SenderType,
        'message': row.MessageText,
        'route': route_payload,
        'created_at': row.CreatedAt.strftime('%Y-%m-%d %H:%M') if row.CreatedAt else '',
    }


def _dispatch_panel_assignment_pending_payload(row):
    route_payload = None
    if getattr(row, 'RouteJson', None):
        try:
            route_payload = json.loads(row.RouteJson)
        except (TypeError, ValueError):
            route_payload = None
    return {
        'id': row.id,
        'driver': row.Driver,
        'assignment_type': row.AssignmentType,
        'primary_order_id': row.PrimaryOrderId,
        'secondary_order_id': row.SecondaryOrderId,
        'primary_jo': row.PrimaryJo,
        'secondary_jo': row.SecondaryJo,
        'primary_container': row.PrimaryContainer,
        'secondary_container': row.SecondaryContainer,
        'route_order_id': row.RouteOrderId,
        'destination_name': row.DestinationName,
        'destination_address': row.DestinationAddress,
        'message_text': row.MessageText,
        'route': route_payload,
        'status': row.Status,
        'created_by': row.CreatedBy,
        'created_at': row.CreatedAt.strftime('%Y-%m-%d %H:%M') if row.CreatedAt else '',
    }


def _driver_open_clock_log(driver_name):
    if not driver_name:
        return None
    return Driverlog.query.filter(
        Driverlog.Driver == driver_name,
        Driverlog.Clockin.isnot(None),
        Driverlog.Clockout.is_(None)
    ).order_by(Driverlog.Clockin.desc()).first()


def _pending_dispatch_assignments(driver_name):
    if not driver_name:
        return []
    rows = DispatchDriverAssignment.query.filter(
        DispatchDriverAssignment.Driver == driver_name,
        DispatchDriverAssignment.Active == 1,
        DispatchDriverAssignment.Status.in_(['pending', 'draft'])
    ).order_by(DispatchDriverAssignment.CreatedAt, DispatchDriverAssignment.id).all()
    return [_dispatch_panel_assignment_pending_payload(row) for row in rows]


def _save_pending_dispatch_assignment(driver_name, order, message_text, route_payload, assignment_type, sender_name):
    route_json = json.dumps(route_payload) if isinstance(route_payload, dict) else None
    route_data = route_payload if isinstance(route_payload, dict) else {}
    destination_name = order.Company2 if order else route_data.get('destination_name')
    destination_address = route_data.get('destination_address')
    assignment = DispatchDriverAssignment.query.filter(
        DispatchDriverAssignment.Driver == driver_name,
        DispatchDriverAssignment.Active == 1,
        DispatchDriverAssignment.Status == 'draft',
        DispatchDriverAssignment.MessageText == message_text
    ).order_by(DispatchDriverAssignment.CreatedAt.desc(), DispatchDriverAssignment.id.desc()).first()
    if assignment is not None:
        assignment.AssignmentType = assignment_type
        assignment.PrimaryOrderId = order.id if order else assignment.PrimaryOrderId
        assignment.PrimaryJo = order.Jo if order else assignment.PrimaryJo
        assignment.PrimaryContainer = order.Container if order else assignment.PrimaryContainer
        assignment.RouteOrderId = order.id if order else assignment.RouteOrderId
        assignment.DestinationName = destination_name
        assignment.DestinationAddress = destination_address
        assignment.RouteJson = route_json
        assignment.Status = 'pending'
        assignment.CreatedBy = sender_name
        db.session.commit()
        return assignment

    assignment = DispatchDriverAssignment(
        Co=scac,
        Driver=driver_name,
        AssignmentType=assignment_type,
        PrimaryOrderId=order.id if order else None,
        SecondaryOrderId=None,
        PrimaryJo=order.Jo if order else None,
        SecondaryJo=None,
        PrimaryContainer=order.Container if order else None,
        SecondaryContainer=None,
        RouteOrderId=order.id if order else None,
        DestinationName=destination_name,
        DestinationAddress=destination_address,
        MessageText=message_text,
        Status='pending',
        CreatedBy=sender_name,
        CreatedAt=datetime.datetime.now(),
        Active=1,
        RouteJson=route_json,
    )
    db.session.add(assignment)
    db.session.commit()
    return assignment


def _dispatch_panel_google_distance(origin, destination):
    api_key = ''
    try:
        api_key = apikeys.get('dkey') or ''
    except Exception:
        api_key = ''
    if not api_key or not origin or not destination:
        return None
    try:
        response = requests.get(
            'https://maps.googleapis.com/maps/api/distancematrix/json',
            params={
                'units': 'imperial',
                'origins': origin,
                'destinations': destination,
                'key': api_key,
            },
            timeout=6,
        )
        payload = response.json()
        element = payload['rows'][0]['elements'][0]
        if element.get('status') != 'OK':
            return None
        return {
            'distance_text': element.get('distance', {}).get('text') or '',
            'duration_text': element.get('duration', {}).get('text') or '',
        }
    except Exception:
        return None


def _dispatch_panel_route_payload(driver_name, order):
    if not driver_name or order is None:
        return None
    open_log = Driverlog.query.filter(
        Driverlog.Driver == driver_name,
        Driverlog.Clockin.isnot(None),
        Driverlog.Clockout.is_(None)
    ).order_by(Driverlog.Clockin.desc()).first()
    destination = _dispatch_panel_clean_text(order.Dropblock2 or order.Company2)
    origin = open_log.GPSin if open_log and open_log.GPSin else ''
    origin_label = open_log.Locationstart if open_log else ''
    if origin and destination:
        from urllib.parse import quote_plus
        maps_url = (
            'https://www.google.com/maps/dir/?api=1'
            f'&origin={quote_plus(origin)}'
            f'&destination={quote_plus(destination)}'
        )
    else:
        maps_url = None
    route = {
        'origin_location': origin_label,
        'origin_coordinates': origin,
        'origin_clockin': open_log.Clockin.strftime('%Y-%m-%d %H:%M') if open_log and open_log.Clockin else '',
        'destination_name': order.Company2,
        'destination_address': destination,
        'maps_url': maps_url,
        'distance_text': '',
        'duration_text': '',
        'eta': '',
        'status': 'ready' if maps_url else 'missing clock-in coordinates or destination',
    }
    route_data = _dispatch_panel_google_distance(origin, destination)
    if route_data:
        route.update(route_data)
        route['eta'] = route_data.get('duration_text') or ''
    elif maps_url:
        route['eta'] = 'Open route link for live ETA'
    return route


def _dispatch_panel_port_origin(order):
    return _dispatch_panel_clean_text(
        getattr(order, 'Dropblock1', None)
        or getattr(order, 'Company', None)
        or getattr(order, 'Pickup', None)
        or getattr(order, 'Location', None)
    )


def _dispatch_panel_route_destination(order):
    dropblock = _dispatch_panel_clean_text(getattr(order, 'Dropblock2', None))
    lines = [line.strip() for line in dropblock.replace('\r', '\n').split('\n') if line.strip()]
    if len(lines) >= 3:
        return _dispatch_panel_clean_text(f'{lines[1]} {lines[2]}')
    if len(lines) == 2:
        return _dispatch_panel_clean_text(f'{lines[0]} {lines[1]}')
    return _dispatch_panel_clean_text(getattr(order, 'Location3', None) or getattr(order, 'Dropblock2', None) or getattr(order, 'Company2', None))


def _dispatch_panel_route_from_port_payload(order):
    if order is None:
        return None
    origin = _dispatch_panel_port_origin(order)
    destination = _dispatch_panel_route_destination(order)
    if origin and destination:
        from urllib.parse import quote_plus
        maps_url = (
            'https://www.google.com/maps/dir/?api=1'
            f'&origin={quote_plus(origin)}'
            f'&destination={quote_plus(destination)}'
        )
    else:
        maps_url = None
    route = {
        'origin_location': origin,
        'origin_coordinates': '',
        'origin_clockin': '',
        'destination_name': order.Company2,
        'destination_address': destination,
        'maps_url': maps_url,
        'distance_text': '',
        'duration_text': '',
        'eta': '',
        'status': 'port gate to delivery' if maps_url else 'missing port gate or destination',
    }
    route_data = _dispatch_panel_google_distance(origin, destination)
    if route_data:
        route.update(route_data)
        route['eta'] = route_data.get('duration_text') or ''
    elif maps_url:
        route['eta'] = 'Open route link for live ETA'
    return route


def _dispatch_panel_assignment_payload(row):
    if row is None:
        return None
    return {
        'id': row.id,
        'driver': row.Driver,
        'assignment_type': row.AssignmentType,
        'primary_order_id': row.PrimaryOrderId,
        'secondary_order_id': row.SecondaryOrderId,
        'primary_jo': row.PrimaryJo,
        'secondary_jo': row.SecondaryJo,
        'primary_container': row.PrimaryContainer,
        'secondary_container': row.SecondaryContainer,
        'route_order_id': row.RouteOrderId,
        'destination_name': row.DestinationName,
        'destination_address': row.DestinationAddress,
        'message_text': row.MessageText,
        'status': row.Status,
        'created_by': row.CreatedBy,
        'created_at': row.CreatedAt.isoformat() if row.CreatedAt else None,
    }


def _is_import_order(order):
    return 'import' in ((order.HaulType or '').lower())


def _dispatch_delivery_lines(order):
    dropblock = _dispatch_panel_clean_text(getattr(order, 'Dropblock2', None))
    lines = [line.strip() for line in dropblock.replace('\r', '\n').split('\n') if line.strip()]
    company = _dispatch_panel_clean_text(getattr(order, 'Company2', None))
    city_state_zip = _dispatch_panel_clean_text(getattr(order, 'Location3', None))

    if not lines and company:
        lines.append(company)
    elif company and lines and company.lower() not in lines[0].lower():
        lines.insert(0, company)

    if city_state_zip and all(city_state_zip.lower() not in line.lower() for line in lines):
        lines.append(city_state_zip)

    while len(lines) < 3:
        lines.append('')
    return lines[:3]


def _dispatch_assignment_text(orders, assignment_type):
    port_in_order = None
    port_out_order = None
    for order in orders:
        hstat = order.Hstat if order.Hstat is not None else -1
        if hstat >= 1 and port_in_order is None:
            port_in_order = order
        if hstat < 1 and port_out_order is None:
            port_out_order = order

    if port_in_order or port_out_order:
        lines = []
        if port_in_order:
            if _is_import_order(port_in_order):
                lines.append(f'Please hook up to empty container {port_in_order.Container} for return to port.')
            else:
                lines.append(f'Please hook up to container {port_in_order.Container} for return to port.')
        if port_out_order:
            pull_label = 'import' if _is_import_order(port_out_order) else 'export'
            if assignment_type == 'pull_deliver':
                lines.append(f'Pull {pull_label} {port_out_order.Container} for delivery to:')
                lines.extend(_dispatch_delivery_lines(port_out_order))
            elif assignment_type == 'pull_store':
                lines.append(f'Pull {pull_label} {port_out_order.Container}.')
                lines.append('Then pull and store.')
            else:
                lines.append(f'Pull {pull_label} {port_out_order.Container}.')
        message = '\n'.join(line for line in lines if line is not None).strip()
        route_order = port_out_order if assignment_type == 'pull_deliver' and port_out_order else port_in_order or port_out_order
        return message, route_order

    lines = []
    for order in orders:
        try:
            dispatch_text = get_dispatch(order)
        except Exception:
            dispatch_text = ''
        if not dispatch_text:
            dispatch_text = f'Dispatch container {order.Container or ""} for job {order.Jo or ""}.'.strip()
        lines.append(dispatch_text)
    if assignment_type == 'pull_deliver':
        lines.insert(0, 'Pull and deliver:')
    elif assignment_type == 'pull_store':
        lines.insert(0, 'Pull and store:')
    route_order = orders[0] if orders else None
    return '\n'.join(lines), route_order


def _dispatch_panel_driver_names():
    active_driver_rows = Drivers.query.filter(
        (Drivers.Active == 1) | (Drivers.Active == None)
    ).all()
    active_driver_names = {
        (driver.Name or '').strip()
        for driver in active_driver_rows
        if (driver.Name or '').strip()
    }

    allowed_names = []
    user_rows = users.query.filter(
        users.authority.in_(['driver', 'admin', 'superuser'])
    ).order_by(users.name).all()
    for user_row in user_rows:
        name = (user_row.name or '').strip()
        authority = (user_row.authority or '').strip().lower()
        if not name:
            continue
        if authority == 'driver' or name in active_driver_names:
            if name not in allowed_names:
                allowed_names.append(name)
    return allowed_names


@main.route('/DriverDispatchPanel', methods=['GET', 'POST'])
@login_required
def DriverDispatchPanel():
    ensure_dispatch_driver_message_table()
    driver_names = _dispatch_panel_driver_names()
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        driver_name = data.get('driver') or ''
        message_text = data.get('message') or ''
        jo = data.get('jo') or ''
        order = Orders.query.filter(Orders.Jo == jo).order_by(Orders.id.desc()).first() if jo else None
        if not driver_name:
            return jsonify({'error': 'Driver is required.'}), 400
        if driver_name not in driver_names:
            return jsonify({'error': 'Driver is not available for dispatch communications.'}), 400
        if not message_text:
            return jsonify({'error': 'Message is required.'}), 400
        assignment_type = data.get('assignment_type') or 'general'
        route_payload = data.get('route')
        route_json = json.dumps(route_payload) if isinstance(route_payload, dict) else None
        sender_name = getattr(current_user, 'name', None) or getattr(current_user, 'username', None) or 'Dispatch'
        if assignment_type != 'general' and _driver_open_clock_log(driver_name) is None:
            assignment = _save_pending_dispatch_assignment(
                driver_name,
                order,
                message_text,
                route_payload,
                assignment_type,
                sender_name,
            )
            return jsonify({
                'message': 'Driver is not clocked in. Dispatch assignment queued.',
                'queued': True,
                'assignment': _dispatch_panel_assignment_pending_payload(assignment),
                'pending_assignments': _pending_dispatch_assignments(driver_name),
            }), 202

        row = DispatchDriverMessage(
            Co=scac,
            OrderId=order.id if order else None,
            Jo=order.Jo if order else jo,
            Container=order.Container if order else data.get('container'),
            Driver=driver_name,
            SenderName=sender_name,
            SenderType='dispatch',
            MessageText=message_text,
            CreatedAt=datetime.datetime.now(),
            ReadAt=None,
            Active=1,
            RouteJson=route_json,
        )
        db.session.add(row)
        db.session.commit()
        return jsonify({
            'message': 'Dispatch message saved.',
            'queued': False,
            'dispatch_message': _dispatch_panel_message_payload(row, sender_name),
        }), 201

    selected_driver = request.args.get('driver') or ''
    selected_jo = request.args.get('jo') or ''
    if selected_driver and selected_driver not in driver_names:
        selected_driver = ''
    assignments = []
    if selected_driver:
        assignment_rows = Orders.query.filter(
            Orders.Driver == selected_driver,
            (Orders.Hstat == None) | (Orders.Hstat < 2)
        ).order_by(Orders.Date3, Orders.Jo).limit(30).all()
        assignments = [_dispatch_panel_order_payload(order) for order in assignment_rows]

    selected_order = Orders.query.filter(Orders.Jo == selected_jo).order_by(Orders.id.desc()).first() if selected_jo else None
    msg_query = DispatchDriverMessage.query.filter(
        DispatchDriverMessage.Active == 1,
        DispatchDriverMessage.Driver == selected_driver,
    )
    if selected_order:
        msg_query = msg_query.filter(DispatchDriverMessage.OrderId == selected_order.id)
    elif selected_jo:
        msg_query = msg_query.filter(DispatchDriverMessage.Jo == selected_jo)
    messages = msg_query.order_by(DispatchDriverMessage.CreatedAt, DispatchDriverMessage.id).limit(100).all() if selected_driver else []
    dispatcher_name = getattr(current_user, 'name', None) or getattr(current_user, 'username', None) or 'Dispatch'
    return jsonify({
        'drivers': driver_names,
        'selected_driver': selected_driver,
        'assignments': assignments,
        'selected_order': _dispatch_panel_order_payload(selected_order) if selected_order else None,
        'route': _dispatch_panel_route_payload(selected_driver, selected_order),
        'pending_assignments': _pending_dispatch_assignments(selected_driver),
        'messages': [_dispatch_panel_message_payload(row, dispatcher_name) for row in messages],
    }), 200


@main.route('/DriverDispatchPanel/assignment-text', methods=['POST'])
@login_required
def DriverDispatchAssignmentText():
    data = request.get_json(silent=True) or {}
    order_ids = data.get('order_ids') or []
    assignment_type = data.get('assignment_type') or 'general'
    driver_name = data.get('driver') or ''
    driver_names = _dispatch_panel_driver_names()
    try:
        order_ids = [int(order_id) for order_id in order_ids]
    except (TypeError, ValueError):
        return jsonify({'error': 'Order ids must be integers.'}), 400

    if driver_name and driver_name not in driver_names:
        return jsonify({'error': 'Driver is not available for dispatch communications.'}), 400
    if not order_ids:
        return jsonify({'error': 'Select one or two order rows first.'}), 400
    if len(order_ids) > 2:
        return jsonify({'error': 'Select no more than two order rows for dispatch copy.'}), 400

    orders = [db.session.get(Orders, order_id) for order_id in order_ids]
    orders = [order for order in orders if order is not None]
    if not orders:
        return jsonify({'error': 'Selected orders were not found.'}), 404

    if len(orders) == 2:
        first, second = orders[0], orders[1]
        first_hstat = first.Hstat if first.Hstat is not None else -1
        second_hstat = second.Hstat if second.Hstat is not None else -1
        if first_hstat > second_hstat:
            orders = [first, second]
        else:
            orders = [second, first]

    message_text, route_order = _dispatch_assignment_text(orders, assignment_type)
    if not message_text:
        return jsonify({'error': 'Unable to generate dispatch assignment text for the selected orders.'}), 400

    assignment = None
    warning = ''
    if driver_name:
        primary_order = orders[0] if orders else None
        secondary_order = orders[1] if len(orders) > 1 else None
        dispatcher_name = getattr(current_user, 'name', None) or getattr(current_user, 'username', None) or 'Dispatch'
        assignment = DispatchDriverAssignment(
            Co=scac,
            Driver=driver_name,
            AssignmentType=assignment_type,
            PrimaryOrderId=primary_order.id if primary_order else None,
            SecondaryOrderId=secondary_order.id if secondary_order else None,
            PrimaryJo=primary_order.Jo if primary_order else None,
            SecondaryJo=secondary_order.Jo if secondary_order else None,
            PrimaryContainer=primary_order.Container if primary_order else None,
            SecondaryContainer=secondary_order.Container if secondary_order else None,
            RouteOrderId=route_order.id if route_order else None,
            DestinationName=route_order.Company2 if route_order else None,
            DestinationAddress=_dispatch_panel_clean_text(route_order.Dropblock2) if route_order else None,
            MessageText=message_text,
            Status='draft',
            CreatedBy=dispatcher_name,
            CreatedAt=datetime.datetime.now(),
            Active=1,
        )
        db.session.add(assignment)
        db.session.commit()
    else:
        warning = 'Select a driver to save this as a dispatch assignment.'

    return jsonify({
        'message_text': message_text,
        'assignment': _dispatch_panel_assignment_payload(assignment),
        'warning': warning,
        'selected_order': _dispatch_panel_order_payload(route_order) if route_order else None,
        'route': _dispatch_panel_route_from_port_payload(route_order) if assignment_type == 'pull_deliver' else _dispatch_panel_route_payload(driver_name, route_order),
    }), 200


@main.route('/DriverLogs', methods=['GET'])
@login_required
def DriverLogs():
    today = datetime.date.today()
    default_start = today - datetime.timedelta(days=14)
    default_end = today + datetime.timedelta(days=1)
    start_date = _driver_log_filter_date(request.args.get('start_date'), default_start)
    end_date = _driver_log_filter_date(request.args.get('end_date'), default_end)
    selected_driver = request.args.get('driver') or 'All Drivers'

    query = Driverlog.query.filter(
        Driverlog.Date >= start_date,
        Driverlog.Date <= end_date,
    )
    if selected_driver != 'All Drivers':
        query = query.filter(Driverlog.Driver.in_(_driver_log_aliases(selected_driver)))

    logs = query.order_by(
        Driverlog.Date.desc(),
        Driverlog.Driver,
        Driverlog.Shift,
        Driverlog.Clockin.desc(),
    ).all()

    driver_rows = Drivers.query.filter(
        (Drivers.Active == 1) | (Drivers.Active == None)
    ).order_by(Drivers.Name).all()
    driver_names = ['All Drivers']
    for driver in driver_rows:
        if driver.Name and driver.Name not in driver_names:
            driver_names.append(driver.Name)
    user_driver_rows = users.query.filter(users.authority == 'driver').order_by(users.name, users.username).all()
    for user_row in user_driver_rows:
        driver_display = user_row.name or user_row.username
        if driver_display and driver_display not in driver_names:
            driver_names.append(driver_display)
    for row in logs:
        if row.Driver and row.Driver not in driver_names:
            driver_names.append(row.Driver)

    return render_template(
        'driver_logs.html',
        cmpdata=cmpdata,
        scac=scac,
        logs=logs,
        driver_names=driver_names,
        selected_driver=selected_driver,
        start_date=start_date,
        end_date=end_date,
        duration=_driver_log_duration,
    )


@main.route('/AssetMaintenance', methods=['GET', 'POST'])
@login_required
def AssetMaintenance():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_compliance_check_tables()
    company_code = cmpdata[10]
    err = []
    msg = []

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add_vehicle_repair':
            repair_err = save_vehicle_repair(company_code, request.form, request.files.get('repair_file'))
            if repair_err:
                err.append(repair_err)
            else:
                msg.append('Vehicle repair note added.')
        elif action == 'delete_vehicle_repair':
            delete_vehicle_repair(company_code, request.form.get('repair_id'))
            msg.append('Vehicle repair note deleted.')

    _, vehicles = compliance_asset_summary(company_code)
    return render_template(
        'asset_maintenance.html',
        cmpdata=cmpdata,
        scac=scac,
        vehicles=vehicles,
        file_link=compliance_file_link,
        err=err,
        msg=msg,
    )


@main.route('/api/compliance/due-alerts', methods=['GET'])
@login_required
def ComplianceDueAlertsApi():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_compliance_check_tables()
    company_code = cmpdata[10]
    profile = compliance_profile(company_code, cmpdata[2], cmpdata[8])
    try:
        lead_days = int(request.args.get('lead_days') or profile.get('AlertLeadDays') or 30)
    except (TypeError, ValueError):
        lead_days = int(profile.get('AlertLeadDays') or 30)
    return jsonify(compliance_due_alert_payload(company_code, lead_days))


@main.route('/GeneralInformation', methods=['GET', 'POST'])
@login_required
def GeneralInformation():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_general_information_tables()
    company_code = cmpdata[10]
    err = []
    msg = []
    revealed_payment_method = None

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save_profile':
            update_general_company_profile(company_code, request.form)
            msg.append('Company profile updated.')
        elif action == 'save_item':
            if (request.form.get('category') or '').strip() == 'Payment Methods':
                item_err = upsert_secure_payment_method_from_general_form(company_code, request.form)
            else:
                item_err = upsert_general_information_item(company_code, request.form)
            if item_err:
                err.append(item_err)
            else:
                msg.append('General information item saved.')
        elif action == 'delete_item':
            if request.form.get('item_kind') == 'payment_method':
                delete_secure_payment_method(company_code, request.form.get('payment_method_id'))
            else:
                delete_general_information_item(company_code, request.form.get('item_id'))
            msg.append('General information item removed.')
        elif action == 'add_category':
            category_err = add_general_information_category(company_code, request.form.get('category_name'))
            if category_err:
                err.append(category_err)
            else:
                msg.append('General information category added.')
        elif action == 'delete_category':
            delete_general_information_category(company_code, request.form.get('category_name'))
            msg.append('General information category removed from selection list.')
        elif action == 'reveal_payment_method':
            confirm_password = request.form.get('confirm_password') or ''
            if not bcrypt.check_password_hash(current_user.password, confirm_password):
                err.append('Password confirmation failed. Secure payment method was not revealed.')
            else:
                revealed_payment_method = reveal_secure_payment_method(
                    company_code,
                    request.form.get('payment_method_id'),
                    current_user.username,
                )
                if not revealed_payment_method:
                    err.append('Secure payment method could not be found.')

    categories = general_information_categories(company_code)
    items = combined_general_information_items(company_code)
    profile = general_company_profile(company_code, scac, cmpdata[7] if len(cmpdata) > 7 else '')
    return render_template(
        'general_information.html',
        cmpdata=cmpdata,
        scac=scac,
        profile=profile,
        categories=categories,
        items=items,
        revealed_payment_method=revealed_payment_method,
        err=err,
        msg=msg,
    )


@main.route('/dispatch/kanban', methods=['GET'])
@login_required
def DispatchKanban():
    ensure_dispatch_kanban_tables()
    return render_template(
        'dispatch_kanban.html',
        cmpdata=cmpdata,
        scac=scac,
        options=kanban_options(),
    )


@main.route('/api/dispatch/calendar/events', methods=['GET'])
@login_required
def DispatchCalendarEvents():
    start = dispatch_parse_date(request.args.get('start'))
    end = dispatch_parse_date(request.args.get('end'))
    filters = {
        'driver': request.args.get('driver'),
        'customer': request.args.get('customer'),
        'status': request.args.get('status'),
        'terminal': request.args.get('terminal'),
        'search': request.args.get('search'),
        'range': request.args.get('range'),
    }
    events = calendar_events(start=start, end=end, filters=filters)
    return jsonify(events)


@main.route('/api/dispatch/calendar/capacity', methods=['GET'])
@login_required
def DispatchCalendarCapacity():
    start = dispatch_parse_date(request.args.get('start'))
    end = dispatch_parse_date(request.args.get('end'))
    filters = {
        'driver': request.args.get('driver'),
        'customer': request.args.get('customer'),
        'status': request.args.get('status'),
        'terminal': request.args.get('terminal'),
        'search': request.args.get('search'),
        'range': request.args.get('range'),
    }
    return jsonify(capacity_summary(start=start, end=end, filters=filters))


@main.route('/api/dispatch/calendar/event/<int:order_id>/move', methods=['POST'])
@login_required
def DispatchCalendarMove(order_id):
    payload = request.get_json(silent=True) or {}
    result, status_code = move_event(order_id, payload.get('start'), username=current_user.username)
    return jsonify(result), status_code


@main.route('/api/dispatch/calendar/event/<int:order_id>/update', methods=['POST'])
@login_required
def DispatchCalendarUpdate(order_id):
    payload = request.get_json(silent=True) or {}
    result, status_code = update_event(order_id, payload, username=current_user.username)
    return jsonify(result), status_code


@main.route('/api/dispatch/calendar/notes', methods=['GET'])
@login_required
def DispatchCalendarNotes():
    start = dispatch_parse_date(request.args.get('start'))
    end = dispatch_parse_date(request.args.get('end'))
    return jsonify(calendar_notes(start=start, end=end))


@main.route('/api/dispatch/calendar/note', methods=['POST'])
@login_required
def DispatchCalendarNoteSave():
    payload = request.get_json(silent=True) or {}
    result, status_code = save_calendar_note(payload, username=current_user.username)
    return jsonify(result), status_code


@main.route('/api/dispatch/calendar/note/<int:note_id>/delete', methods=['POST'])
@login_required
def DispatchCalendarNoteDelete(note_id):
    result, status_code = delete_calendar_note(note_id)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/jobs', methods=['GET'])
@login_required
def DispatchKanbanJobs():
    filters = {
        'company': request.args.get('company'),
        'driver': request.args.get('driver'),
        'customer': request.args.get('customer'),
        'terminal': request.args.get('terminal'),
        'status': request.args.get('status'),
        'range': request.args.get('range'),
    }
    return jsonify(kanban_jobs(filters=filters))


@main.route('/api/dispatch/kanban/options', methods=['GET'])
@login_required
def DispatchKanbanOptions():
    return jsonify(kanban_options())


@main.route('/api/dispatch/kanban/job/<int:order_id>/move', methods=['POST'])
@login_required
def DispatchKanbanMove(order_id):
    payload = request.get_json(silent=True) or {}
    username = getattr(current_user, 'username', None) or getattr(current_user, 'name', None) or 'dispatch'
    result, status_code = kanban_move_job(
        order_id,
        payload.get('workflow_status'),
        username=username,
        override_pin=bool(payload.get('override_pin')),
        reason=payload.get('reason'),
    )
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/job/<int:order_id>/update', methods=['POST'])
@login_required
def DispatchKanbanUpdate(order_id):
    payload = request.get_json(silent=True) or {}
    username = getattr(current_user, 'username', None) or getattr(current_user, 'name', None) or 'dispatch'
    result, status_code = kanban_update_job(order_id, payload, username=username)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/job/<int:order_id>/upload-proof', methods=['POST'])
@login_required
def DispatchKanbanUploadProof(order_id):
    username = getattr(current_user, 'username', None) or getattr(current_user, 'name', None) or 'dispatch'
    result, status_code = kanban_upload_proof(order_id, request.files.get('proof_pdf'), username=username)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/job/<int:order_id>/make-pin-options', methods=['GET'])
@login_required
def DispatchKanbanMakePinOptions(order_id):
    result, status_code = make_pin_candidates(order_id)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/job/<int:order_id>/make-pin', methods=['POST'])
@login_required
def DispatchKanbanMakePin(order_id):
    payload = request.get_json(silent=True) or {}
    username = getattr(current_user, 'username', None) or getattr(current_user, 'name', None) or 'dispatch'
    result, status_code = create_kanban_pin(order_id, payload, username=username)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/pin/<int:pin_id>/activate', methods=['POST'])
@login_required
def DispatchKanbanActivatePin(pin_id):
    result, status_code = activate_pin_pairing(pin_id)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/pin/<int:pin_id>/delete', methods=['POST'])
@login_required
def DispatchKanbanDeletePin(pin_id):
    result, status_code = delete_pin_pairing(pin_id)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/job/<int:order_id>/reviews', methods=['GET'])
@login_required
def DispatchKanbanReviews(order_id):
    result, status_code = kanban_review_logs(order_id)
    return jsonify(result), status_code


@main.route('/api/dispatch/kanban/job/<int:order_id>/review', methods=['POST'])
@login_required
def DispatchKanbanLogReview(order_id):
    payload = request.get_json(silent=True) or {}
    username = getattr(current_user, 'username', None) or getattr(current_user, 'name', None) or 'dispatch'
    result, status_code = kanban_log_review(order_id, payload, username=username)
    return jsonify(result), status_code


@main.route('/financial/collection-kanban', methods=['GET'])
@login_required
@financial_mfa_required
def CollectionKanban():
    return render_template(
        'collection_kanban.html',
        cmpdata=cmpdata,
        scac=scac,
        options=collection_options(),
    )


@main.route('/api/financial/collection-kanban/jobs', methods=['GET'])
@login_required
@financial_mfa_required
def CollectionKanbanJobs():
    filters = {
        'customer': request.args.get('customer'),
        'status': request.args.get('status'),
        'range': request.args.get('range'),
        'search': request.args.get('search'),
    }
    return jsonify(collection_jobs(filters=filters))


@main.route('/api/financial/collection-kanban/job/<int:order_id>/update', methods=['POST'])
@login_required
@financial_mfa_required
def CollectionKanbanUpdate(order_id):
    payload = request.get_json(silent=True) or {}
    result, status_code = update_collection_job(order_id, payload)
    return jsonify(result), status_code


@main.route('/api/financial/collection-kanban/job/<int:order_id>/emails', methods=['GET'])
@login_required
@financial_mfa_required
def CollectionKanbanEmails(order_id):
    result, status_code = collection_email_logs(order_id)
    return jsonify(result), status_code


@main.route('/api/financial/collection-kanban/job/<int:order_id>/email', methods=['POST'])
@login_required
@financial_mfa_required
def CollectionKanbanSendEmail(order_id):
    payload = request.get_json(silent=True) or {}
    result, status_code = send_collection_email(order_id, payload)
    return jsonify(result), status_code


@main.route('/api/financial/collection-kanban/job/<int:order_id>/rate-con', methods=['POST'])
@login_required
@financial_mfa_required
def CollectionKanbanUploadRateCon(order_id):
    result, status_code = upload_collection_rate_con(order_id, request.files.get('rate_con'), request.form)
    return jsonify(result), status_code


@main.route('/api/financial/collection-kanban/job/<int:order_id>/call', methods=['POST'])
@login_required
@financial_mfa_required
def CollectionKanbanLogCall(order_id):
    payload = request.get_json(silent=True) or {}
    username = getattr(current_user, 'username', None) or getattr(current_user, 'name', None) or 'collection'
    result, status_code = log_collection_call(order_id, payload, username=username)
    return jsonify(result), status_code


@main.route('/api/financial/collection-kanban/options', methods=['GET'])
@login_required
@financial_mfa_required
def CollectionKanbanOptions():
    return jsonify(collection_options())

@main.route('/chartdata', methods=['GET', 'POST'])
def chartdata():
    acct = request.values['thisacct']
    timestyle = request.values['thesemonths']
    import ast
    acct = ast.literal_eval(acct)
    print(acct,timestyle)
    print(type(timestyle))
    if timestyle in ['6', '12', '18', '24']:
        nmonths = int(timestyle)
        labeld = []
        datad = []
        lablist=monvals(nmonths)
        rgba = []
        rgb = []
        colors = [[31,105,161], [31,161,65], [161,141,31], [161,57,31], [161,31,141], [161,31,63], [31,53,161], [31,161,126]]
        for ix, plotitem in enumerate(acct):
            datad.append(getmonths(plotitem,nmonths,1))
            labeld.append(plotitem)
            rgba.append(f'rgba({colors[ix][0]}, {colors[ix][1]}, {colors[ix][2]}, 0.3)')
            rgb.append(f'rgb({colors[ix][0]}, {colors[ix][1]}, {colors[ix][2]})')
        return jsonify({'lablist' : lablist,
                        'labeld'  : labeld,
                        'datad'   : datad,
                        'rgba'    : rgba,
                        'rgb'     : rgb
                        })
    elif timestyle == 'lymon' or 'tymon':
        thisyear = datetime.datetime.today().year
        thismonth = datetime.datetime.today().month
        lastyear = thisyear - 1
        print(lastyear,thismonth)
        labeld = []
        datad = []
        lablist=['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        if timestyle == 'lymon':
            start = thismonth
            stop = thismonth+12
            for ix, lab in enumerate(lablist):
                lablist[ix] = f'{lab} {lastyear}'
        else:
            start = 1
            stop = thismonth
            lablist = lablist[:stop-1]
            for ix, lab in enumerate(lablist):
                lablist[ix] = f'{lab} {thisyear}'


        rgba = []
        rgb = []
        colors = [[31,105,161], [31,161,65], [161,141,31], [161,57,31], [161,31,141], [161,31,63], [31,53,161], [31,161,126]]
        for ix, plotitem in enumerate(acct):
            datad.append(getmonths(plotitem,start,stop))
            labeld.append(plotitem)
            rgba.append(f'rgba({colors[ix][0]}, {colors[ix][1]}, {colors[ix][2]}, 0.3)')
            rgb.append(f'rgb({colors[ix][0]}, {colors[ix][1]}, {colors[ix][2]})')
        return jsonify({'lablist' : lablist,
                        'labeld'  : labeld,
                        'datad'   : datad,
                        'rgba'    : rgba,
                        'rgb'     : rgb
                        })


@main.route('/', methods=['GET', 'POST'])
def index():
    info = [''] * 9
    if request.method == 'POST':
        sendnow = request.values.get('sendnow')
        if sendnow is not None:
            name = request.values.get('name')
            email = request.values.get('email')
            phone = request.values.get('phone')
            message = request.values.get('message')
            print(f'Sending email from {name}, {email}, {phone}, {message}')
            info[0] = today_str
            info[1] = message
            info[2] = email
            info[3] = phone
            info[4] = name
            info[5] = True
            info[6] = f'Your contact request has been sent'
            num_finds = check_person(info)
            if num_finds < 2:
                add_person(info)
                emaildata = email_template('contact', info)
                print('Sending Email to', emaildata)
                info_mimemail(emaildata)
            else:
                print('Too many attempts to contact')
                info[5] = True
                info[6] = f'Already received request, please allow time to respond'

    srcpath = statpath('')
    front_template = 'index.html'
    front_page_variant = PublicSite_setup.get(scac, {}).get('front_page_variant', 'classic')
    if scac == 'OSLM' and front_page_variant == 'professional':
        front_template = 'index_professional.html'
    return render_template(f'companysite/{scac}/{front_template}',srcpath=srcpath, cmpdata=cmpdata, scac=scac,info=info)

@main.route('/About')
def About():
    lang = 'English'
    srcpath = statpath('')
    return render_template(f'companysite/{scac}/about.html',srcpath=srcpath,cmpdata=cmpdata, scac=scac, lang=lang)

@main.route('/Whatapp', methods=['GET', 'POST'])
def Whatapp():
    token = os.environ['TWILIO_AUTH_TOKEN']
    print('token=',token)
    msg = request.form.get('Body')
    msg = msg.strip()
    sessionph = request.form.get('From')
    print('sessionph =',sessionph)
    media_files = []
    num_media = int(request.values.get("NumMedia"))
    if num_media > 0:
        for idx in range(num_media):
            media_url = request.values.get(f'MediaUrl{idx}')
            mime_type = request.values.get(f'MediaContentType{idx}')
            req = requests.get(media_url)
            file_extension = mimetypes.guess_extension(mime_type)
            file_extension = file_extension.replace('.jpe', '.jpg')
            file_extension = file_extension.replace('.jpeg', '.jpg')
            media_sid = os.path.basename(urlparse(media_url).path)
            media_files.append(media_sid+file_extension)
            media_path = addpath('static/data/processing/whatsapp/')
            with open(f"{media_path}{media_sid}{file_extension}", 'wb') as f:
                f.write(req.content)
        print(media_files)

    respmsg = msg_analysis(msg, sessionph, media_files)

    resp = MessagingResponse()
    msg = resp.message("{}".format(respmsg))

    lines = respmsg.splitlines()
    line1 = lines[0]

    if 'Attachment' in line1 and len(lines)>1:
        file1 = lines[1].strip()
        my_path = 'https://www.oslbox.com/'
        #my_path = 'https://7223-2601-150-100-8c10-2928-d420-6968-460f.ngrok.io/'
        my_url = my_path + file1
        print('myurl = ',my_url)
        msg.media(my_url)
    return str(resp)


@main.route('/AboutClass8', methods=['GET', 'POST'])
def AboutClass8():
    info = ['']*7
    thisnow = now + timedelta(1)
    info[0] = thisnow.strftime("%Y-%m-%dT%H:%M")

    if request.method == 'POST':
        setappt = request.values.get('setappt')
        if setappt is not None:
            info[0] = request.values.get('date')
            info[1] = request.values.get('location')
            info[2] = request.values.get('email')
            info[3] = request.values.get('phone')
            info[4] = request.values.get('contact')
            info[5] = True
            info[6] = f'Email Sent To {info[2]} at {info[1]} confirming appointment on {info[0]} at {info[1]}'
            emaildata = email_template('class8demo', info)
            print('Sending Email to', emaildata)
            info_mimemail(emaildata)

        cancel = request.values.get('cancel')
        if cancel is not None: info = ['']*7

    for i in info:
        print(i)

    srcpath = statpath('')
    return render_template('AboutClass8.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac, info = info)

@main.route('/People_Forms', methods=['GET', 'POST'])
def People_Forms():
    if request.method == 'POST':
        today = datetime.datetime.today().strftime('%Y-%m-%d')
        appnum = 0
        vals = ['fname', 'mnames', 'lname', 'addr1', 'addr2', 'addr3',
                'idtype', 'tid', 'tel', 'email', 'assoc1', 'assoc2', 'date1', 'yrs']
        a = list(range(len(vals)))
        i = 0
        for v in vals:
            a[i] = request.values.get(v)
            i = i+1
        exporter = request.values.get('exporter')
        consignee = request.values.get('consignee')
        notify = request.values.get('notify')
        driver = request.values.get('driver')
        if exporter is not None:
            ptype = "exporter"
        if consignee is not None:
            ptype = "consignee"
        if notify is not None:
            ptype = "notify"
        if driver is not None:
            ptype = "applicant"
        try:
            company = a[0] + ' ' + a[2]
        except:
            company = a[0]

        input = People(Ptype=ptype, Company=company, First=a[0], Middle=a[1], Last=a[2], Addr1=a[3], Addr2=a[4], Addr3=a[5], Temp1=a[13], Temp2='NewApp',
                       Idtype=a[6], Idnumber=a[7], Telephone=a[8], Email=a[9], Associate1=a[10], Associate2=a[11], Date1=today, Date2=None, Source=None, Accountid=None)
        db.session.add(input)
        db.session.commit()

        if exporter is not None:
            ptype = "consignee"
        if consignee is not None:
            ptype = "notify"
        if notify is not None:
            ptype = "completed"
            pdata = People.query.filter(People.Temp2 == 'NewApp').all()
            for pdat in pdata:
                pdat.Temp2 = '2'
                db.session.commit()
            from email_appl import email_app_exporter
            email_app_exporter(pdata)

        if driver is not None:

            pdat = People.query.filter((People.Ptype == 'applicant') &
                                       (People.Company == company)).first()
            appnum = 'Fapp'+str(pdat.id)
            ptype = "completed"
            from email_appl import email_app
            email_app(pdat)

            return render_template('employment.html', cmpdata=cmpdata, scac=scac, ptype=ptype, appnum=appnum, phone=phone, today=today)
    else:
        ptype = "exporter"

    srcpath = statpath('')
    return render_template(f'companysite/{scac}/pforms.html', cmpdata=cmpdata, scac=scac, ptype=ptype, srcpath=srcpath)

@main.route('/Employment')
def Employment():
    ptype = 'driver'
    srcpath = statpath('')
    return render_template('employment.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac, ptype=ptype, today=today.date(), phone=cmpdata[7],email=cmpdata[8])

@main.route('/support')
def support():
    srcpath = statpath('')
    return render_template('support.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac, today=today.date(), phone=cmpdata[7],email=cmpdata[8])

@main.route('/privacy')
def privacy():
    srcpath = statpath('')
    return render_template('privacy.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac, today=today.date(), phone=cmpdata[7],email=cmpdata[8])

@main.route('/app_overview')
def app_overview():
    srcpath = statpath('')
    return render_template('app_overview.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac, today=today.date(), phone=cmpdata[7],email=cmpdata[8])

@main.route('/agents_overview')
def agents_overview():
    srcpath = statpath('')
    return render_template('agents_overview.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac, today=today.date(), phone=cmpdata[7],email=cmpdata[8])


@main.route('/admin/bot-skills', methods=['GET', 'POST'])
@login_required
def BotSkillSettings():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)

    if request.method == 'POST':
        skill_key = (request.form.get('skill_key') or '').strip()
        if skill_key not in SKILL_DEFINITIONS:
            abort(400)
        enabled = request.form.get('enabled') == '1'
        set_bot_skill_enabled(skill_key, enabled, current_user.username)
        flash(
            f"{SKILL_DEFINITIONS[skill_key]['name']} turned {'on' if enabled else 'off'}.",
            'success',
        )
        return redirect(url_for('main.BotSkillSettings'))

    skills = bot_skill_rows()
    usage = usage_dashboard()
    skill_usage = {row['skill_key']: row for row in usage['skill_rows']}
    for skill in skills:
        skill['usage'] = skill_usage.get(skill['key'], {
            'input_tokens': 0, 'output_tokens': 0, 'cache_read_tokens': 0,
            'total_tokens': 0, 'cost': 0,
        })

    return render_template(
        'bot_skill_settings.html',
        cmpdata=cmpdata,
        scac=scac,
        skills=skills,
        usage=usage,
        skill_definitions=SKILL_DEFINITIONS,
    )

@main.route('/Calculator', methods=['GET', 'POST'])
def Calculator():
    import ast
    from viewfuncs import d2s
    if request.method == 'POST':
        alldata = request.values.get('alldata')
        alldata = ast.literal_eval(alldata)
        print(alldata)
        l = len(alldata)
        print(l)
        a1 = request.form['len']
        a2 = request.form['wid']
        a3 = request.form['hei']
        a4 = float(a1)*float(a2)*float(a3)
        a6 = request.form['unt']
        a7 = request.form['wtunt']
        b1 = request.form['cst']
        b1 = Decimal(b1.strip('$'))
        a6 = int(a6)
        a7 = int(a7)
        if a6 == 1:
            a4 = a4/61023.7
        if a6 == 2:
            a4 = a4/35.3147
        if a6 == 3:
            a4 = a4/1000000.
        wtkg = a4*166.67
        wtlb = wtkg*2.20462
        wtkgstr = d2s(str(wtkg))
        wtlbstr = d2s(str(wtlb))
        a5 = math.ceil(a4)
        a4 = round(a4, 2)
        b2 = a5*float(b1)
        if a7 == 1:
            wt = wtlbstr
        else:
            wt = wtkgstr
        total = float(wt)
        for data in alldata:
            total = total+float(data[3])

        alldata.append([a1, a2, a3, wt])

        # Recalculate all in case units have changed
        newalldata = []
        total = 0
        for data in alldata:
            a1 = data[0]
            a2 = data[1]
            a3 = data[2]
            a4 = float(a1)*float(a2)*float(a3)
            if a6 == 1:
                a4 = a4/61023.7
            if a6 == 2:
                a4 = a4/35.3147
            if a6 == 3:
                a4 = a4/1000000.
            if a7 == 2:
                wt = a4*166.67
            if a7 == 1:
                wt = a4*166.67*2.20462
            total = total+float(wt)
            newalldata.append([a1, a2, a3, d2s(wt)])
        a4 = round(a4, 2)
        finalcost = total*float(b1)
        finalwt = d2s(total)
        finalcost = d2s(finalcost)

    else:
        a1 = 1
        a2 = 1
        a3 = 1
        a4 = 1
        a5 = 1
        a6 = 1
        a7 = 1
        b1 = 25
        b2 = 1
        wtkgstr = ''
        wtlbstr = ''
        alldata = []
        fdata = []
        newalldata = []
        finalwt = ''
        finalcost = ''
    srcpath = statpath('')
    return render_template('calculator.html', srcpath=srcpath,cmpdata=cmpdata, scac=scac, finalcost=finalcost, a1=a1, a2=a2, a3=a3, a4=a4, a5=a5, a6=a6, a7=a7, b1=b1, b2=b2, wtkg=wtkgstr, wtlb=wtlbstr, alldata=newalldata, finalwt=finalwt)


@main.route('/Class8Main/<genre>', methods=['GET', 'POST'])
@login_required

def Class8Main(genre):

    print('routes.py 237: The genre is',genre)
    if (
        genre == 'Trucking'
        and request.method == 'POST'
        and request.values.get('Money Flow') in ['Receive Payments', 'Receive by Acct']
    ):
        return redirect(url_for('main.ReceiveByAccount'))
    if genre == 'Trucking' and request.args.get('collection_receive_order_id'):
        return redirect(url_for('main.ReceiveByAccount'))
    if genre == 'Banking':
        return redirect(url_for('main.Banking'))
    if genre == 'Planning':
        return redirect(url_for('main.DispatchPlanningCalendar'))
    if genre == 'Calendar':
        return redirect(url_for('main.AdminCalendar'))
    if genre == 'Reports':
        return redirect(url_for('main.ProfitLossYearToDate'))
    if genre in FINANCIAL_GENRES:
        redirect_response = financial_mfa_redirect()
        if redirect_response is not None:
            return redirect_response
    if (
        request.method == 'POST'
        and request.values.get('Return') is not None
        and request.values.get('callfrom') == 'collection_kanban'
    ):
        return redirect(url_for('main.CollectionKanban'))
    genre_data, table_data, err, leftsize, tabletitle, table_filters, task_boxes, tfilters, tboxes, jscripts,\
    taskon, task_focus, task_iter, tasktype, holdvec, keydata, entrydata, username, checked_data, viewport, tablesetup = Table_maker(genre)
    if viewport and viewport[0] == 'redirect_collection_kanban':
        return redirect(url_for('main.CollectionKanban'))
    if taskon == 'New': err, viewport = checkfor_fileupload(err, task_iter, viewport)

    rightsize = 12 - leftsize
    runpins = request.values.get('RunPins')
    if runpins is not None:
        return redirect(url_for('main.getpinsnow'))

    return render_template('Class8.html',cmpdata=cmpdata, scac=scac,  genre_data = genre_data, table_data=table_data, err=err, checked_data = checked_data,
                           leftsize=leftsize, rightsize=rightsize, tabletitle=tabletitle, table_filters = table_filters,task_boxes = task_boxes, tfilters=tfilters, tboxes=tboxes, dt1 = jscripts,
                           taskon=taskon, task_focus=task_focus, task_iter=task_iter, tasktype=tasktype, holdvec=holdvec, keydata = keydata, entrydata = entrydata, username=username, genre=genre, viewport=viewport, tablesetup=tablesetup)




@main.route('/Revenue', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def Revenue():
    #print('Made it to the Revenue Data Center')
    title1,col1,data1,title2,col2,data2,title3,col3,data3,tabon = get_revenues()
    return render_template('revenues.html', cmpdata=cmpdata, scac=scac, title1=title1, col1=col1, data1=data1, title2=title2, col2=col2, data2=data2, title3=title3, col3=col3, data3=data3, tabon=tabon)


def ensure_business_report_tables():
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS business_report_allocations (
            id INT AUTO_INCREMENT PRIMARY KEY,
            AccountName VARCHAR(100) NOT NULL,
            AllocationMethod VARCHAR(100) NOT NULL,
            BasisType VARCHAR(45) NOT NULL DEFAULT 'port_entry',
            BasisLabel VARCHAR(100) NOT NULL,
            PeriodStart DATE,
            PeriodEnd DATE,
            BasisQty DECIMAL(12, 2) DEFAULT 0,
            Notes TEXT,
            Active TINYINT DEFAULT 1,
            CreatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UpdatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_business_report_alloc_account (AccountName),
            INDEX idx_business_report_alloc_active (Active)
        )
    """))
    db.session.commit()
    inspector = inspect(db.engine)
    columns = [column['name'] for column in inspector.get_columns('business_report_allocations')]
    if 'BasisType' not in columns:
        db.session.execute(text("""
            ALTER TABLE business_report_allocations
            ADD COLUMN BasisType VARCHAR(45) NOT NULL DEFAULT 'port_entry'
        """))
        db.session.commit()
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS operations_indirect_allocations (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Co VARCHAR(2) NOT NULL,
            AccountName VARCHAR(100) NOT NULL,
            PeriodStart DATE NOT NULL,
            PeriodEnd DATE NOT NULL,
            AmountCents INT NOT NULL DEFAULT 0,
            AllocationBasis VARCHAR(45) NOT NULL DEFAULT 'daily',
            AutoUpdate TINYINT NOT NULL DEFAULT 0,
            Notes TEXT,
            Active TINYINT DEFAULT 1,
            CreatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UpdatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_ops_indirect_alloc_co (Co),
            INDEX idx_ops_indirect_alloc_account (AccountName),
            INDEX idx_ops_indirect_alloc_period (PeriodStart, PeriodEnd),
            INDEX idx_ops_indirect_alloc_active (Active)
        )
    """))
    db.session.commit()
    columns = [column['name'] for column in inspector.get_columns('operations_indirect_allocations')]
    if 'AutoUpdate' not in columns:
        db.session.execute(text("""
            ALTER TABLE operations_indirect_allocations
            ADD COLUMN AutoUpdate TINYINT NOT NULL DEFAULT 0
        """))
        db.session.commit()
    if 'AllocationBasis' not in columns:
        db.session.execute(text("""
            ALTER TABLE operations_indirect_allocations
            ADD COLUMN AllocationBasis VARCHAR(45) NOT NULL DEFAULT 'daily'
        """))
        db.session.commit()
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS operations_ga_allocations (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Co VARCHAR(2) NOT NULL,
            AccountName VARCHAR(100) NOT NULL,
            PeriodStart DATE NOT NULL,
            PeriodEnd DATE NOT NULL,
            AmountCents INT NOT NULL DEFAULT 0,
            AllocationBasis VARCHAR(45) NOT NULL DEFAULT 'daily',
            AutoUpdate TINYINT NOT NULL DEFAULT 0,
            Notes TEXT,
            Active TINYINT DEFAULT 1,
            CreatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UpdatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_ops_ga_alloc_co (Co),
            INDEX idx_ops_ga_alloc_account (AccountName),
            INDEX idx_ops_ga_alloc_period (PeriodStart, PeriodEnd),
            INDEX idx_ops_ga_alloc_active (Active)
        )
    """))
    db.session.commit()
    columns = [column['name'] for column in inspector.get_columns('operations_ga_allocations')]
    if 'AutoUpdate' not in columns:
        db.session.execute(text("""
            ALTER TABLE operations_ga_allocations
            ADD COLUMN AutoUpdate TINYINT NOT NULL DEFAULT 0
        """))
        db.session.commit()
    if 'AllocationBasis' not in columns:
        db.session.execute(text("""
            ALTER TABLE operations_ga_allocations
            ADD COLUMN AllocationBasis VARCHAR(45) NOT NULL DEFAULT 'daily'
        """))
        db.session.commit()
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS business_report_cache (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Co VARCHAR(2) NOT NULL,
            ReportMode VARCHAR(24) NOT NULL,
            DateFrom DATE NOT NULL,
            DateTo DATE NOT NULL,
            PayloadJson LONGTEXT NOT NULL,
            GeneratedBy VARCHAR(120),
            Active TINYINT NOT NULL DEFAULT 1,
            CreatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UpdatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_business_report_cache_lookup (Co, ReportMode, DateFrom, DateTo, Active),
            INDEX idx_business_report_cache_updated (UpdatedAt)
        )
    """))
    db.session.commit()


def business_report_parse_date(value, fallback):
    if not value:
        return fallback
    try:
        return datetime.datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return fallback


def business_report_money(cents):
    return "${:,.2f}".format((cents or 0) / 100)


def business_report_empty_tax_rollup_pl():
    return {
        'income_rows': [],
        'expense_rows': [],
        'income_report_rows': [],
        'expense_report_rows': [],
        'total_income_cents': 0,
        'total_expenses_cents': 0,
        'unassigned_expenses_cents': 0,
        'net_income_cents': 0,
        'total_income': business_report_money(0),
        'total_expenses': business_report_money(0),
        'unassigned_expenses': business_report_money(0),
        'net_income': business_report_money(0),
    }


def business_report_empty_operations_totals():
    return {
        'income_cents': 0,
        'direct_cents': 0,
        'gross_profit_cents': 0,
        'indirect_cents': 0,
        'net_ops_cents': 0,
        'ga_cents': 0,
        'net_profit_cents': 0,
        'income': business_report_money(0),
        'direct': business_report_money(0),
        'gross_profit': business_report_money(0),
        'indirect': business_report_money(0),
        'net_ops': business_report_money(0),
        'ga': business_report_money(0),
        'net_profit': business_report_money(0),
    }


def business_report_jsonable(value):
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return {key: business_report_jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [business_report_jsonable(item) for item in value]
    return value


def business_report_cache_load(company_code, report_mode, start_date, end_date):
    row = db.session.execute(text("""
        SELECT id, PayloadJson, GeneratedBy, UpdatedAt
        FROM business_report_cache
        WHERE Co = :company_code
          AND ReportMode = :report_mode
          AND DateFrom = :start_date
          AND DateTo = :end_date
          AND Active = 1
        ORDER BY UpdatedAt DESC, id DESC
        LIMIT 1
    """), {
        'company_code': company_code,
        'report_mode': report_mode,
        'start_date': start_date,
        'end_date': end_date,
    }).mappings().first()
    if not row:
        return None, None
    payload = json.loads(row['PayloadJson'] or '{}')
    if payload.get('cache_version') != BUSINESS_REPORT_CACHE_VERSION:
        return None, None
    return payload, row


def business_report_cache_save(company_code, report_mode, start_date, end_date, payload):
    now = datetime.datetime.utcnow()
    payload = dict(payload or {})
    payload['cache_version'] = BUSINESS_REPORT_CACHE_VERSION
    db.session.execute(text("""
        UPDATE business_report_cache
        SET Active = 0, UpdatedAt = :updated_at
        WHERE Co = :company_code
          AND ReportMode = :report_mode
          AND DateFrom = :start_date
          AND DateTo = :end_date
          AND Active = 1
    """), {
        'company_code': company_code,
        'report_mode': report_mode,
        'start_date': start_date,
        'end_date': end_date,
        'updated_at': now,
    })
    db.session.execute(text("""
        INSERT INTO business_report_cache
            (Co, ReportMode, DateFrom, DateTo, PayloadJson, GeneratedBy, Active, CreatedAt, UpdatedAt)
        VALUES
            (:company_code, :report_mode, :start_date, :end_date, :payload_json, :generated_by, 1, :created_at, :updated_at)
    """), {
        'company_code': company_code,
        'report_mode': report_mode,
        'start_date': start_date,
        'end_date': end_date,
        'payload_json': json.dumps(business_report_jsonable(payload)),
        'generated_by': getattr(current_user, 'username', None) or getattr(current_user, 'email', None) or '',
        'created_at': now,
        'updated_at': now,
    })
    db.session.commit()
    return business_report_cache_load(company_code, report_mode, start_date, end_date)


def business_report_cache_info(row, source):
    if not row:
        return {'source': source, 'updated_at': '', 'generated_by': ''}
    updated_at = row.get('UpdatedAt') if hasattr(row, 'get') else row['UpdatedAt']
    generated_by = row.get('GeneratedBy') if hasattr(row, 'get') else row['GeneratedBy']
    if isinstance(updated_at, datetime.datetime):
        updated_at = updated_at.strftime('%Y-%m-%d %H:%M')
    return {
        'source': source,
        'updated_at': updated_at or '',
        'generated_by': generated_by or '',
    }


def business_report_number(value, places=2):
    return f"{value or 0:,.{places}f}"


def business_report_as_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return value


def business_report_date_text(value):
    value = business_report_as_date(value)
    return value.strftime('%Y-%m-%d') if value else ''


def business_report_as_datetime(value, end_of_day=False):
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.max if end_of_day else datetime.time.min)
    return value


BUSINESS_REPORT_CACHE_VERSION = 3

BUSINESS_REPORT_BASIS_TYPES = {
    'port_entry': 'Port Entries',
    'day': 'Days',
    'mile': 'Miles',
}

OPERATIONS_ALLOCATION_BASIS_TYPES = {
    'daily': 'Daily Spread',
    'port_trip': 'Port Trips',
}


def business_report_cents(value):
    try:
        clean = str(value).replace('$', '').replace(',', '').strip()
        if clean in ['', '-']:
            return 0
        return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))
    except:
        return 0


def business_report_each_day(start_date, end_date):
    current = start_date
    while current <= end_date:
        yield current
        current += datetime.timedelta(days=1)


def business_report_allocate_cents_by_day(daily_totals, start_date, end_date, amount_cents):
    if start_date is None or end_date is None or end_date < start_date or not amount_cents:
        return
    days = (end_date - start_date).days + 1
    sign = 1 if amount_cents >= 0 else -1
    base_abs, remainder_abs = divmod(abs(int(amount_cents)), days)
    for index, day in enumerate(business_report_each_day(start_date, end_date)):
        cents_for_day = sign * (base_abs + (1 if index < remainder_abs else 0))
        daily_totals[day] = daily_totals.get(day, 0) + cents_for_day


def business_report_allocate_cents_by_period(daily_totals, period_start, period_end, amount_cents, include_start, include_end):
    if period_start is None or period_end is None or period_end < period_start or not amount_cents:
        return
    if include_start is None:
        include_start = period_start
    if include_end is None:
        include_end = period_end
    include_start = max(include_start, period_start)
    include_end = min(include_end, period_end)
    if include_end < include_start:
        return
    days = (period_end - period_start).days + 1
    sign = 1 if amount_cents >= 0 else -1
    base_abs, remainder_abs = divmod(abs(int(amount_cents)), days)
    for index, day in enumerate(business_report_each_day(period_start, period_end)):
        if include_start <= day <= include_end:
            cents_for_day = sign * (base_abs + (1 if index < remainder_abs else 0))
            daily_totals[day] = daily_totals.get(day, 0) + cents_for_day


def business_report_port_trip_counts_by_day(start_date, end_date):
    if not start_date or not end_date or end_date < start_date:
        return {}
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    trip_day = func.date(Interchange.Date)
    rows = db.session.query(
        trip_day.label('trip_day'),
        func.count(func.distinct(Interchange.PortTrip)).label('trip_count'),
    ).filter(
        Interchange.Date >= start_dt,
        Interchange.Date <= end_dt,
        Interchange.PortTrip.isnot(None),
        Interchange.PortTrip > 0,
    ).group_by(trip_day).all()
    counts = {}
    for row in rows:
        day = business_report_parse_date(str(row.trip_day), None)
        if day:
            counts[day] = int(row.trip_count or 0)
    return counts


def business_report_allocate_cents_by_port_trip(daily_totals, period_start, period_end, amount_cents, include_start, include_end):
    if not period_start or not period_end or period_end < period_start:
        return
    include_start = max(period_start, include_start)
    include_end = min(period_end, include_end)
    if include_end < include_start:
        return
    trip_counts = business_report_port_trip_counts_by_day(period_start, period_end)
    total_trips = sum(trip_counts.values())
    if not total_trips:
        return
    included_days = [day for day in business_report_each_day(include_start, include_end) if trip_counts.get(day, 0)]
    included_trip_count = sum(trip_counts.get(day, 0) for day in included_days)
    if not included_trip_count:
        return
    target_amount = int(amount_cents or 0) * included_trip_count // total_trips
    shares = []
    allocated = 0
    for day in included_days:
        numerator = int(amount_cents or 0) * trip_counts[day]
        whole_share = numerator // total_trips
        remainder = numerator % total_trips
        shares.append((day, whole_share, remainder))
        allocated += whole_share
    remainder_cents = target_amount - allocated
    for index, (day, whole_share, remainder) in enumerate(sorted(shares, key=lambda item: item[2], reverse=True)):
        extra = 1 if index < remainder_cents else 0
        daily_totals[day] = daily_totals.get(day, 0) + whole_share + extra


def business_report_week_start(day):
    return day - datetime.timedelta(days=day.weekday())

def ensure_tax_rollup_tables():
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS tax_rollup_categories (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Co VARCHAR(2) NOT NULL,
            Section VARCHAR(20) NOT NULL DEFAULT 'Expense',
            Name VARCHAR(100) NOT NULL,
            ParentId INT NULL,
            SortOrder INT NOT NULL DEFAULT 0,
            Active TINYINT NOT NULL DEFAULT 1,
            CreatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UpdatedAt DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_tax_rollup_categories_co (Co),
            INDEX idx_tax_rollup_categories_section (Section),
            INDEX idx_tax_rollup_categories_parent (ParentId),
            INDEX idx_tax_rollup_categories_active (Active)
        )
    """))
    db.session.commit()
    inspector = inspect(db.engine)
    columns = [column['name'] for column in inspector.get_columns('tax_rollup_categories')]
    if 'Section' not in columns:
        after_clause = ' AFTER Co' if db.engine.dialect.name == 'mysql' else ''
        db.session.execute(text("""
            ALTER TABLE tax_rollup_categories
            ADD COLUMN Section VARCHAR(20) NOT NULL DEFAULT 'Expense'""" + after_clause))
        db.session.execute(text("UPDATE tax_rollup_categories SET Section = 'Expense' WHERE Section IS NULL OR Section = ''"))
        db.session.commit()


def tax_rollup_categories(company_code, section='Expense'):
    rows = db.session.execute(text("""
        SELECT id, Co, Section, Name, ParentId, SortOrder, Active
        FROM tax_rollup_categories
        WHERE Co = :company_code AND Section = :section AND Active = 1
        ORDER BY COALESCE(ParentId, 0), SortOrder, Name
    """), {'company_code': company_code, 'section': section}).mappings().all()
    items = []
    by_id = {}
    for row in rows:
        item = {
            'id': row['id'],
            'co': row['Co'],
            'section': row['Section'] or section,
            'name': row['Name'] or '',
            'parent_id': row['ParentId'],
            'sort_order': row['SortOrder'] or 0,
            'level': 0,
            'path': row['Name'] or '',
        }
        items.append(item)
        by_id[item['id']] = item

    def build_path(item, seen=None):
        seen = seen or set()
        if item['id'] in seen:
            return item['name']
        seen.add(item['id'])
        parent = by_id.get(item['parent_id'])
        if parent is None:
            item['level'] = 0
            return item['name']
        parent_path = build_path(parent, seen)
        item['level'] = parent.get('level', 0) + 1
        return f"{parent_path}:{item['name']}" if parent_path else item['name']

    for item in items:
        item['path'] = build_path(item)

    children = {}
    roots = []
    for item in items:
        children.setdefault(item['parent_id'], []).append(item)
        if item['parent_id'] not in by_id:
            roots.append(item)

    for child_items in children.values():
        child_items.sort(key=lambda item: (item['sort_order'], item['name'].lower(), item['id']))
        for index, item in enumerate(child_items):
            item['can_move_up'] = index > 0
            item['can_move_down'] = index < len(child_items) - 1
    roots.sort(key=lambda item: (item['sort_order'], item['name'].lower(), item['id']))

    ordered = []

    def add_children(parent_item):
        ordered.append(parent_item)
        for child in children.get(parent_item['id'], []):
            add_children(child)

    for root in roots:
        add_children(root)

    return ordered


def next_tax_rollup_sort(company_code, parent_id, section='Expense'):
    max_sort = db.session.execute(text("""
        SELECT COALESCE(MAX(SortOrder), 0)
        FROM tax_rollup_categories
        WHERE Co = :company_code
          AND Section = :section
          AND Active = 1
          AND (
                (:parent_id IS NULL AND ParentId IS NULL)
                OR ParentId = :parent_id
          )
    """), {'company_code': company_code, 'parent_id': parent_id, 'section': section}).scalar() or 0
    return int(max_sort) + 10


def normalize_tax_rollup_sibling_sort(company_code, parent_id, section='Expense'):
    rows = db.session.execute(text("""
        SELECT id
        FROM tax_rollup_categories
        WHERE Co = :company_code
          AND Section = :section
          AND Active = 1
          AND (
                (:parent_id IS NULL AND ParentId IS NULL)
                OR ParentId = :parent_id
          )
        ORDER BY SortOrder, Name, id
    """), {'company_code': company_code, 'parent_id': parent_id, 'section': section}).mappings().all()
    for index, row in enumerate(rows, start=1):
        db.session.execute(text("""
            UPDATE tax_rollup_categories
               SET SortOrder = :sort_order
             WHERE id = :category_id
        """), {'sort_order': index * 10, 'category_id': row['id']})
    db.session.commit()


def move_tax_rollup_category(company_code, category_id, direction, section='Expense'):
    category = db.session.execute(text("""
        SELECT id, ParentId
        FROM tax_rollup_categories
        WHERE id = :category_id
          AND Co = :company_code
          AND Section = :section
          AND Active = 1
    """), {'category_id': category_id, 'company_code': company_code, 'section': section}).mappings().first()
    if category is None:
        return 'Tax category not found.'

    parent_id = category['ParentId']
    normalize_tax_rollup_sibling_sort(company_code, parent_id, section)
    siblings = db.session.execute(text("""
        SELECT id, SortOrder
        FROM tax_rollup_categories
        WHERE Co = :company_code
          AND Section = :section
          AND Active = 1
          AND (
                (:parent_id IS NULL AND ParentId IS NULL)
                OR ParentId = :parent_id
          )
        ORDER BY SortOrder, Name, id
    """), {'company_code': company_code, 'parent_id': parent_id, 'section': section}).mappings().all()
    index = next((idx for idx, row in enumerate(siblings) if str(row['id']) == str(category_id)), None)
    if index is None:
        return 'Tax category sibling group could not be loaded.'
    target_index = index - 1 if direction == 'up' else index + 1
    if target_index < 0 or target_index >= len(siblings):
        return 'Tax category is already at the requested edge.'

    current = siblings[index]
    target = siblings[target_index]
    db.session.execute(text("""
        UPDATE tax_rollup_categories
           SET SortOrder = :sort_order,
               UpdatedAt = :updated_at
         WHERE id = :category_id
    """), {
        'sort_order': target['SortOrder'],
        'updated_at': datetime.datetime.utcnow(),
        'category_id': current['id'],
    })
    db.session.execute(text("""
        UPDATE tax_rollup_categories
           SET SortOrder = :sort_order,
               UpdatedAt = :updated_at
         WHERE id = :category_id
    """), {
        'sort_order': current['SortOrder'],
        'updated_at': datetime.datetime.utcnow(),
        'category_id': target['id'],
    })
    db.session.commit()
    return ''


def tax_rollup_company_options():
    companies = [
        row[0] for row in db.session.query(Accounts.Co)
        .filter(Accounts.Co.isnot(None))
        .distinct()
        .order_by(Accounts.Co)
        .all()
        if row[0]
    ]
    return companies or [cmpdata[10]]


def sync_tax_rollup_account_paths(company_code, old_path_by_id):
    if not old_path_by_id:
        return 0
    new_categories = tax_rollup_categories(company_code)
    new_path_by_id = {item['id']: item['path'] for item in new_categories}
    updated = 0
    for category_id, old_path in old_path_by_id.items():
        new_path = new_path_by_id.get(category_id)
        if not old_path or not new_path or old_path == new_path:
            continue
        rows = Accounts.query.filter(
            (Accounts.Co == company_code) &
            (Accounts.Taxrollup == old_path)
        ).all()
        for account in rows:
            account.Taxrollup = new_path
            updated += 1
    if updated:
        db.session.commit()
    return updated


@main.route('/TaxRollups', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def TaxRollups():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_tax_rollup_tables()
    companies = tax_rollup_company_options()
    selected_company = request.values.get('company') or cmpdata[10]
    if selected_company not in companies:
        selected_company = companies[0]
    err = []
    msg = []

    if request.method == 'POST':
        action = request.form.get('action')
        selected_company = request.form.get('company') or selected_company
        if selected_company not in companies:
            selected_company = companies[0]
        rollup_section = request.form.get('rollup_section') or 'Expense'
        if rollup_section not in ['Income', 'Expense']:
            rollup_section = 'Expense'

        if action == 'add_category':
            name = (request.form.get('category_name') or '').strip()
            parent_id = request.form.get('parent_id') or None
            if parent_id:
                parent_section = db.session.execute(text("""
                    SELECT Section
                    FROM tax_rollup_categories
                    WHERE id = :parent_id
                      AND Co = :co
                      AND Active = 1
                """), {'parent_id': parent_id, 'co': selected_company}).scalar()
                if parent_section != rollup_section:
                    parent_id = None
            sort_order = next_tax_rollup_sort(selected_company, parent_id, rollup_section)
            if not name:
                err.append('Tax category name is required.')
            else:
                db.session.execute(text("""
                    INSERT INTO tax_rollup_categories
                        (Co, Section, Name, ParentId, SortOrder, Active, CreatedAt, UpdatedAt)
                    VALUES
                        (:co, :section, :name, :parent_id, :sort_order, 1, :created_at, :updated_at)
                """), {
                    'co': selected_company,
                    'section': rollup_section,
                    'name': name,
                    'parent_id': parent_id,
                    'sort_order': sort_order,
                    'created_at': datetime.datetime.utcnow(),
                    'updated_at': datetime.datetime.utcnow(),
                })
                db.session.commit()
                msg.append(f'{rollup_section} tax category {name} added.')

        elif action == 'update_category':
            old_path_by_id = {
                item['id']: item['path']
                for item in tax_rollup_categories(selected_company, rollup_section)
            }
            category_id = request.form.get('category_id')
            name = (request.form.get('edit_category_name') or '').strip()
            parent_id = request.form.get('edit_parent_id') or None
            category_row = db.session.execute(text("""
                SELECT ParentId, SortOrder
                FROM tax_rollup_categories
                WHERE id = :category_id
                  AND Co = :co
                  AND Section = :section
                  AND Active = 1
            """), {'category_id': category_id, 'co': selected_company, 'section': rollup_section}).mappings().first()
            old_parent_id = str(category_row['ParentId']) if category_row and category_row['ParentId'] is not None else None
            new_parent_id = str(parent_id) if parent_id is not None else None
            if category_row is not None and old_parent_id == new_parent_id:
                sort_order = category_row['SortOrder'] or 0
            else:
                sort_order = next_tax_rollup_sort(selected_company, parent_id, rollup_section)
            if not category_id or not name:
                err.append('Choose a category and enter a name.')
            elif str(category_id) == str(parent_id):
                err.append('A category cannot be its own parent.')
            elif category_row is None:
                err.append('The selected category could not be found.')
            else:
                if parent_id:
                    parent_section = db.session.execute(text("""
                        SELECT Section
                        FROM tax_rollup_categories
                        WHERE id = :parent_id
                          AND Co = :co
                          AND Active = 1
                    """), {'parent_id': parent_id, 'co': selected_company}).scalar()
                    if parent_section != rollup_section:
                        parent_id = None
                        sort_order = next_tax_rollup_sort(selected_company, parent_id, rollup_section)
                db.session.execute(text("""
                    UPDATE tax_rollup_categories
                       SET Name = :name,
                           ParentId = :parent_id,
                           SortOrder = :sort_order,
                           UpdatedAt = :updated_at
                     WHERE id = :category_id
                       AND Co = :co
                       AND Section = :section
                       AND Active = 1
                """), {
                    'name': name,
                    'parent_id': parent_id,
                    'sort_order': sort_order,
                    'updated_at': datetime.datetime.utcnow(),
                    'category_id': category_id,
                    'co': selected_company,
                    'section': rollup_section,
                })
                db.session.commit()
                account_updates = sync_tax_rollup_account_paths(selected_company, old_path_by_id)
                msg.append(f'Tax category updated. {account_updates} account assignment(s) moved to the new rollup path.')

        elif action == 'delete_category':
            category_id = request.form.get('category_id')
            categories = tax_rollup_categories(selected_company, rollup_section)
            category = next((item for item in categories if str(item['id']) == str(category_id)), None)
            child_count = sum(1 for item in categories if str(item['parent_id']) == str(category_id))
            account_count = 0
            if category is not None:
                account_count = Accounts.query.filter(
                    (Accounts.Co == selected_company) &
                    (Accounts.Taxrollup == category['path'])
                ).count()
            if category is None:
                err.append('Choose a category to delete.')
            elif child_count:
                err.append('Move or delete child categories first.')
            elif account_count:
                err.append('Move expense accounts off this tax category before deleting it.')
            else:
                db.session.execute(text("""
                    UPDATE tax_rollup_categories
                       SET Active = 0,
                           UpdatedAt = :updated_at
                     WHERE id = :category_id
                       AND Co = :co
                       AND Section = :section
                """), {
                    'updated_at': datetime.datetime.utcnow(),
                    'category_id': category_id,
                    'co': selected_company,
                    'section': rollup_section,
                })
                db.session.commit()
                msg.append('Tax category deleted.')

        elif action == 'move_category':
            category_id = request.form.get('category_id')
            direction = request.form.get('direction')
            if direction not in ['up', 'down']:
                err.append('Choose a valid move direction.')
            else:
                move_error = move_tax_rollup_category(selected_company, category_id, direction, rollup_section)
                if move_error:
                    err.append(move_error)
                else:
                    msg.append('Tax category order updated.')

        elif action == 'save_assignments':
            accounts = Accounts.query.filter(
                (Accounts.Co == selected_company) &
                (Accounts.Type.in_(['Income', 'Expense']))
            ).all()
            updated = 0
            income_paths = {item['path'] for item in tax_rollup_categories(selected_company, 'Income')}
            expense_paths = {item['path'] for item in tax_rollup_categories(selected_company, 'Expense')}
            for account in accounts:
                value = (request.form.get(f'taxrollup_{account.id}') or '').strip()
                category_paths = income_paths if account.Type == 'Income' else expense_paths
                if value and value not in category_paths:
                    continue
                if (account.Taxrollup or '') != value:
                    account.Taxrollup = value or None
                    updated += 1
            db.session.commit()
            msg.append(f'Tax rollup assignments updated for {updated} income/expense account(s).')

    income_categories = tax_rollup_categories(selected_company, 'Income')
    expense_categories = tax_rollup_categories(selected_company, 'Expense')
    categories = income_categories + expense_categories
    income_accounts = Accounts.query.filter(
        (Accounts.Co == selected_company) &
        (Accounts.Type == 'Income')
    ).order_by(Accounts.Category, Accounts.Subcategory, Accounts.Name).all()
    expense_accounts = Accounts.query.filter(
        (Accounts.Co == selected_company) &
        (Accounts.Type == 'Expense')
    ).order_by(Accounts.Category, Accounts.Subcategory, Accounts.Name).all()
    return render_template(
        'tax_rollups.html',
        cmpdata=cmpdata,
        scac=scac,
        companies=companies,
        selected_company=selected_company,
        categories=categories,
        income_categories=income_categories,
        expense_categories=expense_categories,
        income_category_paths={category['path'] for category in income_categories},
        expense_category_paths={category['path'] for category in expense_categories},
        category_paths={category['path'] for category in categories},
        income_accounts=income_accounts,
        expense_accounts=expense_accounts,
        err=err,
        msg=msg,
    )


def business_report_decimal(value):
    if value in [None, '']:
        return 0.0
    try:
        cleaned = str(value).replace(',', '').strip()
        return float(cleaned or 0)
    except (TypeError, ValueError):
        return 0.0


def business_report_period_bounds(start_date, end_date):
    return (
        business_report_as_datetime(start_date),
        business_report_as_datetime(end_date, end_of_day=True),
    )


def business_report_account_lookup(company_code):
    accounts = Accounts.query.filter(Accounts.Co == company_code).all()
    by_id = {account.id: account for account in accounts}
    by_name = {}
    for account in accounts:
        if account.Name and account.Name not in by_name:
            by_name[account.Name] = account
    return by_id, by_name


def business_report_account_for_line(line, account_by_id, account_by_name, bill=None):
    # The Bill row owns the expense account, but only the expense side of a
    # bill-payment journal should inherit it. The offsetting bank/payment line
    # must keep its own account or it will show as a duplicate expense.
    if (
        bill is not None
        and (line.Type or '') in ['ED', 'XD', 'AD']
        and (bill.bAccount or '').strip()
    ):
        bill_account = account_by_name.get(bill.bAccount or '')
        if bill_account is not None:
            return bill_account
    # Prefer Aid because account names may have been changed after the ledger
    # line was posted. A join on id-or-name can multiply one ledger line when
    # an old name and a new name both match account records.
    if line.Aid and line.Aid in account_by_id:
        return account_by_id[line.Aid]
    return account_by_name.get(line.Account or '')


def business_report_basis_qty(basis_type, start_date, end_date):
    if not start_date or not end_date:
        return 0.0
    if end_date < start_date:
        return 0.0
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    if basis_type == 'day':
        return float((end_date - start_date).days + 1)
    if basis_type == 'mile':
        assignments = DriverAssign.query.filter(
            DriverAssign.Date >= start_dt,
            DriverAssign.Date <= end_dt,
        ).all()
        return sum(business_report_decimal(row.Miles) for row in assignments)
    # Port entries are distinct Interchange.PortTrip sequence numbers in the period.
    return float(db.session.query(func.count(func.distinct(Interchange.PortTrip))).filter(
        Interchange.Date >= start_dt,
        Interchange.Date <= end_dt,
        Interchange.PortTrip.isnot(None),
        Interchange.PortTrip > 0,
    ).scalar() or 0)


def business_report_expenses(start_date, end_date, company_code=None):
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    query = Gledger.query.filter(
        Gledger.Date >= start_dt,
        Gledger.Date <= end_dt,
    )
    if company_code:
        query = query.filter(Gledger.Com == company_code)
        account_by_id, account_by_name = business_report_account_lookup(company_code)
    else:
        accounts = Accounts.query.all()
        account_by_id = {account.id: account for account in accounts}
        account_by_name = {}
        for account in accounts:
            if account.Co and account.Name and (account.Co, account.Name) not in account_by_name:
                account_by_name[(account.Co, account.Name)] = account

    totals = {}
    for line in query.all():
        if company_code:
            account_obj = business_report_account_for_line(line, account_by_id, account_by_name)
        else:
            account_obj = account_by_id.get(line.Aid) if line.Aid else account_by_name.get((line.Com, line.Account or ''))
        if account_obj is None or account_obj.Type != 'Expense':
            continue
        key = (
            account_obj.Name or line.Account or '',
            account_obj.Category or '',
            account_obj.Subcategory or '',
        )
        totals[key] = totals.get(key, 0) + int(line.Debit or 0) - int(line.Credit or 0)

    output = []
    for (account, category, subcategory), amount_cents in sorted(totals.items(), key=lambda item: (item[0][1], item[0][0])):
        amount_cents = int(amount_cents or 0)
        output.append({
            'account': account or '',
            'category': category or '',
            'subcategory': subcategory or '',
            'amount_cents': amount_cents,
            'amount': business_report_money(amount_cents),
        })
    return output


def business_report_rollup_rows(rows, configured_categories):
    totals = {}
    detail_accounts = {}
    for row in rows:
        path = row['taxrollup'] or 'Unassigned'
        parts = [part.strip() for part in path.split(':') if part.strip()]
        if not parts:
            parts = ['Unassigned']
        for index in range(1, len(parts) + 1):
            prefix = ':'.join(parts[:index])
            totals[prefix] = totals.get(prefix, 0) + row['amount_cents']
        detail_accounts.setdefault(path, []).append(row)

    def contained_accounts(path):
        if path == 'Unassigned':
            return detail_accounts.get(path, [])
        contained = []
        for account_path, accounts in detail_accounts.items():
            if account_path == path or account_path.startswith(f'{path}:'):
                contained.extend(accounts)
        return sorted(contained, key=lambda item: (item['taxrollup'], item['account']))

    output = []
    seen = set()
    for category in configured_categories:
        path = category['path']
        if path in totals:
            output.append({
                'path': path,
                'label': category['name'],
                'level': category['level'],
                'amount_cents': totals[path],
                'amount': business_report_money(totals[path]),
                'accounts': contained_accounts(path),
            })
            seen.add(path)

    for path in sorted(set(totals) - seen):
        level = 0 if path == 'Unassigned' else path.count(':')
        output.append({
            'path': path,
            'label': path.split(':')[-1],
            'level': level,
            'amount_cents': totals[path],
            'amount': business_report_money(totals[path]),
            'accounts': contained_accounts(path),
        })

    return output


def business_report_bill_for_ledger_line(line):
    if line.SourceTable == 'Bills' and line.SourceId:
        bill = Bills.query.get(line.SourceId)
        if bill is not None:
            return bill
    if line.Tcode:
        bill_jo = line.Tcode
        if (line.JournalId or '').startswith('PAYBILL-') and '-' in bill_jo:
            bill_jo = bill_jo.rsplit('-', 1)[0]
        bill = Bills.query.filter(Bills.Jo == bill_jo).first()
        if bill is not None:
            return bill
    return None


def business_report_final_reconciled_value(value):
    # 0/None are unreconciled; 25 is a trial-selected line. Finalized statements
    # replace 25 with the statement month number, so only values outside this
    # open set are treated as reconciled.
    return value not in [None, 0, 25]


def business_report_payment_reconciled_for_bill(bill, fallback_line=None):
    if bill is None:
        return business_report_final_reconciled_value(fallback_line.Reconciled) if fallback_line else False

    payment_filters = [
        (Gledger.SourceTable == 'Bills') & (Gledger.SourceId == bill.id),
    ]
    if bill.Jo:
        payment_filters.extend([
            Gledger.JournalId == f'PAYBILL-{bill.Jo}',
            Gledger.JournalId.like(f'PAYBILL-{bill.Jo}-%'),
            Gledger.Tcode == bill.Jo,
            Gledger.Tcode.like(f'{bill.Jo}-%'),
        ])
    payment_lines = Gledger.query.filter(or_(*payment_filters)).all()

    bank_payment_lines = [
        line for line in payment_lines
        if line.Type in ['PC', 'DD', 'QC', 'XC'] or (line.Account or '') == (bill.pAccount or '')
    ]
    if bank_payment_lines:
        return any(business_report_final_reconciled_value(line.Reconciled) for line in bank_payment_lines)
    return any(business_report_final_reconciled_value(line.Reconciled) for line in payment_lines)


def business_report_payroll_reconciled_for_line(line):
    if line is None or line.SourceTable != 'PayrollBatch':
        return None
    payroll_lines = Gledger.query.filter(
        (Gledger.SourceTable == 'PayrollBatch') &
        (Gledger.JournalId == line.JournalId) &
        (Gledger.Com == line.Com)
    ).all()
    bank_lines = [row for row in payroll_lines if row.Type in ['PC', 'XC']]
    if not bank_lines:
        return any(business_report_final_reconciled_value(row.Reconciled) for row in payroll_lines)
    return any(business_report_final_reconciled_value(row.Reconciled) for row in bank_lines)


def business_report_tax_rollup_detail_lines(company_code, start_date, end_date, account_type):
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    account_by_id, account_by_name = business_report_account_lookup(company_code)
    ledger_rows = Gledger.query.filter(
        Gledger.Com == company_code,
        Gledger.Date >= start_dt,
        Gledger.Date <= end_dt,
    ).order_by(
        Gledger.Date.asc(),
        Gledger.Account.asc(),
        Gledger.id.asc(),
    ).all()

    by_path = {}
    seen_line_paths = set()
    for line in ledger_rows:
        bill = business_report_bill_for_ledger_line(line)
        account = business_report_account_for_line(line, account_by_id, account_by_name, bill)
        if (
            account is None
            or account.Type != account_type
            or not (account.Taxrollup or '').strip()
        ):
            continue
        if account_type == 'Income':
            amount_cents = int(line.Credit or 0) - int(line.Debit or 0)
        else:
            amount_cents = int(line.Debit or 0) - int(line.Credit or 0)
        payroll_reconciled = business_report_payroll_reconciled_for_line(line)
        if payroll_reconciled is not None:
            reconciled = payroll_reconciled
        elif account_type == 'Expense':
            reconciled = business_report_payment_reconciled_for_bill(bill, line)
        else:
            reconciled = business_report_final_reconciled_value(line.Reconciled)
        detail = {
            'date': business_report_date_text(line.Date),
            'account': account.Name or line.Account or '',
            'taxrollup': account.Taxrollup or '',
            'category': account.Category or '',
            'subcategory': account.Subcategory or '',
            'amount_cents': amount_cents,
            'amount': business_report_money(amount_cents),
            'ledger_type': line.Type or '',
            'reconciled': reconciled,
            'journal_id': line.JournalId or '',
            'source_table': line.SourceTable or '',
            'source_id': line.SourceId or '',
            'tcode': line.Tcode or '',
            'ref': line.Ref or '',
            'source': line.Source or '',
            'bill_jo': bill.Jo if bill else '',
            'bill_vendor': bill.Company if bill else '',
            'bill_description': bill.Description if bill else '',
            'bill_amount': bill.bAmount if bill else '',
            'bill_date': business_report_date_text(bill.Date) if bill else '',
            'paid_date': business_report_date_text(bill.pDate) if bill else '',
            'paid_amount': bill.pAmount if bill else '',
            'paid_account': bill.pAccount if bill else '',
            'bill_status': bill.Status if bill else '',
        }
        parts = [part.strip() for part in (account.Taxrollup or '').split(':') if part.strip()]
        for index in range(1, len(parts) + 1):
            prefix = ':'.join(parts[:index])
            line_path_key = (line.id, prefix)
            if line_path_key in seen_line_paths:
                continue
            seen_line_paths.add(line_path_key)
            by_path.setdefault(prefix, []).append(detail)
    return by_path


def business_report_rollup_section_report_rows(rollup_rows):
    output = []
    active_top = None

    def append_top_total():
        if active_top is None:
            return
        output.append({
            'path': active_top['path'],
            'label': f"Total {active_top['label']}",
            'level': 1,
            'amount_cents': None,
            'total_cents': active_top['amount_cents'],
            'amount': '',
            'total': business_report_money(active_top['amount_cents']),
            'accounts': active_top.get('accounts', []),
            'details': active_top.get('details', []),
            'kind': 'top_total',
        })

    for row in rollup_rows:
        if row['level'] == 0:
            append_top_total()
            active_top = row
            display_row = dict(row)
            display_row.update({
                'amount_cents': None,
                'total_cents': None,
                'amount': '',
                'total': '',
                'accounts': row.get('accounts', []),
                'details': row.get('details', []),
                'kind': 'top',
            })
            output.append(display_row)
        else:
            display_row = dict(row)
            display_row.update({
                'total_cents': None,
                'total': '',
                'accounts': row.get('accounts', []),
                'details': row.get('details', []),
                'kind': 'rollup',
            })
            output.append(display_row)

    append_top_total()
    return output


def business_report_tax_rollup_pl(company_code, start_date, end_date):
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    account_by_id, account_by_name = business_report_account_lookup(company_code)
    ledger_rows = Gledger.query.filter(
        Gledger.Com == company_code,
        Gledger.Date >= start_dt,
        Gledger.Date <= end_dt,
    ).all()

    totals = {}
    for line in ledger_rows:
        bill = business_report_bill_for_ledger_line(line)
        account = business_report_account_for_line(line, account_by_id, account_by_name, bill)
        if account is None or account.Type not in ['Income', 'Expense']:
            continue
        key = (
            account.Type,
            account.Taxrollup or '',
            account.Name or line.Account or '',
            account.Category or '',
            account.Subcategory or '',
        )
        debit_cents, credit_cents = totals.get(key, (0, 0))
        totals[key] = (
            debit_cents + int(line.Debit or 0),
            credit_cents + int(line.Credit or 0),
        )

    income_accounts = []
    expense_accounts = []
    unassigned_expense_accounts = []
    for (account_type, taxrollup, name, category, subcategory), values in sorted(totals.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
        debit_cents, credit_cents = values
        if account_type == 'Income':
            amount_cents = credit_cents - debit_cents
            target = income_accounts
        else:
            amount_cents = debit_cents - credit_cents
            target = expense_accounts if (taxrollup or '').strip() else unassigned_expense_accounts
        target.append({
            'account': name or '',
            'category': category or '',
            'subcategory': subcategory or '',
            'taxrollup': taxrollup or '',
            'amount_cents': amount_cents,
            'amount': business_report_money(amount_cents),
        })

    income_categories = tax_rollup_categories(company_code, 'Income')
    expense_categories = tax_rollup_categories(company_code, 'Expense')
    income_rows = business_report_rollup_rows(income_accounts, income_categories)
    expense_rows = business_report_rollup_rows(expense_accounts, expense_categories)
    income_detail_by_path = business_report_tax_rollup_detail_lines(company_code, start_date, end_date, 'Income')
    expense_detail_by_path = business_report_tax_rollup_detail_lines(company_code, start_date, end_date, 'Expense')
    for row in income_rows:
        row['details'] = income_detail_by_path.get(row['path'], [])
    for row in expense_rows:
        row['details'] = expense_detail_by_path.get(row['path'], [])
    total_income = sum(row['amount_cents'] for row in income_accounts)
    total_expenses = sum(row['amount_cents'] for row in expense_accounts)
    unassigned_expenses = sum(row['amount_cents'] for row in unassigned_expense_accounts)
    net_income = total_income - total_expenses

    return {
        'income_rows': income_rows,
        'expense_rows': expense_rows,
        'income_report_rows': business_report_rollup_section_report_rows(income_rows),
        'expense_report_rows': business_report_rollup_section_report_rows(expense_rows),
        'total_income_cents': total_income,
        'total_expenses_cents': total_expenses,
        'unassigned_expenses_cents': unassigned_expenses,
        'net_income_cents': net_income,
        'total_income': business_report_money(total_income),
        'total_expenses': business_report_money(total_expenses),
        'unassigned_expenses': business_report_money(unassigned_expenses),
        'net_income': business_report_money(net_income),
    }


def business_report_tax_rollup_export_rows(tax_rollup_pl):
    rows = [{'label': 'Income', 'level': 0, 'amount_cents': None, 'kind': 'section'}]
    for row in tax_rollup_pl['income_report_rows']:
        rows.append({
            'label': row['label'],
            'path': row['path'],
            'level': row['level'] + 1,
            'amount_cents': row['amount_cents'],
            'total_cents': row.get('total_cents'),
            'kind': row.get('kind', 'rollup'),
        })
    rows.append({
        'label': 'Total Income',
        'level': 1,
        'amount_cents': None,
        'total_cents': tax_rollup_pl['total_income_cents'],
        'kind': 'total',
    })
    rows.append({'label': '', 'level': 0, 'amount_cents': None, 'kind': 'blank'})
    rows.append({'label': 'Expenses', 'level': 0, 'amount_cents': None, 'kind': 'section'})
    for row in tax_rollup_pl['expense_report_rows']:
        rows.append({
            'label': row['label'],
            'path': row['path'],
            'level': row['level'] + 1,
            'amount_cents': row['amount_cents'],
            'total_cents': row.get('total_cents'),
            'kind': row.get('kind', 'rollup'),
        })
    rows.append({
        'label': 'Total Expenses',
        'level': 1,
        'amount_cents': None,
        'total_cents': tax_rollup_pl['total_expenses_cents'],
        'kind': 'total',
    })
    rows.append({'label': '', 'level': 0, 'amount_cents': None, 'kind': 'blank'})
    rows.append({
        'label': 'Net Income',
        'level': 0,
        'amount_cents': None,
        'total_cents': tax_rollup_pl['net_income_cents'],
        'kind': 'net',
    })
    return rows


def business_report_tax_rollup_xlsx(company_code, start_date, end_date, tax_rollup_pl):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tax Rollup PL'

    ws['A1'] = 'Profit & Loss - Tax Rollup'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Company: {company_code}'
    ws['A3'] = f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
    ws.append([])
    ws.append(['Rollup', 'Amount', 'Total'])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')

    for row in business_report_tax_rollup_export_rows(tax_rollup_pl):
        amount = row['amount_cents'] / 100 if row['amount_cents'] is not None else None
        total = row.get('total_cents') / 100 if row.get('total_cents') is not None else None
        ws.append([row['label'], amount, total])
        excel_row = ws.max_row
        ws.cell(excel_row, 1).alignment = Alignment(indent=row['level'])
        ws.row_dimensions[excel_row].outlineLevel = min(row['level'], 7)
        if row['amount_cents'] is not None:
            ws.cell(excel_row, 2).number_format = '$#,##0.00;[Red]($#,##0.00)'
            ws.cell(excel_row, 2).alignment = Alignment(horizontal='right')
        if row.get('total_cents') is not None:
            ws.cell(excel_row, 3).number_format = '$#,##0.00;[Red]($#,##0.00)'
            ws.cell(excel_row, 3).alignment = Alignment(horizontal='right')
        if row['kind'] in ['section', 'top_total', 'total', 'net']:
            ws.cell(excel_row, 1).font = Font(bold=True)
            ws.cell(excel_row, 2).font = Font(bold=True)
            ws.cell(excel_row, 3).font = Font(bold=True)
        if row['kind'] == 'section':
            ws.cell(excel_row, 1).fill = PatternFill('solid', fgColor='E2F0D9')
            ws.cell(excel_row, 2).fill = PatternFill('solid', fgColor='E2F0D9')
            ws.cell(excel_row, 3).fill = PatternFill('solid', fgColor='E2F0D9')
        if row['kind'] == 'net':
            ws.cell(excel_row, 1).fill = PatternFill('solid', fgColor='FFF2CC')
            ws.cell(excel_row, 2).fill = PatternFill('solid', fgColor='FFF2CC')
            ws.cell(excel_row, 3).fill = PatternFill('solid', fgColor='FFF2CC')

    ws.column_dimensions['A'].width = 58
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    wb.save(output)
    output.seek(0)
    filename = f'tax_rollup_profit_loss_{company_code}_{start_date}_{end_date}.xlsx'.replace('/', '-')
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def business_report_tax_rollup_pdf(company_code, start_date, end_date, tax_rollup_pl):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Profit & Loss - Tax Rollup', styles['Title']),
        Paragraph(f'Company: {company_code} &nbsp;&nbsp; Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}', styles['Normal']),
        Spacer(1, 12),
    ]

    table_data = [['Rollup', 'Amount', 'Total']]
    row_styles = []
    for row in business_report_tax_rollup_export_rows(tax_rollup_pl):
        table_data.append([
            row['label'],
            business_report_money(row['amount_cents']) if row['amount_cents'] is not None else '',
            business_report_money(row.get('total_cents')) if row.get('total_cents') is not None else '',
        ])
        idx = len(table_data) - 1
        row_styles.append(('LEFTPADDING', (0, idx), (0, idx), 6 + (row['level'] * 14)))
        if row['kind'] in ['section', 'top_total', 'total', 'net']:
            row_styles.append(('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'))
        if row['kind'] == 'section':
            row_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#E2F0D9')))
        if row['kind'] == 'net':
            row_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFF2CC')))

    table = Table(table_data, colWidths=[330, 85, 85], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9EAF7')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), .25, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ] + row_styles))
    story.append(table)
    doc.build(story)
    output.seek(0)
    filename = f'tax_rollup_profit_loss_{company_code}_{start_date}_{end_date}.pdf'.replace('/', '-')
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')


def business_report_unassigned_expense_lines(company_code, start_date, end_date):
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    rows = db.session.query(
        Gledger.Date,
        Gledger.Account,
        Gledger.Tcode,
        Gledger.Ref,
        Gledger.Source,
        Gledger.Debit,
        Gledger.Credit,
        Accounts.Category,
        Accounts.Subcategory,
    ).join(
        Accounts,
        or_(Gledger.Aid == Accounts.id, (Gledger.Account == Accounts.Name) & (Gledger.Com == Accounts.Co)),
    ).filter(
        Accounts.Co == company_code,
        Accounts.Type == 'Expense',
        or_(Accounts.Taxrollup.is_(None), Accounts.Taxrollup == ''),
        Gledger.Date >= start_dt,
        Gledger.Date <= end_dt,
    ).order_by(
        Gledger.Date.asc(),
        Gledger.Account.asc(),
        Gledger.id.asc(),
    ).all()

    output = []
    for row in rows:
        amount_cents = int(row.Debit or 0) - int(row.Credit or 0)
        output.append({
            'date': business_report_date_text(row.Date),
            'account': row.Account or '',
            'category': row.Category or '',
            'subcategory': row.Subcategory or '',
            'tcode': row.Tcode or '',
            'ref': row.Ref or '',
            'source': row.Source or '',
            'amount_cents': amount_cents,
            'amount': business_report_money(amount_cents),
        })
    return output


def business_report_allocations():
    rows = db.session.execute(text("""
        SELECT id, AccountName, AllocationMethod, BasisType, BasisLabel, PeriodStart, PeriodEnd, BasisQty, Notes, Active
        FROM business_report_allocations
        WHERE Active = 1
        ORDER BY AccountName, AllocationMethod, id
    """)).mappings().all()
    output = []
    for row in rows:
        output.append({
            'id': row.get('id'),
            'account': row.get('AccountName') or '',
            'method': row.get('AllocationMethod') or '',
            'basis_type': row.get('BasisType') or 'port_entry',
            'basis': row.get('BasisLabel') or BUSINESS_REPORT_BASIS_TYPES.get(row.get('BasisType'), 'Port Entries'),
            'period_start': business_report_date_text(row.get('PeriodStart')),
            'period_end': business_report_date_text(row.get('PeriodEnd')),
            'basis_qty': float(row.get('BasisQty') or 0),
            'notes': row.get('Notes') or '',
        })
    return output


def business_report_allocation_calcs(allocations, report_start, report_end):
    output = []
    for item in allocations:
        period_start = business_report_parse_date(item['period_start'], None)
        period_end = business_report_parse_date(item['period_end'], None)
        expense_cents = 0
        if period_start and period_end:
            expenses = business_report_expenses(period_start, period_end)
            expense_cents = next((row['amount_cents'] for row in expenses if row['account'] == item['account']), 0)
        auto_basis_qty = business_report_basis_qty(item['basis_type'], period_start, period_end)
        basis_qty = item['basis_qty'] or auto_basis_qty
        report_basis_qty = business_report_basis_qty(item['basis_type'], report_start, report_end)
        cost_per_basis = (expense_cents / 100) / basis_qty if basis_qty else 0
        allocated_amount = cost_per_basis * report_basis_qty
        calc = dict(item)
        calc.update({
            'basis_name': BUSINESS_REPORT_BASIS_TYPES.get(item['basis_type'], item['basis_type']),
            'auto_basis_qty': auto_basis_qty,
            'source_basis_used': basis_qty,
            'report_basis_qty': report_basis_qty,
            'expense_amount': business_report_money(expense_cents),
            'cost_per_basis': "${:,.4f}".format(cost_per_basis),
            'allocated_amount': "${:,.2f}".format(allocated_amount),
        })
        output.append(calc)
    return output


def operations_indirect_allocations(company_code):
    rows = db.session.execute(text("""
        SELECT id, Co, AccountName, PeriodStart, PeriodEnd, AmountCents, AllocationBasis, AutoUpdate, Notes
        FROM operations_indirect_allocations
        WHERE Co = :company_code AND Active = 1
        ORDER BY PeriodStart DESC, AccountName, id
    """), {'company_code': company_code}).mappings().all()
    output = []
    for row in rows:
        output.append({
            'id': row['id'],
            'account': row['AccountName'] or '',
            'period_start': business_report_date_text(row['PeriodStart']),
            'period_end': business_report_date_text(row['PeriodEnd']),
            'amount_cents': int(row['AmountCents'] or 0),
            'amount': business_report_money(int(row['AmountCents'] or 0)),
            'allocation_basis': row['AllocationBasis'] or 'daily',
            'allocation_basis_label': OPERATIONS_ALLOCATION_BASIS_TYPES.get(row['AllocationBasis'] or 'daily', 'Daily Spread'),
            'auto_update': bool(row['AutoUpdate']),
            'notes': row['Notes'] or '',
        })
    return output


def operations_ga_allocations(company_code):
    rows = db.session.execute(text("""
        SELECT id, Co, AccountName, PeriodStart, PeriodEnd, AmountCents, AllocationBasis, AutoUpdate, Notes
        FROM operations_ga_allocations
        WHERE Co = :company_code AND Active = 1
        ORDER BY PeriodStart DESC, AccountName, id
    """), {'company_code': company_code}).mappings().all()
    output = []
    for row in rows:
        output.append({
            'id': row['id'],
            'account': row['AccountName'] or '',
            'period_start': business_report_date_text(row['PeriodStart']),
            'period_end': business_report_date_text(row['PeriodEnd']),
            'amount_cents': int(row['AmountCents'] or 0),
            'amount': business_report_money(int(row['AmountCents'] or 0)),
            'allocation_basis': row['AllocationBasis'] or 'daily',
            'allocation_basis_label': OPERATIONS_ALLOCATION_BASIS_TYPES.get(row['AllocationBasis'] or 'daily', 'Daily Spread'),
            'auto_update': bool(row['AutoUpdate']),
            'notes': row['Notes'] or '',
        })
    return output


def operations_account_ytd_cents(company_code, account_name, as_of_date, account_category):
    year_start = datetime.date(as_of_date.year, 1, 1)
    start_dt, end_dt = business_report_period_bounds(year_start, as_of_date)
    amount = db.session.query(
        func.sum(func.coalesce(Gledger.Debit, 0) - func.coalesce(Gledger.Credit, 0))
    ).join(
        Accounts,
        or_(Gledger.Aid == Accounts.id, (Gledger.Account == Accounts.Name) & (Gledger.Com == Accounts.Co)),
    ).filter(
        Accounts.Co == company_code,
        Accounts.Type == 'Expense',
        func.lower(Accounts.Category) == account_category,
        Accounts.Name == account_name,
        Gledger.Date >= start_dt,
        Gledger.Date <= end_dt,
    ).scalar()
    return int(amount or 0)


def operations_sync_auto_allocations(company_code, as_of_date, table_name, account_category):
    year_start = datetime.date(as_of_date.year, 1, 1)
    if table_name not in ['operations_indirect_allocations', 'operations_ga_allocations']:
        return
    rows = db.session.execute(text("""
        SELECT id, AccountName
        FROM """ + table_name + """
        WHERE Co = :company_code AND Active = 1 AND AutoUpdate = 1
    """), {'company_code': company_code}).mappings().all()
    for row in rows:
        amount_cents = operations_account_ytd_cents(company_code, row['AccountName'], as_of_date, account_category)
        db.session.execute(text("""
            UPDATE """ + table_name + """
            SET PeriodStart = :period_start,
                PeriodEnd = :period_end,
                AmountCents = :amount_cents,
                UpdatedAt = :updated_at
            WHERE id = :allocation_id AND Co = :company_code
        """), {
            'period_start': year_start,
            'period_end': as_of_date,
            'amount_cents': amount_cents,
            'updated_at': datetime.datetime.utcnow(),
            'allocation_id': row['id'],
            'company_code': company_code,
        })
    if rows:
        db.session.commit()


def operations_account_options(company_code, as_of_date, account_category):
    year_start = datetime.date(as_of_date.year, 1, 1)
    start_dt, end_dt = business_report_period_bounds(year_start, as_of_date)
    rows = db.session.query(
        Accounts.Name,
        Accounts.Category,
        Accounts.Subcategory,
        func.sum(func.coalesce(Gledger.Debit, 0) - func.coalesce(Gledger.Credit, 0)).label('amount_cents'),
    ).outerjoin(
        Gledger,
        (Gledger.Aid == Accounts.id) &
        (Gledger.Com == Accounts.Co) &
        (Gledger.Date >= start_dt) &
        (Gledger.Date <= end_dt),
    ).filter(
        Accounts.Co == company_code,
        Accounts.Type == 'Expense',
        func.lower(Accounts.Category) == account_category,
    ).group_by(
        Accounts.id,
        Accounts.Name,
        Accounts.Category,
        Accounts.Subcategory,
    ).order_by(Accounts.Name).all()

    output = []
    for row in rows:
        amount_cents = int(row.amount_cents or 0)
        output.append({
            'name': row.Name or '',
            'category': row.Category or '',
            'subcategory': row.Subcategory or '',
            'default_start': year_start.strftime('%Y-%m-%d'),
            'default_end': as_of_date.strftime('%Y-%m-%d'),
            'default_amount': f'{amount_cents / 100:.2f}',
            'default_amount_display': business_report_money(amount_cents),
        })
    return output


def operations_allocation_source_row(company_code, allocation_type, allocation_id):
    table_name = 'operations_indirect_allocations' if allocation_type == 'indirect' else 'operations_ga_allocations'
    if allocation_type not in ['indirect', 'ga']:
        return None
    return db.session.execute(text("""
        SELECT id, Co, AccountName, PeriodStart, PeriodEnd, AmountCents, AllocationBasis, AutoUpdate, Notes
        FROM """ + table_name + """
        WHERE Co = :company_code
          AND id = :allocation_id
          AND Active = 1
        LIMIT 1
    """), {
        'company_code': company_code,
        'allocation_id': allocation_id,
    }).mappings().first()


def operations_weekly_allocation_plot(company_code, allocation_type, allocation_id, start_date, end_date):
    if allocation_type in ['builtin_direct', 'builtin_driver_pay']:
        daily_totals = operations_direct_daily_allocations(
            company_code,
            start_date,
            end_date,
            'driver_pay' if allocation_type == 'builtin_driver_pay' else None,
        )
        trip_counts = business_report_port_trip_counts_by_day(start_date, end_date)
        label = 'Driver Pay' if allocation_type == 'builtin_driver_pay' else 'Direct Expenses'
        return operations_weekly_plot_payload(
            allocation_type=allocation_type,
            allocation_id=allocation_id,
            account=label,
            allocation_basis='weekly_direct',
            allocation_basis_label='Weekly Direct Allocation',
            period_start=start_date,
            period_end=end_date,
            amount_cents=sum(daily_totals.values()),
            daily_totals=daily_totals,
            trip_counts=trip_counts,
            start_date=start_date,
            end_date=end_date,
        )

    row = operations_allocation_source_row(company_code, allocation_type, allocation_id)
    if row is None:
        return None
    period_start = business_report_as_date(row['PeriodStart'])
    period_end = business_report_as_date(row['PeriodEnd'])
    daily_totals = {}
    allocation_basis = row['AllocationBasis'] or 'daily'
    if allocation_basis == 'port_trip':
        business_report_allocate_cents_by_port_trip(
            daily_totals,
            period_start,
            period_end,
            int(row['AmountCents'] or 0),
            start_date,
            end_date,
        )
    else:
        business_report_allocate_cents_by_period(
            daily_totals,
            period_start,
            period_end,
            int(row['AmountCents'] or 0),
            start_date,
            end_date,
        )

    trip_counts = business_report_port_trip_counts_by_day(start_date, end_date)
    return operations_weekly_plot_payload(
        allocation_type=allocation_type,
        allocation_id=row['id'],
        account=row['AccountName'] or '',
        allocation_basis=allocation_basis,
        allocation_basis_label=OPERATIONS_ALLOCATION_BASIS_TYPES.get(allocation_basis, 'Daily Spread'),
        period_start=business_report_as_date(row['PeriodStart']),
        period_end=business_report_as_date(row['PeriodEnd']),
        amount_cents=int(row['AmountCents'] or 0),
        daily_totals=daily_totals,
        trip_counts=trip_counts,
        start_date=start_date,
        end_date=end_date,
    )


def operations_weekly_plot_payload(
    allocation_type,
    allocation_id,
    account,
    allocation_basis,
    allocation_basis_label,
    period_start,
    period_end,
    amount_cents,
    daily_totals,
    trip_counts,
    start_date,
    end_date,
):
    weeks = {}
    for day in business_report_each_day(start_date, end_date):
        week_start = business_report_week_start(day)
        week_end = week_start + datetime.timedelta(days=6)
        week = weeks.setdefault(week_start, {
            'week_start': max(week_start, start_date),
            'week_end': min(week_end, end_date),
            'amount_cents': 0,
            'port_trips': 0,
        })
        week['amount_cents'] += int(daily_totals.get(day, 0) or 0)
        week['port_trips'] += int(trip_counts.get(day, 0) or 0)

    points = []
    for week_start in sorted(weeks):
        week = weeks[week_start]
        points.append({
            'week_start': week['week_start'].strftime('%Y-%m-%d'),
            'week_end': week['week_end'].strftime('%Y-%m-%d'),
            'label': f"{week['week_start'].strftime('%m/%d')} - {week['week_end'].strftime('%m/%d')}",
            'amount_cents': week['amount_cents'],
            'amount': business_report_money(week['amount_cents']),
            'amount_value': round(week['amount_cents'] / 100, 2),
            'port_trips': week['port_trips'],
        })

    return {
        'allocation_type': allocation_type,
        'id': allocation_id,
        'account': account or '',
        'period_start': business_report_date_text(period_start),
        'period_end': business_report_date_text(period_end),
        'amount_cents': int(amount_cents or 0),
        'amount': business_report_money(int(amount_cents or 0)),
        'allocation_basis': allocation_basis,
        'allocation_basis_label': allocation_basis_label,
        'points': points,
    }


def operations_direct_account_matches_filter(account, account_filter=None):
    if account_filter != 'driver_pay':
        return True
    name = (account.Name or '').lower()
    taxrollup = (account.Taxrollup or '').lower()
    return (
        'driver labor' in name
        or 'subcontracted trucking' in name
        or 'contracted drivers' in taxrollup
        or 'salaries and wages' in taxrollup
    )


def operations_direct_daily_allocations(company_code, start_date, end_date, account_filter=None):
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    account_by_id, account_by_name = business_report_account_lookup(company_code)
    ledger_rows = Gledger.query.filter(
        Gledger.Com == company_code,
        Gledger.Date >= start_dt,
        Gledger.Date <= end_dt,
    ).order_by(
        Gledger.Date.asc(),
        Gledger.id.asc(),
    ).all()
    totals = {}
    for line in ledger_rows:
        bill = business_report_bill_for_ledger_line(line)
        account = business_report_account_for_line(line, account_by_id, account_by_name, bill)
        if (
            account is None
            or account.Type != 'Expense'
            or (account.Category or '').lower() != 'direct'
            or not operations_direct_account_matches_filter(account, account_filter)
        ):
            continue
        expense_date = business_report_as_date(line.Date)
        if expense_date is None:
            continue
        key = (account.Name or line.Account or '', expense_date)
        totals[key] = totals.get(key, 0) + int(line.Debit or 0) - int(line.Credit or 0)

    direct_daily = {}
    previous_date_by_account = {}
    for (account_name, expense_date), amount_cents in sorted(totals.items(), key=lambda item: (item[0][0], item[0][1])):
        previous_date = previous_date_by_account.get(account_name)
        allocation_start = previous_date + datetime.timedelta(days=1) if previous_date else start_date
        allocation_start = max(allocation_start, start_date)
        allocation_end = min(expense_date, end_date)
        business_report_allocate_cents_by_day(direct_daily, allocation_start, allocation_end, int(amount_cents or 0))
        previous_date_by_account[account_name] = expense_date
    return direct_daily


def operations_weekly_profit_loss(company_code, start_date, end_date):
    start_dt, end_dt = business_report_period_bounds(start_date, end_date)
    daily_income = {}
    income_rows = db.session.query(
        Gledger.Date,
        Gledger.Debit,
        Gledger.Credit,
        Orders.Date2.label('gate_in_date'),
    ).join(
        Accounts,
        or_(Gledger.Aid == Accounts.id, (Gledger.Account == Accounts.Name) & (Gledger.Com == Accounts.Co)),
    ).outerjoin(
        Orders,
        Gledger.Tcode == Orders.Jo,
    ).filter(
        Accounts.Co == company_code,
        Accounts.Type == 'Income',
        or_(
            (Gledger.Date >= start_dt) & (Gledger.Date <= end_dt),
            (Orders.Date2 >= start_dt - datetime.timedelta(days=7)) & (Orders.Date2 <= end_dt - datetime.timedelta(days=7)),
        ),
    ).all()
    for row in income_rows:
        source_date = row.gate_in_date or row.Date
        if source_date is None:
            continue
        recognition_date = business_report_as_date(source_date)
        if row.gate_in_date:
            recognition_date = recognition_date + datetime.timedelta(days=7)
        if start_date <= recognition_date <= end_date:
            amount_cents = int(row.Credit or 0) - int(row.Debit or 0)
            daily_income[recognition_date] = daily_income.get(recognition_date, 0) + amount_cents

    direct_daily = operations_direct_daily_allocations(company_code, start_date, end_date)

    indirect_daily = {}
    for row in db.session.execute(text("""
        SELECT PeriodStart, PeriodEnd, AmountCents, AllocationBasis
        FROM operations_indirect_allocations
        WHERE Co = :company_code
          AND Active = 1
          AND PeriodEnd >= :start_date
          AND PeriodStart <= :end_date
    """), {'company_code': company_code, 'start_date': start_date, 'end_date': end_date}).mappings().all():
        period_start = row['PeriodStart']
        period_end = row['PeriodEnd']
        period_start = business_report_as_date(period_start)
        period_end = business_report_as_date(period_end)
        # Indirect rows represent a payment coverage period, for example an
        # insurance policy from June-to-June. Allocate across the full coverage
        # period, then include only the days that land in the report window.
        if (row['AllocationBasis'] or 'daily') == 'port_trip':
            business_report_allocate_cents_by_port_trip(
                indirect_daily,
                period_start,
                period_end,
                int(row['AmountCents'] or 0),
                start_date,
                end_date,
            )
        else:
            business_report_allocate_cents_by_period(
                indirect_daily,
                period_start,
                period_end,
                int(row['AmountCents'] or 0),
                start_date,
                end_date,
            )

    ga_daily = {}
    for row in db.session.execute(text("""
        SELECT PeriodStart, PeriodEnd, AmountCents, AllocationBasis
        FROM operations_ga_allocations
        WHERE Co = :company_code
          AND Active = 1
          AND PeriodEnd >= :start_date
          AND PeriodStart <= :end_date
    """), {'company_code': company_code, 'start_date': start_date, 'end_date': end_date}).mappings().all():
        period_start = row['PeriodStart']
        period_end = row['PeriodEnd']
        period_start = business_report_as_date(period_start)
        period_end = business_report_as_date(period_end)
        if (row['AllocationBasis'] or 'daily') == 'port_trip':
            business_report_allocate_cents_by_port_trip(
                ga_daily,
                period_start,
                period_end,
                int(row['AmountCents'] or 0),
                start_date,
                end_date,
            )
        else:
            business_report_allocate_cents_by_period(
                ga_daily,
                period_start,
                period_end,
                int(row['AmountCents'] or 0),
                start_date,
                end_date,
            )

    weeks = {}
    for day in business_report_each_day(start_date, end_date):
        week_start = business_report_week_start(day)
        week_end = week_start + datetime.timedelta(days=6)
        week = weeks.setdefault(week_start, {
            'week_start': week_start,
            'week_end': week_end,
            'income_cents': 0,
            'direct_cents': 0,
            'indirect_cents': 0,
            'ga_cents': 0,
        })
        week['income_cents'] += daily_income.get(day, 0)
        week['direct_cents'] += direct_daily.get(day, 0)
        week['indirect_cents'] += indirect_daily.get(day, 0)
        week['ga_cents'] += ga_daily.get(day, 0)

    output = []
    for week_start in sorted(weeks):
        row = weeks[week_start]
        row['week_start_text'] = max(row['week_start'], start_date).strftime('%Y-%m-%d')
        row['week_end_text'] = min(row['week_end'], end_date).strftime('%Y-%m-%d')
        row['gross_profit_cents'] = row['income_cents'] - row['direct_cents']
        row['net_ops_cents'] = row['gross_profit_cents'] - row['indirect_cents']
        row['net_profit_cents'] = row['net_ops_cents'] - row['ga_cents']
        row['income'] = business_report_money(row['income_cents'])
        row['direct'] = business_report_money(row['direct_cents'])
        row['gross_profit'] = business_report_money(row['gross_profit_cents'])
        row['indirect'] = business_report_money(row['indirect_cents'])
        row['net_ops'] = business_report_money(row['net_ops_cents'])
        row['ga'] = business_report_money(row['ga_cents'])
        row['net_profit'] = business_report_money(row['net_profit_cents'])
        output.append(row)

    totals = {
        'income_cents': sum(row['income_cents'] for row in output),
        'direct_cents': sum(row['direct_cents'] for row in output),
        'gross_profit_cents': sum(row['gross_profit_cents'] for row in output),
        'indirect_cents': sum(row['indirect_cents'] for row in output),
        'net_ops_cents': sum(row['net_ops_cents'] for row in output),
        'ga_cents': sum(row['ga_cents'] for row in output),
        'net_profit_cents': sum(row['net_profit_cents'] for row in output),
    }
    totals.update({
        'income': business_report_money(totals['income_cents']),
        'direct': business_report_money(totals['direct_cents']),
        'gross_profit': business_report_money(totals['gross_profit_cents']),
        'indirect': business_report_money(totals['indirect_cents']),
        'net_ops': business_report_money(totals['net_ops_cents']),
        'ga': business_report_money(totals['ga_cents']),
        'net_profit': business_report_money(totals['net_profit_cents']),
    })
    return output, totals


@main.route('/BusinessReports', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def BusinessReports():
    return redirect(url_for('main.ProfitLossYearToDate'))


@main.route('/BusinessReports/Weekly', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def ProfitLossWeekly():
    return business_reports_page('weekly')


@main.route('/api/business-reports/weekly-allocation-plot')
@login_required
@financial_mfa_required
def WeeklyAllocationPlotApi():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_business_report_tables()
    companies = tax_rollup_company_options()
    selected_company = request.args.get('company') or cmpdata[10]
    if selected_company not in companies:
        selected_company = companies[0]
    today_local = datetime.date.today()
    selected_start = business_report_parse_date(request.args.get('date_from'), datetime.date(today_local.year, 1, 1))
    selected_end = business_report_parse_date(request.args.get('date_to'), today_local)
    allocation_type = (request.args.get('allocation_type') or '').strip()
    try:
        allocation_id = int(request.args.get('allocation_id') or 0)
    except ValueError:
        allocation_id = 0
    if allocation_type not in ['indirect', 'ga', 'builtin_direct', 'builtin_driver_pay']:
        return jsonify({'error': 'Choose a valid allocation row.'}), 400
    if allocation_type in ['indirect', 'ga'] and not allocation_id:
        return jsonify({'error': 'Choose a valid allocation row.'}), 400
    try:
        payload = operations_weekly_allocation_plot(
            selected_company,
            allocation_type,
            allocation_id,
            selected_start,
            selected_end,
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 500
    if payload is None:
        return jsonify({'error': 'Allocation row was not found.'}), 404
    return jsonify(payload)


@main.route('/BusinessReports/YearToDate', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def ProfitLossYearToDate():
    return business_reports_page('ytd')


def business_reports_page(report_mode):
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)
    ensure_business_report_tables()
    ensure_tax_rollup_tables()
    companies = tax_rollup_company_options()
    selected_company = request.values.get('company') or cmpdata[10]
    if selected_company not in companies:
        selected_company = companies[0]
    today_local = datetime.date.today()
    default_start = datetime.date(today_local.year, 1, 1)
    default_end = datetime.date(today_local.year, 12, 31)
    selected_start = business_report_parse_date(request.values.get('date_from'), default_start)
    selected_end = business_report_parse_date(request.values.get('date_to'), default_end)
    err = []
    msg = []
    action = None
    force_report_refresh = False

    if request.method == 'POST':
        action = request.form.get('action')
        selected_company = request.form.get('company') or selected_company
        if selected_company not in companies:
            selected_company = companies[0]
        if action == 'refresh_report_cache':
            force_report_refresh = True
        elif action == 'add_allocation':
            account_name = (request.form.get('account_name') or '').strip()
            basis_type = (request.form.get('basis_type') or 'port_entry').strip()
            if basis_type not in BUSINESS_REPORT_BASIS_TYPES:
                basis_type = 'port_entry'
            method = BUSINESS_REPORT_BASIS_TYPES[basis_type]
            basis = (request.form.get('basis_label') or method).strip()
            period_start = business_report_parse_date(request.form.get('period_start'), None)
            period_end = business_report_parse_date(request.form.get('period_end'), None)
            try:
                basis_qty = float(request.form.get('basis_qty') or 0)
            except ValueError:
                basis_qty = 0
            if not account_name or not basis:
                err.append('Account and basis label are required.')
            else:
                db.session.execute(text("""
                    INSERT INTO business_report_allocations
                        (AccountName, AllocationMethod, BasisType, BasisLabel, PeriodStart, PeriodEnd, BasisQty, Notes, Active, CreatedAt, UpdatedAt)
                    VALUES
                        (:account_name, :method, :basis_type, :basis, :period_start, :period_end, :basis_qty, :notes, 1, :created_at, :updated_at)
                """), {
                    'account_name': account_name,
                    'method': method,
                    'basis_type': basis_type,
                    'basis': basis,
                    'period_start': period_start,
                    'period_end': period_end,
                    'basis_qty': basis_qty,
                    'notes': (request.form.get('notes') or '').strip(),
                    'created_at': datetime.datetime.utcnow(),
                    'updated_at': datetime.datetime.utcnow(),
                })
                db.session.commit()
                msg.append('Allocation setup row added.')
        elif action == 'delete_allocation':
            allocation_id = request.form.get('allocation_id')
            db.session.execute(text("""
                UPDATE business_report_allocations
                SET Active = 0, UpdatedAt = :updated_at
                WHERE id = :allocation_id
            """), {'allocation_id': allocation_id, 'updated_at': datetime.datetime.utcnow()})
            db.session.commit()
            msg.append('Allocation setup row removed.')
        elif action == 'add_ops_indirect_allocation':
            account_name = (request.form.get('ops_indirect_account_name') or '').strip()
            period_start = business_report_parse_date(request.form.get('ops_period_start'), None)
            period_end = business_report_parse_date(request.form.get('ops_period_end'), None)
            amount_cents = business_report_cents(request.form.get('ops_amount'))
            allocation_basis = (request.form.get('ops_allocation_basis') or 'daily').strip()
            if allocation_basis not in OPERATIONS_ALLOCATION_BASIS_TYPES:
                allocation_basis = 'daily'
            auto_update = 1 if request.form.get('ops_auto_update') else 0
            if auto_update:
                period_start = datetime.date(today_local.year, 1, 1)
                period_end = today_local
                amount_cents = operations_account_ytd_cents(selected_company, account_name, today_local, 'indirect') if account_name else 0
            if not account_name:
                err.append('Indirect expense account is required.')
            elif period_start is None or period_end is None:
                err.append('Indirect allocation period start and end are required.')
            elif period_end < period_start:
                err.append('Indirect allocation period end must be on or after the start date.')
            elif amount_cents == 0:
                err.append('Indirect allocation amount is required.')
            else:
                db.session.execute(text("""
                    INSERT INTO operations_indirect_allocations
                        (Co, AccountName, PeriodStart, PeriodEnd, AmountCents, AllocationBasis, AutoUpdate, Notes, Active, CreatedAt, UpdatedAt)
                    VALUES
                        (:company_code, :account_name, :period_start, :period_end, :amount_cents, :allocation_basis, :auto_update, :notes, 1, :created_at, :updated_at)
                """), {
                    'company_code': selected_company,
                    'account_name': account_name,
                    'period_start': period_start,
                    'period_end': period_end,
                    'amount_cents': amount_cents,
                    'allocation_basis': allocation_basis,
                    'auto_update': auto_update,
                    'notes': (request.form.get('ops_notes') or '').strip(),
                    'created_at': datetime.datetime.utcnow(),
                    'updated_at': datetime.datetime.utcnow(),
                })
                db.session.commit()
                msg.append('Operations indirect allocation added.')
                force_report_refresh = report_mode == 'weekly'
        elif action == 'delete_ops_indirect_allocation':
            allocation_id = request.form.get('ops_allocation_id')
            db.session.execute(text("""
                UPDATE operations_indirect_allocations
                SET Active = 0, UpdatedAt = :updated_at
                WHERE id = :allocation_id AND Co = :company_code
            """), {
                'allocation_id': allocation_id,
                'company_code': selected_company,
                'updated_at': datetime.datetime.utcnow(),
            })
            db.session.commit()
            msg.append('Operations indirect allocation removed.')
            force_report_refresh = report_mode == 'weekly'
        elif action == 'add_ops_ga_allocation':
            account_name = (request.form.get('ops_ga_account_name') or '').strip()
            period_start = business_report_parse_date(request.form.get('ops_ga_period_start'), None)
            period_end = business_report_parse_date(request.form.get('ops_ga_period_end'), None)
            amount_cents = business_report_cents(request.form.get('ops_ga_amount'))
            allocation_basis = (request.form.get('ops_ga_allocation_basis') or 'daily').strip()
            if allocation_basis not in OPERATIONS_ALLOCATION_BASIS_TYPES:
                allocation_basis = 'daily'
            auto_update = 1 if request.form.get('ops_ga_auto_update') else 0
            if auto_update:
                period_start = datetime.date(today_local.year, 1, 1)
                period_end = today_local
                amount_cents = operations_account_ytd_cents(selected_company, account_name, today_local, 'g-a') if account_name else 0
            if not account_name:
                err.append('G&A expense account is required.')
            elif period_start is None or period_end is None:
                err.append('G&A allocation period start and end are required.')
            elif period_end < period_start:
                err.append('G&A allocation period end must be on or after the start date.')
            elif amount_cents == 0:
                err.append('G&A allocation amount is required.')
            else:
                db.session.execute(text("""
                    INSERT INTO operations_ga_allocations
                        (Co, AccountName, PeriodStart, PeriodEnd, AmountCents, AllocationBasis, AutoUpdate, Notes, Active, CreatedAt, UpdatedAt)
                    VALUES
                        (:company_code, :account_name, :period_start, :period_end, :amount_cents, :allocation_basis, :auto_update, :notes, 1, :created_at, :updated_at)
                """), {
                    'company_code': selected_company,
                    'account_name': account_name,
                    'period_start': period_start,
                    'period_end': period_end,
                    'amount_cents': amount_cents,
                    'allocation_basis': allocation_basis,
                    'auto_update': auto_update,
                    'notes': (request.form.get('ops_ga_notes') or '').strip(),
                    'created_at': datetime.datetime.utcnow(),
                    'updated_at': datetime.datetime.utcnow(),
                })
                db.session.commit()
                msg.append('G&A allocation added.')
                force_report_refresh = report_mode == 'weekly'
        elif action == 'delete_ops_ga_allocation':
            allocation_id = request.form.get('ops_ga_allocation_id')
            db.session.execute(text("""
                UPDATE operations_ga_allocations
                SET Active = 0, UpdatedAt = :updated_at
                WHERE id = :allocation_id AND Co = :company_code
            """), {
                'allocation_id': allocation_id,
                'company_code': selected_company,
                'updated_at': datetime.datetime.utcnow(),
            })
            db.session.commit()
            msg.append('G&A allocation removed.')
            force_report_refresh = report_mode == 'weekly'
        elif action == 'export_expenses_csv':
            expenses = business_report_expenses(selected_start, selected_end, selected_company)
            lines = ['Account,Category,Subcategory,Amount']
            for row in expenses:
                lines.append(f"\"{row['account']}\",\"{row['category']}\",\"{row['subcategory']}\",{row['amount_cents'] / 100:.2f}")
            csv_data = '\n'.join(lines)
            filename = f'business_expenses_{selected_start}_{selected_end}.csv'
            return Response(
                csv_data,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename={filename}'},
            )
        elif action == 'export_tax_rollup_xlsx':
            tax_rollup_pl = business_report_tax_rollup_pl(selected_company, selected_start, selected_end)
            return business_report_tax_rollup_xlsx(selected_company, selected_start, selected_end, tax_rollup_pl)
        elif action == 'export_tax_rollup_pdf':
            tax_rollup_pl = business_report_tax_rollup_pl(selected_company, selected_start, selected_end)
            return business_report_tax_rollup_pdf(selected_company, selected_start, selected_end, tax_rollup_pl)

    operations_end = selected_end
    if report_mode == 'weekly':
        operations_end = min(selected_end, today_local)
        operations_end = operations_end - datetime.timedelta(days=(operations_end.weekday() + 1) % 7)

    if report_mode == 'weekly':
        try:
            operations_sync_auto_allocations(selected_company, today_local, 'operations_indirect_allocations', 'indirect')
            operations_sync_auto_allocations(selected_company, today_local, 'operations_ga_allocations', 'g-a')
        except Exception as exc:
            db.session.rollback()
            err.append(f'Auto allocation update skipped: {exc}')

    expense_rows = []
    total_expenses = 0
    tax_rollup_pl = business_report_empty_tax_rollup_pl()
    operations_rows, operations_totals = [], business_report_empty_operations_totals()
    unassigned_expense_lines = []
    operations_effective_date_to = operations_end.strftime('%Y-%m-%d') if operations_end >= selected_start else ''
    report_cache_info = {'source': 'empty', 'updated_at': '', 'generated_by': ''}
    payload, cache_row = business_report_cache_load(selected_company, report_mode, selected_start, selected_end)
    if force_report_refresh or payload is None:
        try:
            if report_mode == 'weekly':
                operations_rows, operations_totals = operations_weekly_profit_loss(selected_company, selected_start, operations_end)
                payload = {
                    'operations_rows': operations_rows,
                    'operations_totals': operations_totals,
                    'operations_effective_date_to': operations_effective_date_to,
                }
            else:
                tax_rollup_pl = business_report_tax_rollup_pl(selected_company, selected_start, selected_end)
                unassigned_expense_lines = business_report_unassigned_expense_lines(selected_company, selected_start, selected_end)
                payload = {
                    'tax_rollup_pl': tax_rollup_pl,
                    'unassigned_expense_lines': unassigned_expense_lines,
                    'expense_rows': [],
                    'total_expenses_cents': tax_rollup_pl.get('total_expenses_cents', 0),
                }
            payload, cache_row = business_report_cache_save(selected_company, report_mode, selected_start, selected_end, payload)
            report_cache_info = business_report_cache_info(cache_row, 'fresh')
            msg.append('Report cache refreshed.')
        except Exception as exc:
            db.session.rollback()
            payload = None
            if report_mode == 'weekly':
                err.append(f'Weekly operations report could not be loaded: {exc}')
            else:
                err.append(f'Tax rollup report could not be loaded: {exc}')
    else:
        report_cache_info = business_report_cache_info(cache_row, 'saved')

    if payload:
        if report_mode == 'weekly':
            operations_rows = payload.get('operations_rows', [])
            operations_totals = payload.get('operations_totals') or business_report_empty_operations_totals()
            operations_effective_date_to = payload.get('operations_effective_date_to') or operations_effective_date_to
        else:
            tax_rollup_pl = payload.get('tax_rollup_pl') or business_report_empty_tax_rollup_pl()
            unassigned_expense_lines = payload.get('unassigned_expense_lines', [])
            expense_rows = payload.get('expense_rows', [])
            total_expenses = int(payload.get('total_expenses_cents') or tax_rollup_pl.get('total_expenses_cents') or 0)

    try:
        ops_indirect_rows = operations_indirect_allocations(selected_company) if report_mode == 'weekly' else []
    except Exception as exc:
        db.session.rollback()
        ops_indirect_rows = []
        err.append(f'Indirect allocations could not be loaded: {exc}')

    try:
        ops_ga_rows = operations_ga_allocations(selected_company) if report_mode == 'weekly' else []
    except Exception as exc:
        db.session.rollback()
        ops_ga_rows = []
        err.append(f'G&A allocations could not be loaded: {exc}')

    try:
        ops_indirect_accounts = operations_account_options(selected_company, today_local, 'indirect') if report_mode == 'weekly' else []
    except Exception as exc:
        db.session.rollback()
        ops_indirect_accounts = []
        err.append(f'Indirect account options could not be loaded: {exc}')

    try:
        ops_ga_accounts = operations_account_options(selected_company, today_local, 'g-a') if report_mode == 'weekly' else []
    except Exception as exc:
        db.session.rollback()
        ops_ga_accounts = []
        err.append(f'G&A account options could not be loaded: {exc}')
    report_basis = {
        'port_entries': business_report_number(0, 0),
        'days': business_report_number(0, 0),
        'miles': business_report_number(0, 2),
    }
    return render_template(
        'business_reports.html',
        cmpdata=cmpdata,
        scac=scac,
        filters={
            'company': selected_company,
            'date_from': selected_start.strftime('%Y-%m-%d'),
            'date_to': selected_end.strftime('%Y-%m-%d'),
        },
        report_mode=report_mode,
        report_endpoint='main.ProfitLossWeekly' if report_mode == 'weekly' else 'main.ProfitLossYearToDate',
        companies=companies,
        selected_company=selected_company,
        tax_rollup_pl=tax_rollup_pl,
        operations_rows=operations_rows,
        operations_totals=operations_totals,
        operations_effective_date_to=operations_effective_date_to,
        ops_indirect_rows=ops_indirect_rows,
        ops_indirect_accounts=ops_indirect_accounts,
        ops_ga_rows=ops_ga_rows,
        ops_ga_accounts=ops_ga_accounts,
        unassigned_expense_lines=unassigned_expense_lines,
        expense_rows=expense_rows,
        total_expenses=business_report_money(total_expenses),
        basis_types=BUSINESS_REPORT_BASIS_TYPES,
        operations_allocation_basis_types=OPERATIONS_ALLOCATION_BASIS_TYPES,
        report_basis=report_basis,
        report_cache_info=report_cache_info,
        ops_default_start=datetime.date(today_local.year, 1, 1).strftime('%Y-%m-%d'),
        ops_default_end=today_local.strftime('%Y-%m-%d'),
        err=err,
        msg=msg,
    )


@main.route('/IntercompanyEntries', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def IntercompanyEntries():
    company_names = {
        'K': 'One Stop Logistics',
        'J': 'Jays Auto',
        'N': 'Owner Personal',
    }
    allowed_companies = list(company_names.keys())
    owner_cash_account_name = 'Owner Personal Cash'

    def parse_date(value):
        if not value:
            return datetime.datetime.today()
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return None

    def cents(value):
        clean = (value or '').replace('$', '').replace(',', '').strip()
        if not clean:
            return 0
        try:
            return int((Decimal(clean) * 100).quantize(Decimal('1')))
        except Exception:
            return None

    def money(value):
        return "${:,.2f}".format((value or 0) / 100)

    def company_label(code):
        return f"{company_names.get(code, code)} Account {code}"

    def is_transfer_endpoint_account(account):
        if account is None or account.Co not in allowed_companies:
            return False

        account_type = account.Type or ''
        account_text = ' '.join([
            account.Name or '',
            account.Description or '',
            account.Category or '',
            account.Subcategory or '',
        ]).lower()

        if account_type in ['Bank', 'Exch']:
            return True

        if account_type in ['Credit Card']:
            return True

        if account_type == 'Current Liability':
            internal_terms = [
                'accounts payable',
                'due to',
                'payroll',
                'tax',
                'opening balance',
                'accrued',
            ]
            return not any(term in account_text for term in internal_terms)

        return False

    def find_owner_cash_account():
        for account_name in [owner_cash_account_name, 'Owner Cash', 'Owner Personal', 'Owner']:
            account = Accounts.query.filter(
                (Accounts.Co == 'N') &
                (Accounts.Name == account_name)
            ).first()
            if is_transfer_endpoint_account(account):
                return account

        return Accounts.query.filter(
            (Accounts.Co == 'N') &
            (Accounts.Type.in_(['Bank', 'Exch'])) &
            (Accounts.Name.contains('Owner'))
        ).first()

    def ensure_owner_cash_account():
        account = find_owner_cash_account()
        if account is not None:
            return account, None

        account = Accounts(
            Name=owner_cash_account_name,
            Balance=0.00,
            AcctNumber=None,
            Routing=None,
            Payee=None,
            Type='Exch',
            Description='Owner personal cash endpoint for account transfers',
            Category='Assets',
            Subcategory='Cash',
            Taxrollup=None,
            Co='N',
            QBmap=None,
            Shared=None,
        )
        db.session.add(account)
        db.session.flush()
        return account, f'Created {owner_cash_account_name} for owner personal transfers.'

    def find_due_account(book_company, target_company):
        target_name = company_names.get(target_company, target_company)
        base_query = Accounts.query.filter(
            (Accounts.Co == book_company) &
            (Accounts.Name.contains('Due'))
        )
        for term in [target_name, target_company]:
            account = base_query.filter(Accounts.Name.contains(term)).first()
            if account is not None:
                return account
        return None

    def ensure_due_account(book_company, target_company):
        account = find_due_account(book_company, target_company)
        if account is not None:
            return account, None

        target_name = company_names.get(target_company, target_company)
        account = Accounts(
            Name=f'Due to {target_name}',
            Balance=0.00,
            AcctNumber=None,
            Routing=None,
            Payee=None,
            Type='Current Liability',
            Description='Created automatically for account transfers',
            Category='Liabilities',
            Subcategory='Current Liabilities',
            Taxrollup=None,
            Co=book_company,
            QBmap=None,
            Shared=None,
        )
        db.session.add(account)
        db.session.flush()
        return account, f'Created {account.Name} in {company_label(book_company)}.'

    def find_owner_equity_account(company_code):
        query = Accounts.query.filter(
            (Accounts.Co == company_code) &
            (Accounts.Type == 'Equity')
        )
        for term in ['Owner', 'Draw', 'Distribution', 'Capital', 'Equity']:
            account = query.filter(Accounts.Name.contains(term)).first()
            if account is not None:
                return account
        return query.first()

    def ensure_owner_equity_account(company_code):
        account = find_owner_equity_account(company_code)
        if account is not None:
            return account

        if company_code == 'N':
            account_name = 'Owner Equity'
            description = 'Created automatically for owner personal transfers'
            subcategory = 'Owner Equity'
        else:
            account_name = 'Owner Draw / Distributions'
            description = 'Created automatically for owner equity transfers'
            subcategory = 'Owner Draws'

        account = Accounts(
            Name=account_name,
            Balance=0.00,
            AcctNumber=None,
            Routing=None,
            Payee=None,
            Type='Equity',
            Description=description,
            Category='Equity',
            Subcategory=subcategory,
            Taxrollup=None,
            Co=company_code,
            QBmap=None,
            Shared=None,
        )
        db.session.add(account)
        db.session.flush()
        return account

    def build_line(amount, debit, account, source, line_type, tcode, company_code, entry_date, ref):
        return {
            'debit': amount if debit else 0,
            'credit': 0 if debit else amount,
            'account': account.Name,
            'aid': account.id,
            'source': source,
            'sid': 0,
            'type': line_type,
            'tcode': tcode,
            'com': company_code,
            'recorded': datetime.datetime.now(),
            'date': entry_date,
            'ref': ref,
            'match_aid': True,
        }

    def build_transfer_line(amount, debit, account, source_account, line_type, tcode, entry_date, ref):
        return {
            'debit': amount if debit else 0,
            'credit': 0 if debit else amount,
            'account': account.Name,
            'aid': account.id,
            'source': source_account.Name,
            'sid': source_account.id,
            'type': line_type,
            'tcode': tcode,
            'com': account.Co,
            'recorded': datetime.datetime.now(),
            'date': entry_date,
            'ref': ref,
            'match_aid': True,
        }

    def transfer_journal_lines(amount, from_account, to_account, tcode, entry_date, ref, owner_transfer_treatment):
        if from_account.Co == to_account.Co:
            return [
                build_transfer_line(amount, True, to_account, from_account, 'XD', tcode, entry_date, ref),
                build_transfer_line(amount, False, from_account, to_account, 'XC', tcode, entry_date, ref),
            ], []

        owner_involved = 'N' in [from_account.Co, to_account.Co]
        if owner_involved and owner_transfer_treatment == 'equity':
            from_equity = ensure_owner_equity_account(from_account.Co)
            to_equity = ensure_owner_equity_account(to_account.Co)
            errs = []
            if from_equity is None:
                errs.append(f'Missing owner equity account in {company_label(from_account.Co)}.')
            if to_equity is None:
                errs.append(f'Missing owner equity account in {company_label(to_account.Co)}.')
            if errs:
                return [], errs

            return [
                build_transfer_line(amount, True, to_account, from_account, 'XD', tcode, entry_date, ref),
                build_transfer_line(amount, False, to_equity, from_account, 'OC', tcode, entry_date, ref),
                build_transfer_line(amount, True, from_equity, to_account, 'OD', tcode, entry_date, ref),
                build_transfer_line(amount, False, from_account, to_account, 'XC', tcode, entry_date, ref),
            ], []

        due_in_from_company, from_note = ensure_due_account(from_account.Co, to_account.Co)
        due_in_to_company, to_note = ensure_due_account(to_account.Co, from_account.Co)
        errs = []
        if due_in_from_company is None:
            errs.append(f'Missing due-to account in {company_label(from_account.Co)} for {company_label(to_account.Co)}.')
        if due_in_to_company is None:
            errs.append(f'Missing due-to account in {company_label(to_account.Co)} for {company_label(from_account.Co)}.')
        if errs:
            return [], errs

        return [
            build_transfer_line(amount, True, due_in_from_company, to_account, 'IA', tcode, entry_date, ref),
            build_transfer_line(amount, False, from_account, to_account, 'XC', tcode, entry_date, ref),
            build_transfer_line(amount, True, to_account, from_account, 'XD', tcode, entry_date, ref),
            build_transfer_line(amount, False, due_in_to_company, from_account, 'IL', tcode, entry_date, ref),
        ], []

    def transfer_lines_for_journal(journal_id):
        if not journal_id:
            return []
        return Gledger.query.filter(
            (Gledger.JournalId == journal_id) &
            (Gledger.SourceTable == 'AccountTransfer')
        ).order_by(Gledger.JournalSeq, Gledger.id).all()

    def has_final_reconciliation(lines):
        return any(line.Reconciled not in [None, 0, 25] for line in lines)

    def line_has_final_reconciliation(line):
        return line is not None and line.Reconciled not in [None, 0, 25]

    def transfer_lock_state(lines):
        debit_line = next((line for line in lines if line.Type == 'XD'), None)
        credit_line = next((line for line in lines if line.Type == 'XC'), None)
        return {
            'transfer_from_locked': line_has_final_reconciliation(credit_line),
            'transfer_to_locked': line_has_final_reconciliation(debit_line),
            'transfer_core_locked': has_final_reconciliation(lines),
            'transfer_partial_lock_message': (
                'One side of this transfer has been reconciled. Only unreconciled account-side repairs are allowed.'
                if has_final_reconciliation(lines) else ''
            ),
        }

    def transfer_selection_from_lines(lines):
        debit_line = next((line for line in lines if line.Type == 'XD'), None)
        credit_line = next((line for line in lines if line.Type == 'XC'), None)
        if debit_line is None or credit_line is None:
            return None

        line_date = debit_line.Date or credit_line.Date
        line_types = [line.Type for line in lines]
        company_codes = {line.Com for line in lines if line.Com}
        owner_transfer_treatment = ''
        if 'OD' in line_types or 'OC' in line_types:
            owner_transfer_treatment = 'equity'
        elif 'N' in company_codes and len(company_codes) > 1:
            owner_transfer_treatment = 'loan'

        return {
            'edit_transfer_journal_id': debit_line.JournalId or credit_line.JournalId or '',
            'transfer_date': line_date.strftime('%Y-%m-%d') if line_date else datetime.date.today().strftime('%Y-%m-%d'),
            'transfer_from_account_id': str(credit_line.Aid or ''),
            'transfer_to_account_id': str(debit_line.Aid or ''),
            'transfer_amount': money(debit_line.Debit or credit_line.Credit or 0).replace('$', ''),
            'transfer_ref': debit_line.Ref or credit_line.Ref or '',
            'transfer_memo': debit_line.JournalMemo or credit_line.JournalMemo or '',
            'owner_transfer_treatment': owner_transfer_treatment,
            **transfer_lock_state(lines),
        }

    def transfer_date_matches(line, transfer_date):
        if line is None:
            return False
        line_date = line.Date
        if isinstance(line_date, datetime.date) and not isinstance(line_date, datetime.datetime):
            line_date = datetime.datetime.combine(line_date, datetime.time.min)
        return line_date == transfer_date

    def update_open_transfer_side(existing_lines, from_account, to_account, transfer_date, amount, ref, memo):
        debit_line = next((line for line in existing_lines if line.Type == 'XD'), None)
        credit_line = next((line for line in existing_lines if line.Type == 'XC'), None)
        if debit_line is None or credit_line is None or len(existing_lines) != 2:
            return ['Partially reconciled transfer repair is only supported for standard two-line account transfers.']

        from_locked = line_has_final_reconciliation(credit_line)
        to_locked = line_has_final_reconciliation(debit_line)
        if from_locked and to_locked:
            return ['Both sides of this transfer have been reconciled. Reopen the reconciliation statement before editing it.']
        if not from_locked and not to_locked:
            return []

        if credit_line.Aid != from_account.id:
            return ['Pay From has been reconciled and cannot be changed.']
        if (credit_line.Credit or 0) != amount or (debit_line.Debit or 0) != amount:
            return ['Amount cannot be changed because one side of this transfer has been reconciled.']
        if not transfer_date_matches(credit_line, transfer_date) or not transfer_date_matches(debit_line, transfer_date):
            return ['Date cannot be changed because one side of this transfer has been reconciled.']

        if from_locked:
            if to_locked:
                return ['Pay To has been reconciled and cannot be changed.']
            if from_account.Co != to_account.Co:
                return ['Cannot convert a partially reconciled standard transfer into a cross-company transfer.']
            debit_line.Account = to_account.Name
            debit_line.Aid = to_account.id
            debit_line.Source = from_account.Name
            debit_line.Sid = from_account.id
            debit_line.Com = to_account.Co
            debit_line.Ref = ref
            debit_line.JournalMemo = memo or debit_line.JournalMemo
            debit_line.PostedBy = 'account_transfer_repair'
            debit_line.PostedAt = datetime.datetime.now()
            db.session.commit()
            return []

        if to_locked:
            if from_account.Co != to_account.Co:
                return ['Cannot convert a partially reconciled standard transfer into a cross-company transfer.']
            credit_line.Account = from_account.Name
            credit_line.Aid = from_account.id
            credit_line.Source = to_account.Name
            credit_line.Sid = to_account.id
            credit_line.Com = from_account.Co
            credit_line.Ref = ref
            credit_line.JournalMemo = memo or credit_line.JournalMemo
            credit_line.PostedBy = 'account_transfer_repair'
            credit_line.PostedAt = datetime.datetime.now()
            db.session.commit()
            return []

        return ['This transfer could not be repaired.']

    selected = {
        'entry_date': datetime.date.today().strftime('%Y-%m-%d'),
        'entry_type': request.values.get('entry_type', 'Expense'),
        'cash_account_id': request.values.get('cash_account_id', ''),
        'operating_company': request.values.get('operating_company', 'K'),
        'operating_account_id': request.values.get('operating_account_id', ''),
        'transfer_date': request.values.get('transfer_date', datetime.date.today().strftime('%Y-%m-%d')),
        'transfer_from_account_id': request.values.get('transfer_from_account_id', ''),
        'transfer_to_account_id': request.values.get('transfer_to_account_id', ''),
        'transfer_amount': request.values.get('transfer_amount', ''),
        'transfer_ref': request.values.get('transfer_ref', ''),
        'transfer_memo': request.values.get('transfer_memo', ''),
        'owner_transfer_treatment': request.values.get('owner_transfer_treatment', ''),
        'edit_transfer_journal_id': request.values.get('edit_transfer_journal_id', ''),
        'transfer_from_locked': False,
        'transfer_to_locked': False,
        'transfer_core_locked': False,
        'transfer_partial_lock_message': '',
        'amount': request.values.get('amount', ''),
        'source': request.values.get('source', ''),
        'ref': request.values.get('ref', ''),
        'memo': request.values.get('memo', ''),
    }
    err = []
    msg = ''

    if request.method == 'POST' and request.values.get('create_entry'):
        entry_date = parse_date(request.values.get('entry_date'))
        entry_type = request.values.get('entry_type', 'Expense')
        amount = cents(request.values.get('amount'))
        cash_account = Accounts.query.get(request.values.get('cash_account_id') or 0)
        operating_account = Accounts.query.get(request.values.get('operating_account_id') or 0)
        operating_company = request.values.get('operating_company')
        source = request.values.get('source', '').strip()
        ref = request.values.get('ref', '').strip()
        memo = request.values.get('memo', '').strip()

        if entry_date is None:
            err.append('Enter a valid transaction date.')
        if amount is None or amount <= 0:
            err.append('Enter an amount greater than zero.')
        if cash_account is None or cash_account.Co not in allowed_companies:
            err.append('Choose a valid K, J, or N cash account.')
        if operating_company not in allowed_companies:
            err.append('Choose a valid operating company.')
        if operating_account is None or operating_account.Co != operating_company:
            err.append('Choose an operating account for the selected operating company.')
        if operating_account is not None and entry_type == 'Expense' and operating_account.Type != 'Expense':
            err.append('Expense entries must use an expense operating account.')
        if operating_account is not None and entry_type == 'Income' and operating_account.Type != 'Income':
            err.append('Income entries must use an income operating account.')
        if not source:
            err.append('Enter a source, payee, customer, or owner name.')

        if not err:
            cash_company = cash_account.Co
            journal_memo = memo or f'{entry_type} for {company_label(operating_company)} through {company_label(cash_company)}'
            tcode = None
            journal_id = None
            lines = []

            if cash_company == operating_company:
                tcode = newjo('IC', entry_date.strftime('%Y-%m-%d'))
                journal_id = f'INTERCOMPANY-{tcode}'
                if entry_type == 'Expense':
                    lines = [
                        build_line(amount, True, operating_account, source, 'IE', tcode, operating_company, entry_date, ref),
                        build_line(amount, False, cash_account, source, 'PC', tcode, cash_company, entry_date, ref),
                    ]
                else:
                    lines = [
                        build_line(amount, True, cash_account, source, 'DD', tcode, cash_company, entry_date, ref),
                        build_line(amount, False, operating_account, source, 'IR', tcode, operating_company, entry_date, ref),
                    ]
            elif entry_type == 'Expense':
                due_in_cash_company = find_due_account(cash_company, operating_company)
                due_in_operating_company = find_due_account(operating_company, cash_company)
                if due_in_cash_company is None:
                    err.append(f'Missing due-to account in {company_label(cash_company)} for {company_label(operating_company)}.')
                if due_in_operating_company is None:
                    err.append(f'Missing due-to account in {company_label(operating_company)} for {company_label(cash_company)}.')
                if not err:
                    tcode = newjo('IC', entry_date.strftime('%Y-%m-%d'))
                    journal_id = f'INTERCOMPANY-{tcode}'
                    lines = [
                        build_line(amount, True, operating_account, source, 'IE', tcode, operating_company, entry_date, ref),
                        build_line(amount, False, due_in_operating_company, source, 'IL', tcode, operating_company, entry_date, ref),
                        build_line(amount, True, due_in_cash_company, source, 'IA', tcode, cash_company, entry_date, ref),
                        build_line(amount, False, cash_account, source, 'PC', tcode, cash_company, entry_date, ref),
                    ]
            else:
                due_in_cash_company = find_due_account(cash_company, operating_company)
                due_in_operating_company = find_due_account(operating_company, cash_company)
                if due_in_cash_company is None:
                    err.append(f'Missing due-to account in {company_label(cash_company)} for {company_label(operating_company)}.')
                if due_in_operating_company is None:
                    err.append(f'Missing due-to account in {company_label(operating_company)} for {company_label(cash_company)}.')
                if not err:
                    tcode = newjo('IC', entry_date.strftime('%Y-%m-%d'))
                    journal_id = f'INTERCOMPANY-{tcode}'
                    lines = [
                        build_line(amount, True, cash_account, source, 'DD', tcode, cash_company, entry_date, ref),
                        build_line(amount, False, due_in_cash_company, source, 'IL', tcode, cash_company, entry_date, ref),
                        build_line(amount, True, due_in_operating_company, source, 'IA', tcode, operating_company, entry_date, ref),
                        build_line(amount, False, operating_account, source, 'IR', tcode, operating_company, entry_date, ref),
                    ]

            if not err:
                post_err = post_balanced_journal(
                    lines,
                    journal_id=journal_id,
                    journal_memo=journal_memo,
                    posted_by='intercompany',
                    source_table='IntercompanyEntry',
                )
                if post_err:
                    err.extend(post_err)
                else:
                    msg = f'Recorded intercompany entry {tcode}.'
                    selected = {
                        'entry_date': datetime.date.today().strftime('%Y-%m-%d'),
                        'entry_type': entry_type,
                        'cash_account_id': '',
                        'operating_company': operating_company,
                        'operating_account_id': '',
                        'amount': '',
                        'source': '',
                        'ref': '',
                        'memo': '',
                    }

    if request.method == 'POST' and request.values.get('load_transfer'):
        journal_id = request.values.get('selected_transfer_journal_id', '').strip()
        lines = transfer_lines_for_journal(journal_id)
        if not lines:
            err.append('Choose an account transfer to edit.')
        else:
            transfer_selected = transfer_selection_from_lines(lines)
            if transfer_selected is None:
                err.append('The selected transfer could not be loaded because its ledger lines are incomplete.')
            else:
                selected.update(transfer_selected)
                msg = f'Loaded transfer {lines[0].Tcode} for editing.'

    if request.method == 'POST' and request.values.get('delete_transfer'):
        journal_id = request.values.get('selected_transfer_journal_id', '').strip()
        lines = transfer_lines_for_journal(journal_id)
        if not lines:
            err.append('Choose an account transfer to delete.')
        elif has_final_reconciliation(lines):
            err.append('This transfer has been reconciled. Reopen the reconciliation statement before deleting it.')
        else:
            tcode = lines[0].Tcode
            for line in lines:
                db.session.delete(line)
            db.session.commit()
            selected['edit_transfer_journal_id'] = ''
            msg = f'Deleted account transfer {tcode}.'

    if request.method == 'POST' and request.values.get('create_transfer'):
        transfer_date = parse_date(request.values.get('transfer_date'))
        amount = cents(request.values.get('transfer_amount'))
        from_account = Accounts.query.get(request.values.get('transfer_from_account_id') or 0)
        to_account = Accounts.query.get(request.values.get('transfer_to_account_id') or 0)
        ref = request.values.get('transfer_ref', '').strip()
        memo = request.values.get('transfer_memo', '').strip()
        owner_transfer_treatment = request.values.get('owner_transfer_treatment', '').strip()
        edit_journal_id = request.values.get('edit_transfer_journal_id', '').strip()
        existing_transfer_lines = transfer_lines_for_journal(edit_journal_id) if edit_journal_id else []

        if transfer_date is None:
            err.append('Enter a valid transfer date.')
        if amount is None or amount <= 0:
            err.append('Enter a transfer amount greater than zero.')
        if from_account is None:
            err.append('Choose a valid account to pay from.')
        if to_account is None:
            err.append('Choose a valid account to pay to.')
        if from_account is not None and not is_transfer_endpoint_account(from_account):
            err.append('Pay From must be a bank, exchange, merchant, or credit card account.')
        if to_account is not None and not is_transfer_endpoint_account(to_account):
            err.append('Pay To must be a bank, exchange, merchant, or credit card account.')
        if from_account is not None and to_account is not None:
            if from_account.id == to_account.id:
                err.append('The transfer accounts must be different.')
            if from_account.Co != to_account.Co and 'N' in [from_account.Co, to_account.Co] and owner_transfer_treatment not in ['loan', 'equity']:
                err.append('Choose whether this owner transfer is repayable or owner equity.')
        if edit_journal_id and not existing_transfer_lines:
            err.append('The transfer being edited could not be found.')
        if existing_transfer_lines:
            selected.update(transfer_lock_state(existing_transfer_lines))

        if not err:
            if existing_transfer_lines:
                tcode = existing_transfer_lines[0].Tcode
                journal_id = existing_transfer_lines[0].JournalId or f'TRANSFER-{tcode}'
            else:
                tcode = newjo('XF', transfer_date.strftime('%Y-%m-%d'))
                journal_id = f'TRANSFER-{tcode}'
            journal_memo = memo or f'Transfer from {from_account.Name} to {to_account.Name}'
            lines, transfer_err = transfer_journal_lines(amount, from_account, to_account, tcode, transfer_date, ref, owner_transfer_treatment)
            if transfer_err:
                err.extend(transfer_err)
            elif existing_transfer_lines and has_final_reconciliation(existing_transfer_lines):
                repair_err = update_open_transfer_side(existing_transfer_lines, from_account, to_account, transfer_date, amount, ref, journal_memo)
                if repair_err:
                    err.extend(repair_err)
                else:
                    msg = f'Updated unreconciled side of account transfer {tcode}.'
                    selected.update(transfer_selection_from_lines(transfer_lines_for_journal(journal_id)) or {})
            else:
                for existing_line in existing_transfer_lines:
                    db.session.delete(existing_line)
                if existing_transfer_lines:
                    db.session.flush()
                post_err = post_balanced_journal(
                    lines,
                    journal_id=journal_id,
                    journal_memo=journal_memo,
                    posted_by='account_transfer',
                    source_table='AccountTransfer',
                )
                if post_err:
                    err.extend(post_err)
                else:
                    msg = f'Updated account transfer {tcode}.' if existing_transfer_lines else f'Recorded account transfer {tcode}.'
                    selected.update({
                        'transfer_date': datetime.date.today().strftime('%Y-%m-%d'),
                        'transfer_from_account_id': '',
                        'transfer_to_account_id': '',
                        'transfer_amount': '',
                        'transfer_ref': '',
                        'transfer_memo': '',
                        'owner_transfer_treatment': '',
                        'edit_transfer_journal_id': '',
                    })

    setup_warnings = []
    setup_changed = False
    owner_cash_account, owner_cash_note = ensure_owner_cash_account()
    if owner_cash_note:
        setup_changed = True
        setup_warnings.append(owner_cash_note)
    for book_company in allowed_companies:
        for target_company in allowed_companies:
            if book_company == target_company:
                continue
            account, note = ensure_due_account(book_company, target_company)
            if note:
                setup_changed = True
                setup_warnings.append(note)
    if setup_changed:
        db.session.commit()

    bank_accounts = Accounts.query.filter(
        (Accounts.Co.in_(allowed_companies)) &
        (Accounts.Type.in_(['Bank', 'Asset', 'Exch']))
    ).order_by(Accounts.Co, Accounts.Name).all()
    operating_accounts = Accounts.query.filter(
        (Accounts.Co.in_(allowed_companies)) &
        (Accounts.Type.in_(['Income', 'Expense']))
    ).order_by(Accounts.Co, Accounts.Type, Accounts.Name).all()
    transfer_account_candidates = Accounts.query.filter(
        Accounts.Co.in_(allowed_companies)
    ).order_by(Accounts.Co, Accounts.Type, Accounts.Name).all()
    transfer_accounts = [
        account for account in transfer_account_candidates
        if is_transfer_endpoint_account(account)
    ]

    bank_companies = {account.Co for account in bank_accounts}
    for code in allowed_companies:
        if code not in bank_companies:
            setup_warnings.append(f'No cash/bank account is configured for {company_label(code)}.')

    raw_recent = Gledger.query.filter(
        Gledger.SourceTable.in_(['IntercompanyEntry', 'AccountTransfer'])
    ).order_by(Gledger.Date.desc(), Gledger.id.desc()).limit(120).all()
    journal_map = {}
    for line in raw_recent:
        key = line.JournalId or line.Tcode
        item = journal_map.setdefault(key, {
            'journal_id': key,
            'entry_kind': 'Transfer' if line.SourceTable == 'AccountTransfer' else 'Intercompany',
            'source_table': line.SourceTable,
            'date': line.Date,
            'tcode': line.Tcode,
            'memo': line.JournalMemo,
            'ref': line.Ref,
            'source': line.Source,
            'companies': set(),
            'debit': 0,
            'credit': 0,
            'bank_amount': 0,
            'reconciled': False,
        })
        if line.SourceTable == 'AccountTransfer':
            item['entry_kind'] = 'Transfer'
        item['companies'].add(line.Com)
        item['debit'] += line.Debit or 0
        item['credit'] += line.Credit or 0
        if line.Reconciled not in [None, 0, 25]:
            item['reconciled'] = True
        if line.Type in ['PC', 'DD', 'XD', 'XC']:
            item['bank_amount'] = (line.Debit or line.Credit or 0)
    recent_entries = sorted(journal_map.values(), key=lambda item: item['date'] or datetime.datetime.min, reverse=True)
    for item in recent_entries:
        item['companies'] = ', '.join(sorted(item['companies']))
        item['amount_fmt'] = money(item['bank_amount'] or max(item['debit'], item['credit']))
        item['balanced'] = item['debit'] == item['credit']

    due_accounts = Accounts.query.filter(
        (Accounts.Co.in_(allowed_companies)) &
        (Accounts.Name.contains('Due'))
    ).order_by(Accounts.Co, Accounts.Name).all()
    due_balances = []
    for account in due_accounts:
        totals = db.session.query(
            func.coalesce(func.sum(Gledger.Debit), 0),
            func.coalesce(func.sum(Gledger.Credit), 0),
        ).filter(Gledger.Aid == account.id).first()
        debit_total = int(totals[0] or 0)
        credit_total = int(totals[1] or 0)
        due_balances.append({
            'company': account.Co,
            'account': account.Name,
            'debit': money(debit_total),
            'credit': money(credit_total),
            'net_credit': money(credit_total - debit_total),
        })

    return render_template(
        'intercompany_entries.html',
        cmpdata=cmpdata,
        scac=scac,
        company_names=company_names,
        bank_accounts=bank_accounts,
        operating_accounts=operating_accounts,
        transfer_accounts=transfer_accounts,
        selected=selected,
        err='\n'.join(err),
        msg=msg,
        setup_warnings=setup_warnings,
        recent_entries=recent_entries[:40],
        due_balances=due_balances,
    )


@main.route('/GeneralLedger', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def GeneralLedger():
    if session.get('authority') not in ['admin', 'superuser']:
        abort(403)

    def parse_date(value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return None

    def money(value):
        return "${:,.2f}".format((value or 0) / 100)

    def open_reconciliation_value(value):
        return value in [None, 0, 25]

    def selected_ledger_lines(selection_key):
        if not selection_key:
            return []
        if selection_key.startswith('legacy-'):
            try:
                line_id = int(selection_key.replace('legacy-', '', 1))
            except ValueError:
                return []
            line = Gledger.query.get(line_id)
            return [line] if line is not None else []
        return Gledger.query.filter(
            Gledger.JournalId == selection_key
        ).order_by(Gledger.JournalSeq.asc(), Gledger.id.asc()).all()

    def line_is_reconciled(line):
        return not open_reconciliation_value(line.Reconciled)

    def format_selected_lines(lines):
        for line in lines:
            line.debit_fmt = money(line.Debit)
            line.credit_fmt = money(line.Credit)
        return lines

    def bill_for_ledger_line(line):
        if line.SourceTable == 'Bills' and line.SourceId:
            bill = Bills.query.get(line.SourceId)
            if bill is not None:
                return bill
        if line.Tcode:
            return Bills.query.filter(Bills.Jo == line.Tcode).first()
        return None

    def bills_for_ledger_lines(lines):
        bills = []
        for line in lines:
            bill = bill_for_ledger_line(line)
            if bill is not None and bill.id not in [item.id for item in bills]:
                bills.append(bill)
        return bills

    def bill_related_ledger_lines(bills):
        filters_for_bills = []
        for bill in bills:
            if not bill or not bill.Jo:
                continue
            filters_for_bills.extend([
                Gledger.Tcode == bill.Jo,
                Gledger.Tcode.like(f'{bill.Jo}-%'),
                Gledger.JournalId == f'PAYBILL-{bill.Jo}',
                Gledger.JournalId.like(f'PAYBILL-{bill.Jo}-%'),
                (Gledger.SourceTable == 'Bills') & (Gledger.SourceId == bill.id),
            ])
        if not filters_for_bills:
            return []
        return Gledger.query.filter(or_(*filters_for_bills)).order_by(
            Gledger.Date.asc(),
            Gledger.JournalId.asc(),
            Gledger.JournalSeq.asc(),
            Gledger.id.asc(),
        ).all()

    def ledger_source_tables(lines):
        return {line.SourceTable for line in lines if line.SourceTable}

    def has_bill_payment_lines(lines):
        return any(line.Type in ['PC', 'PD', 'QD', 'QC'] or (line.JournalId or '').startswith('PAYBILL-') for line in lines)

    def selected_review_context(lines):
        source_tables = ledger_source_tables(lines)
        bills = bills_for_ledger_lines(lines)
        review_lines = bill_related_ledger_lines(bills) if bills else lines
        review_source_tables = ledger_source_tables(review_lines)
        receive_payment_locked = bool(review_source_tables.intersection({
            'ReceivePaymentBatch',
            'ReceivePaymentAllocation',
            'CounterDepositItem',
        }))
        can_delete_payment = bool(
            lines and not receive_payment_locked and (
                ('Bills' in source_tables and has_bill_payment_lines(lines)) or
                source_tables == {'PayrollBatch'} or
                (source_tables and source_tables.issubset({'ManualDeposit', 'AccountTransfer', 'IntercompanyEntry'})) or
                (not source_tables and has_bill_payment_lines(lines))
            )
        )
        can_delete_bill = bool(bills and not receive_payment_locked)
        return {
            'lines': review_lines,
            'bills': bills,
            'can_delete_payment': can_delete_payment,
            'can_delete_bill': can_delete_bill,
            'receive_payment_locked': receive_payment_locked,
            'all_reconciled_locked': any(line_is_reconciled(line) for line in review_lines),
        }

    def reset_bill_payment(bill):
        bill.Status = None
        bill.pDate = None
        bill.pDate2 = None
        bill.pAmount = None
        bill.pAmount2 = None
        bill.pMulti = None
        bill.pAccount = None
        bill.Check = None
        bill.PmtList = None
        bill.PacctList = None
        bill.RefList = None
        bill.MemoList = None
        bill.PdateList = None
        bill.CheckList = None
        bill.MethList = None
        bill.Pcache = None
        bill.pMeth = None
        bill.QBi = None

    def delete_selected_ledger_lines(lines, delete_scope='payment'):
        if not lines:
            return False, 'Choose a journal to delete.'
        reconciled_lines = [line for line in lines if line_is_reconciled(line)]
        if reconciled_lines:
            return False, 'This ledger transaction has been reconciled. Reopen the reconciliation statement before editing or deleting it.'

        source_tables = ledger_source_tables(lines)
        journal_id = lines[0].JournalId

        if source_tables == {'PayrollBatch'}:
            for bill in Bills.query.filter(
                (Bills.Temp1 == 'PayrollBatch') &
                (Bills.Temp2 == journal_id)
            ).all():
                db.session.delete(bill)
            for line in lines:
                db.session.delete(line)
            db.session.commit()
            return True, 'Deleted the payroll batch ledger journal and payroll bill records.'

        if source_tables and source_tables.issubset({'ManualDeposit', 'AccountTransfer', 'IntercompanyEntry'}):
            for line in lines:
                db.session.delete(line)
            db.session.commit()
            return True, 'Deleted the selected ledger journal.'

        if 'Bills' in source_tables or (not source_tables and has_bill_payment_lines(lines)):
            bills = bills_for_ledger_lines(lines)
            if not bills:
                return False, 'The bill payment could not be matched to a bill record, so no changes were made.'
            if delete_scope == 'bill':
                all_bill_lines = bill_related_ledger_lines(bills)
                reconciled_bill_lines = [line for line in all_bill_lines if line_is_reconciled(line)]
                if reconciled_bill_lines:
                    return False, 'One or more related bill ledger rows have been reconciled. Reopen the reconciliation statement before deleting the bill.'
                for line in all_bill_lines:
                    db.session.delete(line)
                for bill in bills:
                    db.session.delete(bill)
                db.session.commit()
                return True, f'Deleted {len(bills)} bill item(s) and all related bill/payment ledger rows.'
            if not has_bill_payment_lines(lines):
                return False, 'Only bill payment ledger journals can be deleted here. Bill entry/accrual rows should be corrected from Bill Payments.'
            for bill in bills:
                reset_bill_payment(bill)
            for line in lines:
                db.session.delete(line)
            db.session.commit()
            return True, f'Deleted the bill payment journal and reopened {len(bills)} bill item(s).'

        source_name = ', '.join(sorted(source_tables)) if source_tables else 'legacy ledger rows'
        return False, f'{source_name} is review-only from General Ledger. Correct or delete it from its source workflow.'

    err = []
    msg = []
    selected_key = ''
    selected_lines = []
    selected_context = None
    selected_summary = None

    filters = {
        'account': request.values.get('account', '').strip(),
        'tcode': request.values.get('tcode', '').strip(),
        'journal_id': request.values.get('journal_id', '').strip(),
        'com': request.values.get('com', '').strip(),
        'date_from': request.values.get('date_from', '').strip(),
        'date_to': request.values.get('date_to', '').strip(),
        'unbalanced': request.values.get('unbalanced', ''),
        'limit': request.values.get('limit', '500').strip() or '500',
    }

    if request.method == 'POST':
        selections = [item for item in request.form.getlist('selected_journal_key') if item]
        if len(selections) != 1:
            err.append('Choose one ledger journal to review or delete.')
        else:
            selected_key = selections[0]
            selected_lines = selected_ledger_lines(selected_key)
            if request.form.get('delete_bill_and_payment') is not None:
                ok, text = delete_selected_ledger_lines(selected_lines, delete_scope='bill')
                if ok:
                    msg.append(text)
                    selected_key = ''
                    selected_lines = []
                else:
                    err.append(text)
            elif request.form.get('delete_journal') is not None:
                ok, text = delete_selected_ledger_lines(selected_lines, delete_scope='payment')
                if ok:
                    msg.append(text)
                    selected_key = ''
                    selected_lines = []
                else:
                    err.append(text)
            elif request.form.get('review_journal') is not None:
                if selected_lines:
                    msg.append('Loaded the selected ledger journal for review.')
                else:
                    err.append('The selected ledger journal could not be found.')

    try:
        limit = min(max(int(filters['limit']), 1), 5000)
    except ValueError:
        limit = 500
        filters['limit'] = '500'

    date_from = parse_date(filters['date_from'])
    date_to = parse_date(filters['date_to'])
    if date_to:
        date_to = date_to + datetime.timedelta(days=1)

    query = Gledger.query
    if filters['account']:
        query = query.filter(Gledger.Account == filters['account'])
    if filters['tcode']:
        query = query.filter(Gledger.Tcode.ilike(f"%{filters['tcode']}%"))
    if filters['journal_id']:
        query = query.filter(Gledger.JournalId.ilike(f"%{filters['journal_id']}%"))
    if filters['com']:
        query = query.filter(Gledger.Com == filters['com'])
    if date_from:
        query = query.filter(Gledger.Date >= date_from)
    if date_to:
        query = query.filter(Gledger.Date < date_to)
    if filters['unbalanced']:
        unbalanced_ids = [
            item[0] for item in db.session.query(Gledger.JournalId)
            .filter(Gledger.JournalId.isnot(None))
            .group_by(Gledger.JournalId)
            .having(func.coalesce(func.sum(Gledger.Debit), 0) != func.coalesce(func.sum(Gledger.Credit), 0))
            .all()
        ]
        if unbalanced_ids:
            query = query.filter(Gledger.JournalId.in_(unbalanced_ids))
        else:
            query = query.filter(Gledger.id == -1)

    entries = query.order_by(Gledger.Date.desc(), Gledger.id.desc()).limit(limit).all()

    debit_total = sum((entry.Debit or 0) for entry in entries)
    credit_total = sum((entry.Credit or 0) for entry in entries)
    journal_map = {}
    for entry in entries:
        key = entry.JournalId or f"legacy-{entry.id}"
        item = journal_map.setdefault(key, {
            'selection_key': key,
            'journal_id': entry.JournalId or '(legacy row)',
            'date': entry.Date,
            'memo': entry.JournalMemo or '',
            'source': entry.SourceTable or entry.Source or '',
            'source_id': entry.SourceId or entry.Sid or '',
            'debit': 0,
            'credit': 0,
            'rows': 0,
            'reconciled': False,
        })
        item['debit'] += entry.Debit or 0
        item['credit'] += entry.Credit or 0
        item['rows'] += 1
        if line_is_reconciled(entry):
            item['reconciled'] = True
        if entry.Date and (not item['date'] or entry.Date > item['date']):
            item['date'] = entry.Date

    journal_summaries = sorted(
        journal_map.values(),
        key=lambda item: item['date'] or datetime.datetime.min,
        reverse=True,
    )
    for item in journal_summaries:
        item['balanced'] = item['debit'] == item['credit']
        item['debit_fmt'] = money(item['debit'])
        item['credit_fmt'] = money(item['credit'])
        item['variance_fmt'] = money(item['debit'] - item['credit'])

    for entry in entries:
        entry.debit_fmt = money(entry.Debit)
        entry.credit_fmt = money(entry.Credit)

    if selected_lines:
        selected_context = selected_review_context(selected_lines)
        selected_review_lines = selected_context['lines']
    else:
        selected_review_lines = []
    selected_lines = format_selected_lines(selected_review_lines)
    if selected_lines:
        selected_debit = sum(line.Debit or 0 for line in selected_lines)
        selected_credit = sum(line.Credit or 0 for line in selected_lines)
        selected_summary = {
            'journal_id': selected_lines[0].JournalId or '(legacy row)',
            'rows': len(selected_lines),
            'debit': money(selected_debit),
            'credit': money(selected_credit),
            'variance': money(selected_debit - selected_credit),
            'balanced': selected_debit == selected_credit,
            'reconciled': any(line_is_reconciled(line) for line in selected_lines),
            'can_delete_payment': bool(selected_context and selected_context['can_delete_payment']),
            'can_delete_bill': bool(selected_context and selected_context['can_delete_bill']),
            'receive_payment_locked': bool(selected_context and selected_context['receive_payment_locked']),
            'all_reconciled_locked': bool(selected_context and selected_context['all_reconciled_locked']),
            'bill_count': len(selected_context['bills']) if selected_context else 0,
        }

    accounts = Accounts.query.order_by(Accounts.Name).all()
    companies = [
        item[0] for item in db.session.query(Gledger.Com)
        .filter(Gledger.Com.isnot(None))
        .distinct()
        .order_by(Gledger.Com)
        .all()
    ]

    totals = {
        'debit': money(debit_total),
        'credit': money(credit_total),
        'variance': money(debit_total - credit_total),
        'balanced': debit_total == credit_total,
        'row_count': len(entries),
        'limit': limit,
    }

    return render_template(
        'general_ledger.html',
        cmpdata=cmpdata,
        scac=scac,
        entries=entries,
        accounts=accounts,
        companies=companies,
        filters=filters,
        totals=totals,
        journal_summaries=journal_summaries,
        selected_key=selected_key,
        selected_lines=selected_lines,
        selected_summary=selected_summary,
        err=err,
        msg=msg,
    )


@main.route('/IncomeExpenseReview', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def IncomeExpenseReview():
    def parse_date(value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return None

    def money(cents):
        return "${:,.2f}".format((cents or 0) / 100)

    def money_number(cents):
        return float(cents or 0) / 100

    def build_rollup_rows(account_summaries):
        rows = []
        for section in ['Income', 'Expense']:
            section_items = [item for item in account_summaries if item['type'] == section]
            if not section_items:
                continue
            section_total = sum(item['net'] for item in section_items)
            rows.append({'level': 0, 'label': section.upper(), 'amount': section_total, 'kind': 'section'})
            categories = sorted({item['category'] or '(No Category)' for item in section_items})
            for category in categories:
                category_items = [item for item in section_items if (item['category'] or '(No Category)') == category]
                category_total = sum(item['net'] for item in category_items)
                rows.append({'level': 1, 'label': category, 'amount': category_total, 'kind': 'category'})
                subcategories = sorted({item['subcategory'] or '(No Subcategory)' for item in category_items})
                for subcategory in subcategories:
                    subcategory_items = [item for item in category_items if (item['subcategory'] or '(No Subcategory)') == subcategory]
                    subcategory_total = sum(item['net'] for item in subcategory_items)
                    rows.append({'level': 2, 'label': subcategory, 'amount': subcategory_total, 'kind': 'subcategory'})
                    for item in sorted(subcategory_items, key=lambda row: row['account']):
                        rows.append({'level': 3, 'label': item['account'], 'amount': item['net'], 'kind': 'account'})
            rows.append({'level': 1, 'label': f'Total {section}', 'amount': section_total, 'kind': 'total'})
            rows.append({'level': 0, 'label': '', 'amount': None, 'kind': 'blank'})
        income_total = sum(item['net'] for item in account_summaries if item['type'] == 'Income')
        expense_total = sum(item['net'] for item in account_summaries if item['type'] == 'Expense')
        rows.append({'level': 0, 'label': 'NET INCOME', 'amount': income_total - expense_total, 'kind': 'net'})
        return rows

    def export_rollup_xlsx(rollup_rows, filters):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Income Expense'
        ws['A1'] = 'Income & Expense Review'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Period: {filters['date_from']} to {filters['date_to']}"
        ws['A3'] = f"Company: {filters['com'] or 'All'}"
        ws.append([])
        ws.append(['Description', 'Amount'])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9EAF7')

        for row in rollup_rows:
            label = ('    ' * row['level']) + row['label']
            ws.append([label, money_number(row['amount']) if row['amount'] is not None else None])
            excel_row = ws.max_row
            if row['kind'] in ['section', 'category', 'total', 'net']:
                ws.cell(excel_row, 1).font = Font(bold=True)
                ws.cell(excel_row, 2).font = Font(bold=True)
            if row['kind'] == 'section':
                ws.cell(excel_row, 1).fill = PatternFill('solid', fgColor='E2F0D9')
                ws.cell(excel_row, 2).fill = PatternFill('solid', fgColor='E2F0D9')
            if row['kind'] == 'net':
                ws.cell(excel_row, 1).fill = PatternFill('solid', fgColor='FFF2CC')
                ws.cell(excel_row, 2).fill = PatternFill('solid', fgColor='FFF2CC')
            if row['amount'] is not None:
                ws.cell(excel_row, 2).number_format = '$#,##0.00;[Red]($#,##0.00)'

        ws.column_dimensions['A'].width = 52
        ws.column_dimensions['B'].width = 18
        for row in ws.iter_rows(min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='right')
        wb.save(output)
        output.seek(0)
        filename = f"income_expense_{filters['com'] or 'all'}_{filters['date_from']}_{filters['date_to']}.xlsx".replace('/', '-')
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def export_rollup_pdf(rollup_rows, filters):
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        story = [
            Paragraph('Income & Expense Review', styles['Title']),
            Paragraph(f"Period: {filters['date_from']} to {filters['date_to']} &nbsp;&nbsp; Company: {filters['com'] or 'All'}", styles['Normal']),
            Spacer(1, 12),
        ]
        table_data = [['Description', 'Amount']]
        row_styles = []
        for row in rollup_rows:
            table_data.append([('    ' * row['level']) + row['label'], money(row['amount']) if row['amount'] is not None else ''])
            idx = len(table_data) - 1
            if row['kind'] in ['section', 'category', 'total', 'net']:
                row_styles.append(('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'))
            if row['kind'] == 'section':
                row_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#E2F0D9')))
            if row['kind'] == 'net':
                row_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFF2CC')))
        table = Table(table_data, colWidths=[380, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9EAF7')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), .25, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ] + row_styles))
        story.append(table)
        doc.build(story)
        output.seek(0)
        filename = f"income_expense_{filters['com'] or 'all'}_{filters['date_from']}_{filters['date_to']}.pdf".replace('/', '-')
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

    today_value = datetime.date.today()
    preset = request.values.get('preset', 'this_year')
    custom_from = request.values.get('date_from', '').strip()
    custom_to = request.values.get('date_to', '').strip()

    if preset == 'this_month':
        date_from = datetime.datetime(today_value.year, today_value.month, 1)
        date_to = datetime.datetime.combine(today_value, datetime.time.min) + datetime.timedelta(days=1)
    elif preset == 'last_month':
        first_this_month = datetime.date(today_value.year, today_value.month, 1)
        last_month_end = first_this_month - datetime.timedelta(days=1)
        date_from = datetime.datetime(last_month_end.year, last_month_end.month, 1)
        date_to = datetime.datetime.combine(first_this_month, datetime.time.min)
    elif preset == 'last_year':
        date_from = datetime.datetime(today_value.year - 1, 1, 1)
        date_to = datetime.datetime(today_value.year, 1, 1)
    elif preset == 'custom':
        date_from = parse_date(custom_from)
        date_to = parse_date(custom_to)
        if date_to:
            date_to = date_to + datetime.timedelta(days=1)
    else:
        preset = 'this_year'
        date_from = datetime.datetime(today_value.year, 1, 1)
        date_to = datetime.datetime.combine(today_value, datetime.time.min) + datetime.timedelta(days=1)

    company_filter = request.values.get('com', cmpdata[10]).strip()
    account_type_filter = request.values.get('account_type', 'all').strip()

    account_rows = Accounts.query.filter(Accounts.Type.in_(['Income', 'Expense']))
    if company_filter:
        account_rows = account_rows.filter(Accounts.Co == company_filter)
    account_rows = account_rows.order_by(Accounts.Type, Accounts.Name).all()
    account_by_name = {account.Name: account for account in account_rows}
    account_names = list(account_by_name.keys())

    query = Gledger.query
    if company_filter:
        query = query.filter(Gledger.Com == company_filter)
    if date_from:
        query = query.filter(Gledger.Date >= date_from)
    if date_to:
        query = query.filter(Gledger.Date < date_to)
    if account_names:
        query = query.filter(Gledger.Account.in_(account_names))
    else:
        query = query.filter(Gledger.id == -1)
    if account_type_filter in ['Income', 'Expense']:
        filtered_names = [
            account.Name for account in account_rows
            if account.Type == account_type_filter
        ]
        query = query.filter(Gledger.Account.in_(filtered_names if filtered_names else ['__none__']))

    entries = query.order_by(Gledger.Date.desc(), Gledger.id.desc()).all()

    grouped = {}
    income_total = 0
    expense_total = 0
    for entry in entries:
        account = account_by_name.get(entry.Account)
        if account is None:
            continue
        group = grouped.setdefault(entry.Account, {
            'account': entry.Account,
            'type': account.Type,
            'category': account.Category or '',
            'subcategory': account.Subcategory or '',
            'debit': 0,
            'credit': 0,
            'net': 0,
            'rows': 0,
        })
        debit = entry.Debit or 0
        credit = entry.Credit or 0
        group['debit'] += debit
        group['credit'] += credit
        group['rows'] += 1
        if account.Type == 'Income':
            net = credit - debit
            income_total += net
        else:
            net = debit - credit
            expense_total += net

    for group in grouped.values():
        if group['type'] == 'Income':
            group['net'] = group['credit'] - group['debit']
        else:
            group['net'] = group['debit'] - group['credit']
        group['debit_fmt'] = money(group['debit'])
        group['credit_fmt'] = money(group['credit'])
        group['net_fmt'] = money(group['net'])

    account_summaries = sorted(
        grouped.values(),
        key=lambda item: (item['type'], item['category'], item['account'])
    )

    category_map = {}
    for item in account_summaries:
        key = (
            item['type'],
            item['category'] or '(No Category)',
            item['subcategory'] or '',
        )
        category = category_map.setdefault(key, {
            'type': item['type'],
            'category': item['category'] or '(No Category)',
            'subcategory': item['subcategory'] or '',
            'debit': 0,
            'credit': 0,
            'net': 0,
            'accounts': 0,
            'rows': 0,
        })
        category['debit'] += item['debit']
        category['credit'] += item['credit']
        category['net'] += item['net']
        category['accounts'] += 1
        category['rows'] += item['rows']

    category_summaries = sorted(
        category_map.values(),
        key=lambda item: (item['type'], item['category'], item['subcategory'])
    )
    for category in category_summaries:
        category['debit_fmt'] = money(category['debit'])
        category['credit_fmt'] = money(category['credit'])
        category['net_fmt'] = money(category['net'])

    for entry in entries:
        account = account_by_name.get(entry.Account)
        entry.account_type = account.Type if account is not None else ''
        entry.account_category = account.Category if account is not None else ''
        entry.debit_fmt = money(entry.Debit)
        entry.credit_fmt = money(entry.Credit)
        if account is not None and account.Type == 'Income':
            entry.net_fmt = money((entry.Credit or 0) - (entry.Debit or 0))
        else:
            entry.net_fmt = money((entry.Debit or 0) - (entry.Credit or 0))

    companies = [
        item[0] for item in db.session.query(Gledger.Com)
        .filter(Gledger.Com.isnot(None))
        .distinct()
        .order_by(Gledger.Com)
        .all()
    ]

    filters = {
        'preset': preset,
        'date_from': custom_from or (date_from.strftime('%Y-%m-%d') if date_from else ''),
        'date_to': custom_to or ((date_to - datetime.timedelta(days=1)).strftime('%Y-%m-%d') if date_to else ''),
        'com': company_filter,
        'account_type': account_type_filter,
    }
    totals = {
        'income': money(income_total),
        'expenses': money(expense_total),
        'net_income': money(income_total - expense_total),
        'net_income_positive': income_total >= expense_total,
        'row_count': len(entries),
        'account_count': len(account_summaries),
    }

    rollup_rows = build_rollup_rows(account_summaries)
    export_type = request.values.get('export', '').strip().lower()
    if export_type == 'xlsx':
        return export_rollup_xlsx(rollup_rows, filters)
    if export_type == 'pdf':
        return export_rollup_pdf(rollup_rows, filters)

    return render_template(
        'income_expense_review.html',
        cmpdata=cmpdata,
        scac=scac,
        filters=filters,
        totals=totals,
        companies=companies,
        category_summaries=category_summaries,
        account_summaries=account_summaries,
        entries=entries,
    )


@main.route('/BalanceSheet', methods=['GET'])
@login_required
@financial_mfa_required
def BalanceSheet():
    def parse_date(value):
        if not value:
            return datetime.datetime.combine(datetime.date.today(), datetime.time.min)
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return datetime.datetime.combine(datetime.date.today(), datetime.time.min)

    def money(cents):
        return "${:,.2f}".format((cents or 0) / 100)

    def account_section(account):
        account_type = account.Type or ''
        category = account.Category or ''
        if account_type in ['Bank', 'Asset', 'Exch'] or category == 'Assets':
            return 'Assets'
        if 'Liability' in account_type or category == 'Liabilities':
            return 'Liabilities'
        if account_type == 'Equity' or category == 'Equity':
            return 'Equity'
        return 'Other Balance Accounts'

    def account_balance(account, debit, credit):
        section = account_section(account)
        if section in ['Liabilities', 'Equity']:
            return credit - debit
        return debit - credit

    today_value = datetime.date.today().strftime('%Y-%m-%d')
    as_of_text = request.values.get('as_of', today_value).strip() or today_value
    as_of_date = parse_date(as_of_text)
    as_of_next = as_of_date + datetime.timedelta(days=1)
    company_filter = request.values.get('com', cmpdata[10]).strip()

    all_account_rows = Accounts.query
    if company_filter:
        all_account_rows = all_account_rows.filter(Accounts.Co == company_filter)
    all_account_rows = all_account_rows.order_by(Accounts.Co, Accounts.Type, Accounts.Name).all()
    all_accounts_by_id = {account.id: account for account in all_account_rows}
    all_accounts_by_company_name = {(account.Co, account.Name): account for account in all_account_rows}
    balance_accounts_by_id = {
        account.id: account for account in all_account_rows
        if account.Type not in ['Income', 'Expense']
    }
    balance_accounts_by_company_name = {
        (account.Co, account.Name): account for account in all_account_rows
        if account.Type not in ['Income', 'Expense']
    }

    ledger_query = Gledger.query.filter(Gledger.Date < as_of_next)
    if company_filter:
        ledger_query = ledger_query.filter(Gledger.Com == company_filter)
    entries = ledger_query.all()

    grouped = {}
    income_total = 0
    expense_total = 0
    row_count = 0
    for entry in entries:
        account = balance_accounts_by_id.get(entry.Aid) or balance_accounts_by_company_name.get((entry.Com, entry.Account))
        debit = entry.Debit or 0
        credit = entry.Credit or 0

        if account is None:
            income_expense_account = all_accounts_by_id.get(entry.Aid) or all_accounts_by_company_name.get((entry.Com, entry.Account))
            if income_expense_account is None:
                continue
            if income_expense_account.Type == 'Income':
                income_total += credit - debit
            else:
                expense_total += debit - credit
            continue

        key = (account.Co, account.Name)
        group = grouped.setdefault(key, {
            'company': account.Co,
            'account': account.Name,
            'type': account.Type or '',
            'category': account.Category or '',
            'subcategory': account.Subcategory or '',
            'section': account_section(account),
            'debit': 0,
            'credit': 0,
            'balance': 0,
            'rows': 0,
        })
        group['debit'] += debit
        group['credit'] += credit
        group['rows'] += 1
        row_count += 1

    for group in grouped.values():
        account = balance_accounts_by_company_name.get((group['company'], group['account']))
        group['balance'] = account_balance(account, group['debit'], group['credit'])

    current_earnings = income_total - expense_total
    if current_earnings:
        grouped[('__earnings__', company_filter or 'All')] = {
            'company': company_filter or 'All',
            'account': 'Current Earnings',
            'type': 'Equity',
            'category': 'Equity',
            'subcategory': '',
            'section': 'Equity',
            'debit': expense_total,
            'credit': income_total,
            'balance': current_earnings,
            'rows': 0,
        }

    section_order = ['Assets', 'Liabilities', 'Equity', 'Other Balance Accounts']
    account_summaries = sorted(
        grouped.values(),
        key=lambda item: (section_order.index(item['section']) if item['section'] in section_order else 99,
                          item['company'], item['category'], item['account'])
    )
    for item in account_summaries:
        item['debit_fmt'] = money(item['debit'])
        item['credit_fmt'] = money(item['credit'])
        item['balance_fmt'] = money(item['balance'])

    section_summaries = []
    for section in section_order:
        section_items = [item for item in account_summaries if item['section'] == section]
        if not section_items:
            continue
        total = sum(item['balance'] for item in section_items)
        section_summaries.append({
            'section': section,
            'balance': total,
            'balance_fmt': money(total),
            'accounts': len(section_items),
        })

    category_map = {}
    for item in account_summaries:
        key = (item['section'], item['category'] or '(No Category)', item['subcategory'] or '')
        category = category_map.setdefault(key, {
            'section': item['section'],
            'category': item['category'] or '(No Category)',
            'subcategory': item['subcategory'] or '',
            'balance': 0,
            'accounts': 0,
        })
        category['balance'] += item['balance']
        category['accounts'] += 1
    category_summaries = sorted(
        category_map.values(),
        key=lambda item: (section_order.index(item['section']) if item['section'] in section_order else 99,
                          item['category'], item['subcategory'])
    )
    for item in category_summaries:
        item['balance_fmt'] = money(item['balance'])

    assets_total = sum(item['balance'] for item in account_summaries if item['section'] == 'Assets')
    liabilities_total = sum(item['balance'] for item in account_summaries if item['section'] == 'Liabilities')
    equity_total = sum(item['balance'] for item in account_summaries if item['section'] == 'Equity')
    other_total = sum(item['balance'] for item in account_summaries if item['section'] == 'Other Balance Accounts')
    liabilities_equity_total = liabilities_total + equity_total
    difference = assets_total - liabilities_equity_total - other_total

    companies = [
        item[0] for item in db.session.query(Gledger.Com)
        .filter(Gledger.Com.isnot(None))
        .distinct()
        .order_by(Gledger.Com)
        .all()
    ]

    filters = {
        'as_of': as_of_date.strftime('%Y-%m-%d'),
        'com': company_filter,
    }
    totals = {
        'assets': money(assets_total),
        'liabilities': money(liabilities_total),
        'equity': money(equity_total),
        'liabilities_equity': money(liabilities_equity_total),
        'other': money(other_total),
        'difference': money(difference),
        'balanced': difference == 0,
        'row_count': row_count,
        'account_count': len(account_summaries),
    }

    return render_template(
        'balance_sheet.html',
        cmpdata=cmpdata,
        scac=scac,
        filters=filters,
        totals=totals,
        companies=companies,
        section_summaries=section_summaries,
        category_summaries=category_summaries,
        account_summaries=account_summaries,
    )


@main.route('/OpeningBalances', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def OpeningBalances():
    def parse_date(value):
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except:
            return None

    def money(cents):
        return "{:,.2f}".format((cents or 0) / 100)

    def money_to_cents(value):
        try:
            clean = str(value).replace('$', '').replace(',', '').strip()
            if clean in ['', '-']:
                return 0
            return int((Decimal(clean) * 100).quantize(Decimal('1')))
        except:
            return None

    def account_section(account):
        account_type = account.Type or ''
        category = account.Category or ''
        if account_type in ['Bank', 'Asset', 'Exch'] or category == 'Assets':
            return 'Assets'
        if 'Liability' in account_type or category == 'Liabilities':
            return 'Liabilities'
        if account_type == 'Equity' or category == 'Equity':
            return 'Equity'
        return 'Other Balance Accounts'

    def line_for_normal_balance(account, cents):
        section = account_section(account)
        if section in ['Liabilities', 'Equity']:
            debit = abs(cents) if cents < 0 else 0
            credit = cents if cents > 0 else 0
        else:
            debit = cents if cents > 0 else 0
            credit = abs(cents) if cents < 0 else 0
        return debit, credit

    def normal_balance_for_line(account, line):
        section = account_section(account)
        debit = line.Debit or 0
        credit = line.Credit or 0
        if section in ['Liabilities', 'Equity']:
            return credit - debit
        return debit - credit

    def ensure_opening_equity(company_code):
        account = Accounts.query.filter(
            (Accounts.Co == company_code) &
            (Accounts.Name == 'Opening Balance Equity')
        ).first()
        if account is not None:
            return account
        account = Accounts(
            Name='Opening Balance Equity',
            Balance=0.00,
            AcctNumber=None,
            Routing=None,
            Payee=None,
            Type='Equity',
            Description='Created automatically for opening balance journals',
            Category='Equity',
            Subcategory='Opening Balances',
            Taxrollup=None,
            Co=company_code,
            QBmap=None,
            Shared=None,
        )
        db.session.add(account)
        db.session.flush()
        return account

    today_value = datetime.date.today()
    default_date = datetime.date(today_value.year, 1, 1).strftime('%Y-%m-%d')
    selected = {
        'company': request.values.get('company', cmpdata[10]).strip(),
        'opening_date': request.values.get('opening_date', default_date).strip() or default_date,
    }
    err = []

    companies = [
        item[0] for item in db.session.query(Accounts.Co)
        .filter(Accounts.Co.isnot(None))
        .distinct()
        .order_by(Accounts.Co)
        .all()
    ]
    if selected['company'] not in companies and companies:
        selected['company'] = companies[0]

    opening_date = parse_date(selected['opening_date'])
    journal_id = f"OPENBAL-{selected['company']}-{selected['opening_date'].replace('-', '')}"

    accounts = Accounts.query.filter(
        (Accounts.Co == selected['company']) &
        (~Accounts.Type.in_(['Income', 'Expense'])) &
        (Accounts.Name != 'Opening Balance Equity')
    ).order_by(Accounts.Type, Accounts.Name).all()

    existing_lines = Gledger.query.filter(
        (Gledger.JournalId == journal_id) &
        (Gledger.SourceTable == 'OpeningBalance') &
        (Gledger.Com == selected['company'])
    ).all()
    existing_by_aid = {line.Aid: line for line in existing_lines}

    if request.method == 'POST' and request.values.get('save_opening_balances') is not None:
        if opening_date is None:
            err.append('Opening balance date is invalid.')

        reconciled_lines = [line for line in existing_lines if line.Reconciled not in [None, 0, 25]]
        if reconciled_lines:
            err.append('Cannot replace these opening balances because at least one line has already been reconciled.')

        lines = []
        debit_total = 0
        credit_total = 0
        recorded = datetime.datetime.now()
        if not err:
            for index, account in enumerate(accounts, start=1):
                raw_value = request.values.get(f'balance_{account.id}', '')
                amount = money_to_cents(raw_value)
                if amount is None:
                    err.append(f'Opening balance for {account.Name} is invalid.')
                    continue
                if amount == 0:
                    continue
                debit, credit = line_for_normal_balance(account, amount)
                debit_total += debit
                credit_total += credit
                lines.append(Gledger(
                    Debit=debit,
                    Credit=credit,
                    Account=account.Name,
                    Aid=account.id,
                    Source='Opening Balance',
                    Sid=0,
                    Type='DD' if debit else 'PC',
                    Tcode=f'OB{selected["company"]}',
                    Com=selected['company'],
                    Recorded=recorded,
                    Reconciled=0,
                    Date=opening_date,
                    Ref='Opening Balance',
                    JournalId=journal_id,
                    JournalSeq=index,
                    JournalMemo=f'Opening balances as of {selected["opening_date"]}',
                    PostedBy='opening_balances',
                    PostedAt=recorded,
                    SourceTable='OpeningBalance',
                    SourceId=account.id,
                ))

        if not err and lines:
            equity = ensure_opening_equity(selected['company'])
            difference = debit_total - credit_total
            if difference > 0:
                equity_debit, equity_credit = 0, difference
            elif difference < 0:
                equity_debit, equity_credit = abs(difference), 0
            else:
                equity_debit, equity_credit = 0, 0
            if equity_debit or equity_credit:
                lines.append(Gledger(
                    Debit=equity_debit,
                    Credit=equity_credit,
                    Account=equity.Name,
                    Aid=equity.id,
                    Source='Opening Balance',
                    Sid=0,
                    Type='OE',
                    Tcode=f'OB{selected["company"]}',
                    Com=selected['company'],
                    Recorded=recorded,
                    Reconciled=0,
                    Date=opening_date,
                    Ref='Opening Balance',
                    JournalId=journal_id,
                    JournalSeq=len(lines) + 1,
                    JournalMemo=f'Opening balances as of {selected["opening_date"]}',
                    PostedBy='opening_balances',
                    PostedAt=recorded,
                    SourceTable='OpeningBalance',
                    SourceId=equity.id,
                ))

            for line in existing_lines:
                db.session.delete(line)
            for line in lines:
                db.session.add(line)
            db.session.commit()
            err.append(f'Saved opening balance journal {journal_id} with {len(lines)} ledger row(s).')
            existing_lines = Gledger.query.filter(
                (Gledger.JournalId == journal_id) &
                (Gledger.SourceTable == 'OpeningBalance') &
                (Gledger.Com == selected['company'])
            ).all()
            existing_by_aid = {line.Aid: line for line in existing_lines}
        elif not err:
            for line in existing_lines:
                db.session.delete(line)
            db.session.commit()
            err.append(f'Removed opening balance journal {journal_id}; all entered balances were zero.')
            existing_lines = []
            existing_by_aid = {}

    account_rows = []
    total_debits = 0
    total_credits = 0
    for account in accounts:
        existing_line = existing_by_aid.get(account.id)
        existing_balance = normal_balance_for_line(account, existing_line) if existing_line is not None else 0
        debit, credit = line_for_normal_balance(account, existing_balance)
        total_debits += debit
        total_credits += credit
        account_rows.append({
            'id': account.id,
            'name': account.Name,
            'type': account.Type or '',
            'category': account.Category or '',
            'section': account_section(account),
            'balance': money(existing_balance) if existing_balance else '',
        })

    existing_offset = next((line for line in existing_lines if line.Type == 'OE'), None)
    offset_balance = 0
    if existing_offset is not None:
        offset_balance = (existing_offset.Credit or 0) - (existing_offset.Debit or 0)

    totals = {
        'journal_id': journal_id,
        'debits': "${:,.2f}".format(total_debits / 100),
        'credits': "${:,.2f}".format(total_credits / 100),
        'offset': "${:,.2f}".format(offset_balance / 100),
        'rows': len(existing_lines),
    }

    return render_template(
        'opening_balances.html',
        cmpdata=cmpdata,
        scac=scac,
        companies=companies,
        selected=selected,
        account_rows=account_rows,
        totals=totals,
        err='\n'.join(err),
    )


@main.route('/CashFlowStatement', methods=['GET'])
@login_required
@financial_mfa_required
def CashFlowStatement():
    def parse_date(value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return None

    def money(cents):
        return "${:,.2f}".format((cents or 0) / 100)

    def classify_cash_flow(offset_accounts, source_table):
        names = ' '.join((account.Name or '') for account in offset_accounts).lower()
        types = {account.Type for account in offset_accounts if account is not None}
        categories = {account.Category for account in offset_accounts if account is not None}

        if not offset_accounts:
            return 'Transfers', 'Cash transfer or opening balance'
        if any(name in names for name in ['accounts receivable', 'accounts payable', 'payroll']):
            return 'Operating Activities', 'Receipts/payments from operations'
        if 'Income' in types or 'Expense' in types:
            return 'Operating Activities', 'Income and expense activity'
        if source_table in ['PayrollBatch', 'ManualDeposit']:
            return 'Operating Activities', source_table
        if any(account.Type in ['Asset'] and account.Type not in ['Bank', 'Exch'] for account in offset_accounts):
            return 'Investing Activities', 'Non-cash asset activity'
        if any(('Liability' in (account.Type or '')) or account.Type == 'Equity' for account in offset_accounts):
            return 'Financing Activities', 'Debt, equity, owner, or intercompany activity'
        return 'Operating Activities', 'Other cash activity'

    def cash_balance_before(cash_account_names, company_filter, cutoff_date):
        query = Gledger.query.filter(
            (Gledger.Account.in_(cash_account_names)) &
            (Gledger.Date < cutoff_date)
        )
        if company_filter:
            query = query.filter(Gledger.Com == company_filter)
        return sum((row.Debit or 0) - (row.Credit or 0) for row in query.all())

    today_value = datetime.date.today()
    default_from = datetime.date(today_value.year, 1, 1).strftime('%Y-%m-%d')
    default_to = today_value.strftime('%Y-%m-%d')
    filters = {
        'date_from': request.values.get('date_from', default_from).strip() or default_from,
        'date_to': request.values.get('date_to', default_to).strip() or default_to,
        'com': request.values.get('com', cmpdata[10]).strip(),
        'cash_account': request.values.get('cash_account', '').strip(),
    }
    date_from = parse_date(filters['date_from']) or datetime.datetime(today_value.year, 1, 1)
    date_to = parse_date(filters['date_to']) or datetime.datetime.combine(today_value, datetime.time.min)
    date_to_exclusive = date_to + datetime.timedelta(days=1)

    cash_accounts_query = Accounts.query.filter(Accounts.Type.in_(['Bank', 'Exch']))
    if filters['com']:
        cash_accounts_query = cash_accounts_query.filter(Accounts.Co == filters['com'])
    cash_accounts = cash_accounts_query.order_by(Accounts.Co, Accounts.Name).all()
    cash_account_names = [account.Name for account in cash_accounts]
    if filters['cash_account']:
        cash_account_names = [filters['cash_account']]

    account_rows = Accounts.query.all()
    accounts_by_id = {account.id: account for account in account_rows}
    accounts_by_company_name = {(account.Co, account.Name): account for account in account_rows}

    cash_query = Gledger.query.filter(
        (Gledger.Account.in_(cash_account_names if cash_account_names else ['__none__'])) &
        (Gledger.Date >= date_from) &
        (Gledger.Date < date_to_exclusive)
    )
    if filters['com']:
        cash_query = cash_query.filter(Gledger.Com == filters['com'])
    cash_lines = cash_query.order_by(Gledger.Date.asc(), Gledger.id.asc()).all()

    detail_rows = []
    processed = set()
    section_totals = {
        'Operating Activities': 0,
        'Investing Activities': 0,
        'Financing Activities': 0,
        'Transfers': 0,
    }
    for cash_line in cash_lines:
        key = cash_line.JournalId or (f'TCODE-{cash_line.Com}-{cash_line.Tcode}' if cash_line.Tcode else f'LINE-{cash_line.id}')
        if key in processed:
            continue
        processed.add(key)

        if cash_line.JournalId:
            journal_lines = Gledger.query.filter(Gledger.JournalId == cash_line.JournalId).all()
        elif cash_line.Tcode:
            journal_lines = Gledger.query.filter(
                (Gledger.Com == cash_line.Com) &
                (Gledger.Tcode == cash_line.Tcode)
            ).all()
        else:
            journal_lines = [cash_line]

        if filters['com']:
            journal_lines = [line for line in journal_lines if line.Com == filters['com']]

        selected_cash_lines = [
            line for line in journal_lines
            if line.Account in cash_account_names and date_from <= line.Date < date_to_exclusive
        ]
        cash_change = sum((line.Debit or 0) - (line.Credit or 0) for line in selected_cash_lines)

        offset_lines = [line for line in journal_lines if line.Account not in cash_account_names]
        offset_accounts = []
        for line in offset_lines:
            account = accounts_by_id.get(line.Aid) or accounts_by_company_name.get((line.Com, line.Account))
            if account is not None:
                offset_accounts.append(account)
        section, classification = classify_cash_flow(offset_accounts, cash_line.SourceTable)

        if cash_change == 0 and section != 'Transfers':
            section = 'Transfers'
            classification = 'Net cash change is zero'
        section_totals[section] += cash_change

        detail_rows.append({
            'date': cash_line.Date,
            'section': section,
            'classification': classification,
            'cash_accounts': ', '.join(sorted({line.Account for line in selected_cash_lines})),
            'offset_accounts': ', '.join(sorted({line.Account for line in offset_lines})) or '(cash transfer)',
            'source': cash_line.Source or cash_line.SourceTable or '',
            'ref': cash_line.Ref or '',
            'journal': cash_line.JournalId or cash_line.Tcode or '',
            'memo': cash_line.JournalMemo or '',
            'amount': cash_change,
            'amount_fmt': money(cash_change),
        })

    beginning_cash = cash_balance_before(cash_account_names, filters['com'], date_from)
    ending_cash = cash_balance_before(cash_account_names, filters['com'], date_to_exclusive)
    net_change = ending_cash - beginning_cash

    section_summaries = []
    for section in ['Operating Activities', 'Investing Activities', 'Financing Activities']:
        section_summaries.append({
            'section': section,
            'amount': section_totals[section],
            'amount_fmt': money(section_totals[section]),
        })
    transfer_total = section_totals['Transfers']

    companies = [
        item[0] for item in db.session.query(Gledger.Com)
        .filter(Gledger.Com.isnot(None))
        .distinct()
        .order_by(Gledger.Com)
        .all()
    ]
    totals = {
        'beginning_cash': money(beginning_cash),
        'ending_cash': money(ending_cash),
        'net_change': money(net_change),
        'reported_change': money(section_totals['Operating Activities'] + section_totals['Investing Activities'] + section_totals['Financing Activities']),
        'transfers': money(transfer_total),
        'row_count': len(detail_rows),
    }

    return render_template(
        'cash_flow_statement.html',
        cmpdata=cmpdata,
        scac=scac,
        filters=filters,
        companies=companies,
        cash_accounts=cash_accounts,
        section_summaries=section_summaries,
        detail_rows=detail_rows,
        totals=totals,
    )


@main.route('/DepreciationSchedules', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def DepreciationSchedules():
    def ensure_depreciation_schema():
        if db.engine.dialect.name != 'mysql':
            DepreciationAsset.__table__.create(bind=db.engine, checkfirst=True)
            return

        columns = {
            'SourceTable': 'VARCHAR(45)',
            'SourceId': 'INT',
            'Company': 'VARCHAR(2)',
            'AssetName': 'VARCHAR(100)',
            'AssetIdentifier': 'VARCHAR(100)',
            'AssetAccount': 'VARCHAR(50)',
            'AccumDepAccount': 'VARCHAR(50)',
            'DepExpenseAccount': 'VARCHAR(50)',
            'InServiceDate': 'DATE',
            'CostBasis': 'INT',
            'SalvageValue': 'INT',
            'BookMethod': 'VARCHAR(45)',
            'BookLifeMonths': 'INT',
            'TaxMethod': 'VARCHAR(45)',
            'TaxClass': 'VARCHAR(45)',
            'TaxLifeMonths': 'INT',
            'Section179': 'INT',
            'BonusDepreciation': 'INT',
            'PriorBookAccum': 'INT',
            'PriorTaxAccum': 'INT',
            'Status': 'VARCHAR(25)',
            'CreatedAt': 'DATETIME',
            'UpdatedAt': 'DATETIME',
        }

        create_columns = ',\n                '.join(
            [f'`{name}` {definition}' for name, definition in columns.items()]
        )
        with db.engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE IF NOT EXISTS `depreciation_assets` (
                    `id` INT NOT NULL AUTO_INCREMENT,
                    {create_columns},
                    PRIMARY KEY (`id`)
                )
            """))
            existing_columns = {
                row[0] for row in conn.execute(text('SHOW COLUMNS FROM `depreciation_assets`'))
            }
            for name, definition in columns.items():
                if name not in existing_columns:
                    conn.execute(text(
                        f'ALTER TABLE `depreciation_assets` ADD COLUMN `{name}` {definition}'
                    ))

    ensure_depreciation_schema()

    def parse_date(value):
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d').date()
        except:
            return None

    def money_to_cents(value):
        try:
            clean = str(value).replace('$', '').replace(',', '').strip()
            if clean in ['', '-']:
                return 0
            return int((Decimal(clean) * 100).quantize(Decimal('1')))
        except:
            return None

    def money(cents):
        return "${:,.2f}".format((cents or 0) / 100)

    def source_label(source_table, source_id):
        if source_table == 'Vehicles':
            item = Vehicles.query.get(source_id)
            if item is None:
                return 'Missing vehicle'
            return f'Vehicle {item.Unit or item.id} {item.Year or ""} {item.Make or ""} {item.Model or ""}'.strip()
        if source_table == 'Autos':
            item = Autos.query.get(source_id)
            if item is None:
                return 'Missing auto'
            return f'Auto {item.Jo or item.id} {item.Year or ""} {item.Make or ""} {item.Model or ""}'.strip()
        return f'{source_table} #{source_id}'

    def source_identifier(source_table, source_id):
        if source_table == 'Vehicles':
            item = Vehicles.query.get(source_id)
            if item is None:
                return ''
            return item.VIN or item.Plate or item.Unit or str(item.id)
        if source_table == 'Autos':
            item = Autos.query.get(source_id)
            if item is None:
                return ''
            return item.VIN or item.Title or item.Jo or str(item.id)
        return str(source_id)

    def schedule_rows(asset, mode):
        in_service = asset.InServiceDate
        if in_service is None:
            return []
        cost = asset.CostBasis or 0
        salvage = asset.SalvageValue or 0
        if mode == 'book':
            life_months = asset.BookLifeMonths or 60
            basis = max(cost - salvage - (asset.PriorBookAccum or 0), 0)
            first_year_extra = 0
            prior_accum = asset.PriorBookAccum or 0
        else:
            life_months = asset.TaxLifeMonths or 60
            section179 = min(asset.Section179 or 0, cost)
            bonus = min(asset.BonusDepreciation or 0, max(cost - section179, 0))
            basis = max(cost - section179 - bonus - (asset.PriorTaxAccum or 0), 0)
            first_year_extra = section179 + bonus
            prior_accum = asset.PriorTaxAccum or 0

        if life_months <= 0:
            return []
        monthly = basis / life_months
        rows = []
        accumulated = prior_accum
        months_used = 0
        year = in_service.year
        while months_used < life_months or first_year_extra:
            start_month = in_service.month if year == in_service.year else 1
            months = min(12 - start_month + 1, life_months - months_used)
            if months < 0:
                months = 0
            depreciation = int(round(monthly * months))
            if first_year_extra:
                depreciation += first_year_extra
                first_year_extra = 0
            if depreciation <= 0 and months_used >= life_months:
                break
            months_used += months
            accumulated += depreciation
            rows.append({
                'year': year,
                'depreciation': depreciation,
                'depreciation_fmt': money(depreciation),
                'accumulated': accumulated,
                'accumulated_fmt': money(accumulated),
                'basis': max(cost - accumulated, 0),
                'basis_fmt': money(max(cost - accumulated, 0)),
            })
            year += 1
            if len(rows) > 50:
                break
        return rows

    today_value = datetime.date.today().strftime('%Y-%m-%d')
    selected_id = request.values.get('current_asset_id', '').strip()
    if request.values.get('load_asset') is not None:
        selected_id = request.values.get('asset_id', '').strip()
    err = []
    selected = {
        'asset_id': selected_id,
        'source': request.values.get('source', ''),
        'company': request.values.get('company', cmpdata[10]),
        'asset_name': request.values.get('asset_name', ''),
        'asset_identifier': request.values.get('asset_identifier', ''),
        'asset_account': request.values.get('asset_account', ''),
        'accum_dep_account': request.values.get('accum_dep_account', ''),
        'dep_expense_account': request.values.get('dep_expense_account', ''),
        'in_service_date': request.values.get('in_service_date', today_value),
        'cost_basis': request.values.get('cost_basis', ''),
        'salvage_value': request.values.get('salvage_value', '0.00'),
        'book_method': request.values.get('book_method', 'Straight Line'),
        'book_life_months': request.values.get('book_life_months', '60'),
        'tax_method': request.values.get('tax_method', 'Straight Line'),
        'tax_class': request.values.get('tax_class', ''),
        'tax_life_months': request.values.get('tax_life_months', '60'),
        'section179': request.values.get('section179', '0.00'),
        'bonus_depreciation': request.values.get('bonus_depreciation', '0.00'),
        'prior_book_accum': request.values.get('prior_book_accum', '0.00'),
        'prior_tax_accum': request.values.get('prior_tax_accum', '0.00'),
        'status': request.values.get('status', 'Active'),
    }

    if request.method == 'POST' and request.values.get('load_asset') is not None:
        asset = DepreciationAsset.query.get(selected_id or 0)
        if asset is not None:
            selected.update({
                'asset_id': str(asset.id),
                'source': f'{asset.SourceTable}:{asset.SourceId}',
                'company': asset.Company or cmpdata[10],
                'asset_name': asset.AssetName or '',
                'asset_identifier': asset.AssetIdentifier or '',
                'asset_account': asset.AssetAccount or '',
                'accum_dep_account': asset.AccumDepAccount or '',
                'dep_expense_account': asset.DepExpenseAccount or '',
                'in_service_date': asset.InServiceDate.strftime('%Y-%m-%d') if asset.InServiceDate else today_value,
                'cost_basis': money(asset.CostBasis).replace('$', ''),
                'salvage_value': money(asset.SalvageValue).replace('$', ''),
                'book_method': asset.BookMethod or 'Straight Line',
                'book_life_months': str(asset.BookLifeMonths or 60),
                'tax_method': asset.TaxMethod or 'Straight Line',
                'tax_class': asset.TaxClass or '',
                'tax_life_months': str(asset.TaxLifeMonths or 60),
                'section179': money(asset.Section179).replace('$', ''),
                'bonus_depreciation': money(asset.BonusDepreciation).replace('$', ''),
                'prior_book_accum': money(asset.PriorBookAccum).replace('$', ''),
                'prior_tax_accum': money(asset.PriorTaxAccum).replace('$', ''),
                'status': asset.Status or 'Active',
            })

    if request.method == 'POST' and request.values.get('use_source') is not None:
        source = selected['source']
        try:
            source_table, source_id = source.split(':', 1)
            source_id = int(source_id)
            selected['asset_name'] = source_label(source_table, source_id)
            selected['asset_identifier'] = source_identifier(source_table, source_id)
        except:
            err.append('Choose a valid operational source asset.')

    if request.method == 'POST' and request.values.get('save_asset') is not None:
        try:
            source_table, source_id = selected['source'].split(':', 1)
            source_id = int(source_id)
        except:
            source_table, source_id = None, None
            err.append('Choose a source asset to link.')

        in_service = parse_date(selected['in_service_date'])
        if in_service is None:
            err.append('In-service date is invalid.')
        cost_basis = money_to_cents(selected['cost_basis'])
        salvage_value = money_to_cents(selected['salvage_value'])
        section179 = money_to_cents(selected['section179'])
        bonus_depreciation = money_to_cents(selected['bonus_depreciation'])
        prior_book_accum = money_to_cents(selected['prior_book_accum'])
        prior_tax_accum = money_to_cents(selected['prior_tax_accum'])
        for label, value in [
            ('Cost basis', cost_basis),
            ('Salvage value', salvage_value),
            ('Section 179', section179),
            ('Bonus depreciation', bonus_depreciation),
            ('Prior book accumulated depreciation', prior_book_accum),
            ('Prior tax accumulated depreciation', prior_tax_accum),
        ]:
            if value is None:
                err.append(f'{label} is invalid.')
        try:
            book_life_months = int(selected['book_life_months'])
            tax_life_months = int(selected['tax_life_months'])
        except:
            book_life_months, tax_life_months = 0, 0
            err.append('Book and tax life must be month counts.')
        if book_life_months <= 0 or tax_life_months <= 0:
            err.append('Book and tax life must be greater than zero.')
        if cost_basis is not None and cost_basis <= 0:
            err.append('Cost basis must be greater than zero.')
        if not selected['asset_name'].strip():
            err.append('Asset name is required.')

        if not err:
            now = datetime.datetime.now()
            asset = DepreciationAsset.query.get(selected_id or 0)
            if asset is None:
                asset = DepreciationAsset(
                    SourceTable=source_table,
                    SourceId=source_id,
                    Company=selected['company'],
                    AssetName=selected['asset_name'].strip(),
                    AssetIdentifier=selected['asset_identifier'].strip(),
                    AssetAccount=selected['asset_account'].strip(),
                    AccumDepAccount=selected['accum_dep_account'].strip(),
                    DepExpenseAccount=selected['dep_expense_account'].strip(),
                    InServiceDate=in_service,
                    CostBasis=cost_basis,
                    SalvageValue=salvage_value,
                    BookMethod=selected['book_method'],
                    BookLifeMonths=book_life_months,
                    TaxMethod=selected['tax_method'],
                    TaxClass=selected['tax_class'].strip(),
                    TaxLifeMonths=tax_life_months,
                    Section179=section179,
                    BonusDepreciation=bonus_depreciation,
                    PriorBookAccum=prior_book_accum,
                    PriorTaxAccum=prior_tax_accum,
                    Status=selected['status'],
                    CreatedAt=now,
                    UpdatedAt=now,
                )
                db.session.add(asset)
            else:
                asset.SourceTable = source_table
                asset.SourceId = source_id
                asset.Company = selected['company']
                asset.AssetName = selected['asset_name'].strip()
                asset.AssetIdentifier = selected['asset_identifier'].strip()
                asset.AssetAccount = selected['asset_account'].strip()
                asset.AccumDepAccount = selected['accum_dep_account'].strip()
                asset.DepExpenseAccount = selected['dep_expense_account'].strip()
                asset.InServiceDate = in_service
                asset.CostBasis = cost_basis
                asset.SalvageValue = salvage_value
                asset.BookMethod = selected['book_method']
                asset.BookLifeMonths = book_life_months
                asset.TaxMethod = selected['tax_method']
                asset.TaxClass = selected['tax_class'].strip()
                asset.TaxLifeMonths = tax_life_months
                asset.Section179 = section179
                asset.BonusDepreciation = bonus_depreciation
                asset.PriorBookAccum = prior_book_accum
                asset.PriorTaxAccum = prior_tax_accum
                asset.Status = selected['status']
                asset.UpdatedAt = now
            db.session.commit()
            selected['asset_id'] = str(asset.id)
            err.append(f'Depreciation asset {asset.AssetName} saved.')

    source_options = []
    for item in Vehicles.query.order_by(Vehicles.Unit, Vehicles.id).all():
        source_options.append({
            'value': f'Vehicles:{item.id}',
            'label': source_label('Vehicles', item.id),
        })
    for item in Autos.query.order_by(Autos.Jo, Autos.id).limit(200).all():
        source_options.append({
            'value': f'Autos:{item.id}',
            'label': source_label('Autos', item.id),
        })

    companies = [
        item[0] for item in db.session.query(Accounts.Co)
        .filter(Accounts.Co.isnot(None))
        .distinct()
        .order_by(Accounts.Co)
        .all()
    ]
    asset_accounts = Accounts.query.filter(
        (Accounts.Type.in_(['Asset', 'Bank', 'Exch'])) |
        (Accounts.Category == 'Assets')
    ).order_by(Accounts.Co, Accounts.Name).all()
    accum_accounts = Accounts.query.filter(
        (Accounts.Type.in_(['Asset', 'Current Liability', 'Equity'])) |
        (Accounts.Name.contains('Depreciation'))
    ).order_by(Accounts.Co, Accounts.Name).all()
    expense_accounts = Accounts.query.filter(
        Accounts.Type == 'Expense'
    ).order_by(Accounts.Co, Accounts.Name).all()

    assets = DepreciationAsset.query.order_by(DepreciationAsset.Company, DepreciationAsset.AssetName).all()
    for asset in assets:
        asset.cost_fmt = money(asset.CostBasis)
        asset.book_accum_fmt = money(sum(row['depreciation'] for row in schedule_rows(asset, 'book')))
        asset.tax_accum_fmt = money(sum(row['depreciation'] for row in schedule_rows(asset, 'tax')))
        asset.source_label = source_label(asset.SourceTable, asset.SourceId)

    selected_asset = DepreciationAsset.query.get(selected.get('asset_id') or 0)
    book_schedule = schedule_rows(selected_asset, 'book') if selected_asset is not None else []
    tax_schedule = schedule_rows(selected_asset, 'tax') if selected_asset is not None else []

    return render_template(
        'depreciation_schedules.html',
        cmpdata=cmpdata,
        scac=scac,
        selected=selected,
        source_options=source_options,
        companies=companies,
        asset_accounts=asset_accounts,
        accum_accounts=accum_accounts,
        expense_accounts=expense_accounts,
        assets=assets,
        book_schedule=book_schedule,
        tax_schedule=tax_schedule,
        err='\n'.join(err),
    )


@main.route('/Banking', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def Banking():
    from webapp.iso_Bank import isoBank, is_credit_reconciliation_account
    odata, oder, err, modata, modlink, leftscreen, leftsize, today, now, docref, cache, acdata, thismuch, acctinfo, hv = isoBank()

    def reconciliation_history_start(range_key):
        today_date = date.today()
        if range_key == 'last_12':
            return today_date - timedelta(days=365)
        if range_key == 'last_24':
            return today_date - timedelta(days=730)
        if range_key == 'all':
            return None
        return date(today_date.year, 1, 1)

    def reconciliation_status_label(status):
        if status == 1:
            return 'Finalized'
        if status == 25:
            return 'Trial'
        if status == 0:
            return 'Reopened'
        return f'Status {status}' if status is not None else ''

    def reconciliation_history(account_name, range_key):
        start_date = reconciliation_history_start(range_key)
        query = Reconciliations.query.filter(Reconciliations.Account == account_name)
        if start_date is not None:
            query = query.filter(Reconciliations.Rdate >= start_date)
        rows = query.order_by(Reconciliations.Rdate.desc(), Reconciliations.id.desc()).all()
        history = []
        for row in rows:
            history.append({
                'date': row.Rdate,
                'beginning': row.Bbal or '',
                'ending': row.Ebal or '',
                'deposits': row.Deposits or '',
                'withdrawals': row.Withdraws or '',
                'servicefees': row.Servicefees or '',
                'difference': row.Diff or '',
                'status': row.Status,
                'status_label': reconciliation_status_label(row.Status),
            })
        return history

    def date_only(value):
        if isinstance(value, datetime.datetime):
            return value.date()
        return value

    def sort_datetime(value):
        if isinstance(value, datetime.datetime):
            return value
        if isinstance(value, datetime.date):
            return datetime.datetime.combine(value, datetime.time.min)
        return datetime.datetime.min

    bank_rows = []
    grouped = {}
    for item in odata:
        if item.Type in ['DD', 'XD']:
            key = (
                date_only(item.Date),
                item.Type or '',
                item.Source or '',
                item.Ref or '',
                item.Reconciled,
            )
        else:
            key = ('single', item.id)
        row = grouped.setdefault(key, {
            'ids': [],
            'date': item.Date,
            'recorded': item.Recorded,
            'type': item.Type or '',
            'tcodes': [],
            'source': item.Source or '',
            'ref': item.Ref or '',
            'debit': 0,
            'credit': 0,
            'reconciled': item.Reconciled,
            'memos': [],
        })
        row['ids'].append(item.id)
        if item.Tcode and item.Tcode not in row['tcodes']:
            row['tcodes'].append(item.Tcode)
        row['debit'] += item.Debit or 0
        row['credit'] += item.Credit or 0
        if item.JournalMemo and item.JournalMemo not in row['memos']:
            row['memos'].append(item.JournalMemo)
        if item.Recorded and (not row['recorded'] or item.Recorded > row['recorded']):
            row['recorded'] = item.Recorded
    for index, row in enumerate(grouped.values()):
        row['index'] = index
        row['id_value'] = ','.join(str(row_id) for row_id in row['ids'])
        row['checked'] = row['ids'] and all(row_id in hv[1] for row_id in row['ids'])
        row['row_count'] = len(row['ids'])
        row['tcode'] = row['tcodes'][0] if row['tcodes'] else ''
        row['all_tcodes'] = ', '.join(row['tcodes'])
        row['has_more_tcodes'] = len(row['tcodes']) > 1
        row['memo'] = '; '.join(row['memos'])
        row['debit_fmt'] = "${:,.2f}".format(float(row['debit'] or 0) / 100)
        row['credit_fmt'] = "${:,.2f}".format(float(row['credit'] or 0) / 100)
        row['has_detail'] = bool(row['ids'])
        row['has_payment_mismatch'] = False
        bank_rows.append(row)
    bank_rows = sorted(
        bank_rows,
        key=lambda row: (sort_datetime(row['date']), row['source'], row['ref']),
        reverse=True,
    )
    reconciliation_history_range = request.values.get('reconciliation_history_range', 'current_year')
    reconciliation_history_rows = reconciliation_history(acctinfo[6], reconciliation_history_range)
    return render_template(
        'banking.html',
        cmpdata=cmpdata,
        scac=scac,
        odata=odata,
        oder=oder,
        err=err,
        modata=modata,
        modlink=modlink,
        leftscreen=leftscreen,
        leftsize=leftsize,
        today=today,
        now=now,
        docref=docref,
        cache=cache,
        acdata=acdata,
        thismuch=thismuch,
        acctinfo=acctinfo,
        hv=hv,
        bank_rows=bank_rows,
        is_credit_reconciliation=is_credit_reconciliation_account(acctinfo[6]),
        reconciliation_history_rows=reconciliation_history_rows,
        reconciliation_history_range=reconciliation_history_range,
    )


@main.route('/PlaidConnections', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def PlaidConnections():
    ready, plaid_env = plaid_ready()
    err = []

    if request.method == 'POST':
        action = request.values.get('action')
        if action == 'map_accounts':
            for plaid_account in plaid_accounts_for_current_scac():
                local_id = request.values.get(f'local_account_{plaid_account.id}')
                try:
                    plaid_account.LocalAccountId = int(local_id) if local_id else None
                except:
                    plaid_account.LocalAccountId = None
                PlaidTransaction.query.filter_by(PlaidAccountId=plaid_account.id).update(
                    {'LocalAccountId': plaid_account.LocalAccountId},
                    synchronize_session=False,
                )
            db.session.commit()
            flash('Plaid account mappings updated.', 'success')
            return redirect(url_for('main.PlaidConnections'))

        if action == 'sync_item':
            item_id = request.values.get('item_id')
            item = PlaidItem.query.get(item_id)
            if item is None:
                flash('Plaid connection not found.', 'danger')
                return redirect(url_for(
                    'main.PlaidConnections',
                    sync_item_id=item_id or '',
                    sync_status='danger',
                    sync_message='Plaid connection not found.',
                ))
            else:
                try:
                    update_item_status(item)
                    added, modified, removed = sync_item_transactions(item)
                    message = f'Added {added}, modified {modified}, removed {removed}.'
                    flash(f'Plaid sync complete. {message}', 'success')
                    return redirect(url_for(
                        'main.PlaidConnections',
                        sync_item_id=item.id,
                        sync_status='success',
                        sync_message=message,
                    ))
                except Exception as exc:
                    message = f'Plaid sync failed: {exc}'
                    flash(message, 'danger')
                    return redirect(url_for(
                        'main.PlaidConnections',
                        sync_item_id=item.id,
                        sync_status='danger',
                        sync_message=message,
                    ))

        if action == 'create_bill_from_plaid':
            tx_id = request.values.get('transaction_id')
            expense_account_id = request.values.get('expense_account_id')
            vendor_name = request.values.get('vendor_name')
            err, bill = create_bill_from_plaid_transaction(
                tx_id,
                expense_account_id,
                vendor_name,
                current_user.username,
            )
            if err:
                flash(' '.join(err), 'danger')
            else:
                flash(f'Created bill payment {bill.Jo} from Plaid transaction.', 'success')
            return redirect(url_for('main.PlaidConnections'))

        if action == 'match_bill_from_plaid':
            tx_id = request.values.get('transaction_id')
            bill_id = request.values.get('bill_id')
            expense_account_id = request.values.get('expense_account_id')
            vendor_name = request.values.get('vendor_name')
            err = match_plaid_transaction_to_bill(
                tx_id,
                bill_id,
                expense_account_id,
                vendor_name,
                current_user.username,
            )
            if err:
                flash(' '.join(err), 'danger')
            else:
                flash('Plaid transaction matched to existing bill payment.', 'success')
            return redirect(url_for('main.PlaidConnections'))

        if action == 'create_transfer_from_plaid':
            tx_id = request.values.get('transaction_id')
            other_account_id = request.values.get('other_account_id')
            owner_transfer_treatment = request.values.get('owner_transfer_treatment')
            err, tcode = create_transfer_from_plaid_transaction(
                tx_id,
                other_account_id,
                owner_transfer_treatment,
                current_user.username,
            )
            if err:
                flash(' '.join(err), 'danger')
            else:
                flash(f'Created account transfer {tcode} from Plaid transaction.', 'success')
            return redirect(url_for('main.PlaidConnections'))

        if action == 'ignore_plaid_transaction':
            tx_id = request.values.get('transaction_id')
            note = request.values.get('review_note')
            err = ignore_plaid_transaction(tx_id, current_user.username, note)
            if err:
                flash(' '.join(err), 'danger')
            else:
                flash('Plaid transaction marked ignored.', 'success')
            return redirect(url_for('main.PlaidConnections'))

    items, plaid_accounts, transactions, local_accounts, plaid_account_lookup, mapping_options = plaid_dashboard_data()
    review_transactions = plaid_review_transactions()
    expense_account_options = plaid_expense_account_options()
    transfer_account_options = plaid_transfer_account_options()
    vendor_rule_lookup = plaid_vendor_rule_lookup(review_transactions)
    bill_match_options = plaid_bill_match_options(review_transactions)
    processed_transactions = plaid_processed_transactions()
    sync_result = {
        'item_id': request.args.get('sync_item_id'),
        'status': request.args.get('sync_status'),
        'message': request.args.get('sync_message'),
    }
    return render_template(
        'plaid_connections.html',
        cmpdata=cmpdata,
        scac=scac,
        ready=ready,
        plaid_env=plaid_env,
        items=items,
        plaid_accounts=plaid_accounts,
        transactions=transactions,
        local_accounts=local_accounts,
        plaid_account_lookup=plaid_account_lookup,
        mapping_options=mapping_options,
        review_transactions=review_transactions,
        expense_account_options=expense_account_options,
        transfer_account_options=transfer_account_options,
        vendor_rule_lookup=vendor_rule_lookup,
        bill_match_options=bill_match_options,
        processed_transactions=processed_transactions,
        sync_result=sync_result,
        err=err,
    )


@main.route('/PlaidConnections/link_token', methods=['POST'])
@login_required
@financial_mfa_required
def PlaidLinkToken():
    try:
        data = create_link_token(current_user.id, current_user.username)
        return jsonify({'link_token': data.get('link_token')})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


@main.route('/PlaidConnections/exchange_public_token', methods=['POST'])
@login_required
@financial_mfa_required
def PlaidExchangePublicToken():
    payload = request.get_json(silent=True) or {}
    public_token = payload.get('public_token')
    if not public_token:
        return jsonify({'error': 'Missing public token'}), 400
    try:
        access_token, item_id = exchange_public_token(public_token)
        item = upsert_item_from_exchange(access_token, item_id)
        return jsonify({'ok': True, 'item_id': item.ItemId, 'institution_name': item.InstitutionName})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


@main.route('/Banking/payment_detail', methods=['GET'])
@login_required
@financial_mfa_required
def BankingPaymentDetail():
    def cents_from_value(value):
        try:
            clean = str(value).replace(',', '').replace('$', '').strip()
            return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))
        except:
            return 0

    def money(cents):
        return "${:,.2f}".format((cents or 0) / 100)

    def forwarding_name(job):
        if job is None:
            return ''
        for value in [job.Exporter, job.Consignee, job.Notify]:
            if value:
                return str(value).splitlines()[0].strip()
        return ''

    ids = []
    for item in request.values.get('ids', '').split(','):
        try:
            ids.append(int(item))
        except:
            pass

    detail_rows = []
    invoice_total = 0
    received_total = 0
    balance_total = 0
    header = {'date': '', 'source': '', 'ref': '', 'amount': '$0.00'}

    for ledger_id in ids:
        ledger_line = Gledger.query.get(ledger_id)
        if ledger_line is None:
            continue
        if header['date'] == '' and ledger_line.Date is not None:
            header['date'] = ledger_line.Date.strftime('%Y-%m-%d')
            header['source'] = ledger_line.Source or ''
            header['ref'] = ledger_line.Ref or ''
        if ledger_line.SourceTable == 'ReceivePaymentBatch' and ledger_line.Sid:
            allocation_rows = Gledger.query.filter(
                (Gledger.Sid == ledger_line.Sid) &
                (Gledger.Type == 'IC') &
                (Gledger.Com == ledger_line.Com)
            ).order_by(Gledger.JournalSeq.asc(), Gledger.id.asc()).all()
            received_total += ledger_line.Debit or 0
            for allocation in allocation_rows:
                order = Orders.query.filter(Orders.Jo == allocation.Tcode).first()
                invoice = Invoices.query.filter(Invoices.Jo == allocation.Tcode).first()
                forwarding_job = None if order is not None else OverSeas.query.filter(OverSeas.Jo == allocation.Tcode).first()
                invo_cents = cents_from_value(
                    invoice.Total if invoice is not None
                    else (order.InvoTotal if order is not None
                          else (forwarding_job.InvoTotal if forwarding_job is not None else 0))
                )
                received_cents = allocation.Credit or 0
                balance_cents = invo_cents - received_cents
                invoice_total += invo_cents
                balance_total += balance_cents
                detail_rows.append({
                    'ledger_id': allocation.id,
                    'jo': allocation.Tcode or '',
                    'order_id': order.id if order is not None else (forwarding_job.id if forwarding_job is not None else ''),
                    'shipper': order.Shipper if order is not None else (forwarding_name(forwarding_job) or allocation.Source),
                    'container': order.Container if order is not None else (forwarding_job.Container if forwarding_job is not None else ''),
                    'invoice': money(invo_cents),
                    'received': money(received_cents),
                    'balance': money(balance_cents),
                    'balance_cents': balance_cents,
                    'pay_ref': allocation.Ref or ledger_line.Ref,
                    'paid_date': allocation.Date.strftime('%Y-%m-%d') if allocation.Date is not None else '',
                    'manual': False,
                })
            continue
        order = Orders.query.filter(Orders.Jo == ledger_line.Tcode).first()
        forwarding_job = None if order is not None else OverSeas.query.filter(OverSeas.Jo == ledger_line.Tcode).first()
        income = Deposits.query.filter(Deposits.Jo == ledger_line.Tcode).first()
        invoice = Invoices.query.filter(Invoices.Jo == ledger_line.Tcode).first()
        is_manual = ledger_line.SourceTable == 'ManualDeposit'
        original_payment = None
        if ledger_line.SourceTable == 'CounterDepositItem' and ledger_line.SourceId:
            original_payment = Gledger.query.get(ledger_line.SourceId)
        invo_cents = 0 if is_manual else cents_from_value(
            invoice.Total if invoice is not None
            else (order.InvoTotal if order is not None
                  else (forwarding_job.InvoTotal if forwarding_job is not None else 0))
        )
        received_cents = ledger_line.Debit or 0
        balance_cents = 0 if is_manual else invo_cents - received_cents
        invoice_total += invo_cents
        received_total += received_cents
        balance_total += balance_cents
        detail_rows.append({
            'ledger_id': ledger_line.id,
            'jo': ledger_line.Tcode or '',
            'order_id': order.id if order is not None else (forwarding_job.id if forwarding_job is not None else ''),
            'shipper': order.Shipper if order is not None else (forwarding_name(forwarding_job) or ledger_line.Source),
            'container': order.Container if order is not None else (forwarding_job.Container if forwarding_job is not None else ''),
            'invoice': '' if is_manual else money(invo_cents),
            'received': money(received_cents),
            'balance': '' if is_manual else money(balance_cents),
            'balance_cents': balance_cents,
            'pay_ref': original_payment.Ref if original_payment is not None else (income.Ref if income is not None else ledger_line.Ref),
            'paid_date': original_payment.Date.strftime('%Y-%m-%d') if original_payment is not None and original_payment.Date is not None else (income.Date.strftime('%Y-%m-%d') if income is not None and income.Date is not None else (ledger_line.Date.strftime('%Y-%m-%d') if ledger_line.Date is not None else '')),
            'manual': is_manual,
        })

    header['amount'] = money(received_total)
    return jsonify({
        'header': header,
        'invoice_total': money(invoice_total),
        'received_total': money(received_total),
        'balance_total': money(balance_total),
        'has_payment_mismatch': any(row['balance_cents'] != 0 for row in detail_rows if not row['manual']),
        'rows': detail_rows,
    })


@main.route('/ReceiveByAccount', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def ReceiveByAccount():
    from webapp.class8_tasks import ReceiveByAccount_task
    from webapp.viewfuncs import d2s

    if request.method == 'POST' and request.values.get('Return') is not None:
        return redirect(url_for('main.Class8Main', genre='Trucking'))

    def receive_batch_date_range():
        today = datetime.date.today()
        default_start = today - datetime.timedelta(days=30)
        start_text = request.values.get('batch_start_date') or default_start.strftime('%Y-%m-%d')
        end_text = request.values.get('batch_end_date') or today.strftime('%Y-%m-%d')
        try:
            start_date = datetime.datetime.strptime(start_text, '%Y-%m-%d')
        except:
            start_date = datetime.datetime.combine(default_start, datetime.time.min)
            start_text = default_start.strftime('%Y-%m-%d')
        try:
            end_date = datetime.datetime.strptime(end_text, '%Y-%m-%d')
        except:
            end_date = datetime.datetime.combine(today, datetime.time.min)
            end_text = today.strftime('%Y-%m-%d')
        if end_date < start_date:
            start_date, end_date = end_date, start_date
            start_text, end_text = end_text, start_text
        end_exclusive = end_date + datetime.timedelta(days=1)
        return start_text, end_text, start_date, end_exclusive

    def receive_order_stopdate():
        lookbacktime = request.values.get('lookbacktime')
        if lookbacktime == 'Two Years':
            lookback = 728
        elif lookbacktime == 'Three Years':
            lookback = 1092
        else:
            lookback = 364
            lookbacktime = 'One Year'
        return lookbacktime, datetime.date.today() - datetime.timedelta(days=lookback)

    def cents_to_money(cents):
        try:
            return d2s(float(cents or 0) / 100)
        except:
            return '0.00'

    def money_to_cents(value):
        try:
            clean = str(value).replace('$', '').replace(',', '').strip()
            return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))
        except:
            return 0

    def selected_payment_ledger_ids():
        ids = []
        for selected in request.values.getlist('payment_batch'):
            for row_id in selected.split(','):
                try:
                    ids.append(int(row_id))
                except:
                    pass
        return ids

    def receive_success_message(holdvec):
        try:
            checked_count = sum(1 for value in holdvec[3] if value == 1)
        except:
            checked_count = 0
        customer = holdvec[0] or 'selected account'
        total = holdvec[5] or '0.00'
        ref = holdvec[8] or ''
        ref_text = f' Reference {ref}.' if ref else ''
        return f'Recorded received payment batch for {customer}: {checked_count} invoice(s), ${total}.{ref_text}'

    def create_counter_deposit(err_list):
        payment_ids = selected_payment_ledger_ids()
        if not payment_ids:
            err_list.append('Select one or more received check batches for the counter deposit.')
            return err_list

        bank_account_name = request.values.get('counter_deposit_account')
        deposit_date = request.values.get('counter_deposit_date')
        deposit_ref = (request.values.get('counter_deposit_ref') or '').strip()
        try:
            deposit_date_value = datetime.datetime.strptime(deposit_date, '%Y-%m-%d')
        except:
            err_list.append('Counter deposit date is invalid.')
            return err_list
        if not bank_account_name:
            err_list.append('Choose the bank account for the counter deposit.')
            return err_list
        if not deposit_ref:
            err_list.append('Enter a counter deposit reference number.')
            return err_list

        bank_account = Accounts.query.filter(
            (Accounts.Name == bank_account_name) &
            (Accounts.Co == cmpdata[10])
        ).first()
        undeposited_account = Accounts.query.filter(
            (Accounts.Name == 'Undeposited Funds') &
            (Accounts.Co == cmpdata[10])
        ).first()
        if bank_account is None or undeposited_account is None:
            err_list.append('Required bank or undeposited-funds account could not be found.')
            return err_list

        payment_rows = Gledger.query.filter(
            (Gledger.id.in_(payment_ids)) &
            (Gledger.Type.in_(['ID', 'DD'])) &
            (Gledger.Account == 'Undeposited Funds') &
            (Gledger.Com == cmpdata[10])
        ).order_by(Gledger.Date.asc(), Gledger.id.asc()).all()
        if not payment_rows:
            err_list.append('Selected rows do not include undeposited check payments.')
            return err_list
        if len({row.id for row in payment_rows}) != len(set(payment_ids)):
            err_list.append('Only undeposited check/deposit batches can be included in a counter deposit.')
            return err_list

        already_deposited = Gledger.query.filter(
            (Gledger.SourceTable == 'CounterDepositItem') &
            (Gledger.SourceId.in_([row.id for row in payment_rows])) &
            (Gledger.Type == 'XD')
        ).first()
        if already_deposited is not None:
            err_list.append('At least one selected check is already assigned to a counter deposit.')
            return err_list

        recorded = datetime.datetime.now()
        journal_id = f'CD-{cmpdata[10]}-{deposit_ref}'
        total_amount = 0
        for seq, row in enumerate(payment_rows, start=1):
            amount = row.Debit or 0
            if amount <= 0:
                continue
            total_amount += amount
            memo = f'Counter deposit {deposit_ref}; original check {row.Ref or ""}'.strip()

            bank_line = Gledger(
                Debit=amount,
                Credit=0,
                Account=bank_account.Name,
                Aid=bank_account.id,
                Source='Counter Deposit',
                Sid=row.Sid,
                Type='XD',
                Tcode=row.Tcode,
                Com=cmpdata[10],
                Recorded=recorded,
                Reconciled=0,
                Date=deposit_date_value,
                Ref=deposit_ref,
                JournalId=journal_id,
                JournalSeq=seq * 2 - 1,
                JournalMemo=memo,
                PostedBy='receive_by_account',
                PostedAt=recorded,
                SourceTable='CounterDepositItem',
                SourceId=row.id,
            )
            db.session.add(bank_line)

            clearing_line = Gledger(
                Debit=0,
                Credit=amount,
                Account='Undeposited Funds',
                Aid=undeposited_account.id,
                Source='Counter Deposit',
                Sid=row.Sid,
                Type='XC',
                Tcode=row.Tcode,
                Com=cmpdata[10],
                Recorded=recorded,
                Reconciled=0,
                Date=deposit_date_value,
                Ref=deposit_ref,
                JournalId=journal_id,
                JournalSeq=seq * 2,
                JournalMemo=memo,
                PostedBy='receive_by_account',
                PostedAt=recorded,
                SourceTable='CounterDepositItem',
                SourceId=row.id,
            )
            db.session.add(clearing_line)

        if total_amount <= 0:
            err_list.append('Counter deposit total is zero.')
            return err_list

        db.session.commit()
        err_list.append(f'Counter deposit {deposit_ref} posted to {bank_account.Name} for ${total_amount / 100:,.2f}.')
        return err_list

    def refresh_order_payment_status(order):
        payment_rows = Gledger.query.filter(
            (Gledger.Tcode == order.Jo) &
            (Gledger.Type == 'IC')
        ).order_by(Gledger.Date.desc(), Gledger.id.desc()).all()
        paid_cents = sum(row.Credit or 0 for row in payment_rows)
        latest = payment_rows[0] if payment_rows else None
        invoice = Invoices.query.filter(Invoices.Jo == order.Jo).first()
        invoice_total = 0.0
        try:
            invoice_total = float(invoice.Total if invoice is not None else order.InvoTotal or 0)
        except:
            invoice_total = 0.0
        paid_total = float(paid_cents) / 100
        balance_due = invoice_total - paid_total
        order.Payments = d2s(paid_total)
        order.BalDue = d2s(balance_due)
        if latest is not None:
            pay_record = PaymentsRec.query.get(latest.Sid) if latest.Sid else None
            order.PaidDate = latest.Date
            order.PaidAmt = d2s(float(latest.Credit or 0) / 100)
            order.PayRef = latest.Ref
            order.PayMeth = pay_record.Type if pay_record is not None else order.PayMeth
            order.PayAcct = pay_record.Account if pay_record is not None else order.PayAcct
            order.QBi = latest.Sid
        else:
            order.PaidDate = None
            order.PaidAmt = '0.00'
            order.PayRef = None
            order.QBi = None

        if paid_total > 0:
            if balance_due > .01:
                order.Istat = 4
            elif order.Istat in [6, 7, 8]:
                order.Istat = 8
            else:
                order.Istat = 5

    def update_reloaded_receive_batch(err_list):
        selected_ids = request.values.get('reloaded_batch_ids', '')
        try:
            original_ids = [int(row_id) for row_id in selected_ids.split(',') if row_id.strip()]
        except:
            err_list.append('Reloaded payment batch has invalid row data.')
            return err_list, selected_ids
        if not original_ids:
            err_list.append('Reload a payment batch before updating it.')
            return err_list, selected_ids

        old_debit_rows = Gledger.query.filter(
            (Gledger.id.in_(original_ids)) &
            (Gledger.Type.in_(['DD', 'ID'])) &
            (Gledger.Com == cmpdata[10])
        ).all()
        if not old_debit_rows:
            err_list.append('Could not find the reloaded payment batch to update.')
            return err_list, selected_ids

        first = old_debit_rows[0]
        payment_ref_id = first.Sid
        selected_payment_ref_ids = {row.Sid for row in old_debit_rows if row.Sid}
        pay_date = request.values.get('thisdate')
        pay_ref = request.values.get('thisref')
        pay_method = request.values.get('paymethod')
        deposit_account = request.values.get('acctfordeposit')
        try:
            pay_date_value = datetime.datetime.strptime(pay_date, '%Y-%m-%d')
        except:
            err_list.append('Payment date is invalid.')
            return err_list, selected_ids
        if not pay_method or not deposit_account:
            err_list.append('Choose payment method and deposit account before updating.')
            return err_list, selected_ids

        new_lines = {}
        for key in request.values:
            if not key.startswith('oder'):
                continue
            try:
                order_id = int(key.replace('oder', ''))
            except:
                continue
            order = Orders.query.get(order_id)
            if order is None:
                continue
            amount = money_to_cents(request.values.get(f'amount{order.id}', '0.00'))
            if amount != 0:
                new_lines[order.Jo] = (order, amount)
        if not new_lines:
            err_list.append('Select at least one invoice with a non-zero received amount.')
            return err_list, selected_ids
        updated_total_amount = sum(amount for order, amount in new_lines.values())
        if updated_total_amount <= 0:
            err_list.append('Updated payment batch total must be greater than zero.')
            return err_list, selected_ids

        payment_record = PaymentsRec.query.get(payment_ref_id) if payment_ref_id else None
        if payment_record is None:
            payment_record = PaymentsRec(
                Amount=0,
                Account=deposit_account,
                Source=first.Source,
                Type=pay_method,
                Com=cmpdata[10],
                Recorded=datetime.datetime.now(),
                Date=pay_date_value,
                Ref=pay_ref,
            )
            db.session.add(payment_record)
            db.session.flush()
            payment_ref_id = payment_record.id

        old_credit_rows = Gledger.query.filter(
            (Gledger.Sid == payment_ref_id) &
            (Gledger.Type == 'IC') &
            (Gledger.Com == cmpdata[10])
        ).all()
        old_jos = {row.Tcode for row in old_credit_rows if row.Tcode} or {row.Tcode for row in old_debit_rows if row.Tcode and not str(row.Tcode).startswith('RP-')}
        new_jos = set(new_lines.keys())
        affected_jos = old_jos | new_jos

        dtype = 'ID' if deposit_account in ['Undeposited Funds'] or pay_method in ['Cash', 'Check'] else 'DD'
        debit_account = Accounts.query.filter(
            (Accounts.Name == deposit_account) &
            (Accounts.Co == cmpdata[10])
        ).first()
        credit_account = Accounts.query.filter(
            (Accounts.Name == 'Accounts Receivable') &
            (Accounts.Co == cmpdata[10])
        ).first()
        if debit_account is None or credit_account is None:
            err_list.append('Required ledger accounts could not be found.')
            return err_list, selected_ids

        total_amount = 0
        recorded = datetime.datetime.now()
        journal_id = f'RP-{cmpdata[10]}-{payment_ref_id}'
        Gledger.query.filter(
            (Gledger.Sid == payment_ref_id) &
            (Gledger.Type.in_(['DD', 'ID', 'IC'])) &
            (Gledger.Com == cmpdata[10])
        ).delete(synchronize_session=False)
        if selected_payment_ref_ids:
            Gledger.query.filter(
                (Gledger.Sid.in_(selected_payment_ref_ids)) &
                (Gledger.Type.in_(['DD', 'ID', 'IC'])) &
                (Gledger.Com == cmpdata[10])
            ).delete(synchronize_session=False)
            PaymentsRec.query.filter(
                (PaymentsRec.id.in_([ref_id for ref_id in selected_payment_ref_ids if ref_id != payment_ref_id])) &
                (PaymentsRec.Com == cmpdata[10])
            ).delete(synchronize_session=False)

        for jo, (order, amount) in new_lines.items():
            total_amount += amount
        first_order = next(iter(new_lines.values()))[0]
        debit_line = Gledger(
            Debit=total_amount,
            Credit=0,
            Account=deposit_account,
            Aid=debit_account.id,
            Source=first_order.Shipper,
            Sid=payment_ref_id,
            Type=dtype,
            Tcode=journal_id,
            Com=cmpdata[10],
            Recorded=recorded,
            Reconciled=0,
            Date=pay_date_value,
            Ref=pay_ref,
            JournalId=journal_id,
            JournalSeq=1,
            JournalMemo=f'Received payment batch {pay_ref or payment_ref_id}',
            PostedBy='receive_by_account',
            PostedAt=recorded,
            SourceTable='ReceivePaymentBatch',
            SourceId=payment_ref_id,
        )
        db.session.add(debit_line)

        for seq, (jo, (order, amount)) in enumerate(new_lines.items(), start=2):
            credit_line = Gledger(
                Debit=0,
                Credit=amount,
                Account='Accounts Receivable',
                Aid=credit_account.id,
                Source=order.Shipper,
                Sid=payment_ref_id,
                Type='IC',
                Tcode=jo,
                Com=cmpdata[10],
                Recorded=recorded,
                Reconciled=0,
                Date=pay_date_value,
                Ref=pay_ref,
                JournalId=journal_id,
                JournalSeq=seq,
                JournalMemo=f'Received payment allocation {pay_ref or payment_ref_id}',
                PostedBy='receive_by_account',
                PostedAt=recorded,
                SourceTable='ReceivePaymentAllocation',
                SourceId=payment_ref_id,
            )
            db.session.add(credit_line)

        payment_record.Amount = total_amount
        payment_record.Account = deposit_account
        payment_record.Type = pay_method
        payment_record.Com = cmpdata[10]
        payment_record.Recorded = recorded
        payment_record.Date = pay_date_value
        payment_record.Ref = pay_ref

        for jo in affected_jos:
            order = Orders.query.filter(Orders.Jo == jo).first()
            if order is not None:
                refresh_order_payment_status(order)

        db.session.commit()
        updated_rows = Gledger.query.filter(
            (Gledger.Sid == payment_ref_id) &
            (Gledger.Type.in_(['DD', 'ID'])) &
            (Gledger.Com == cmpdata[10])
        ).order_by(Gledger.Date.asc(), Gledger.id.asc()).all()
        updated_ids = ','.join(str(row.id) for row in updated_rows)
        err_list.append(f'Updated reloaded payment batch {payment_ref_id} for ${total_amount / 100:,.2f}.')
        return err_list, updated_ids

    def reload_receive_batch(holdvec, selected_batches, err_list, use_posted_values=False):
        selected_batches = [batch for batch in dict.fromkeys(selected_batches) if batch]
        if not selected_batches:
            err_list.append('Select one or more received payment rows to reload.')
            return holdvec, err_list
        try:
            row_ids = []
            for selected_batch in selected_batches:
                row_ids.extend([int(row_id) for row_id in selected_batch.split(',') if row_id.strip()])
            row_ids = list(dict.fromkeys(row_ids))
        except:
            err_list.append('Selected payment batch has invalid row data.')
            return holdvec, err_list
        if not row_ids:
            err_list.append('Selected payment batch has no ledger rows.')
            return holdvec, err_list

        batch_rows = Gledger.query.filter(
            (Gledger.id.in_(row_ids)) &
            (Gledger.Type.in_(['DD', 'ID'])) &
            (Gledger.Com == cmpdata[10])
        ).order_by(Gledger.Date.asc(), Gledger.id.asc()).all()
        if not batch_rows:
            err_list.append('Could not find the selected payment batch.')
            return holdvec, err_list

        first = batch_rows[0]
        batch_signature = (
            first.Date.strftime('%Y-%m-%d') if first.Date else '',
            first.Source or '',
            first.Ref or '',
            first.Account or '',
            first.Type or '',
        )
        for row in batch_rows[1:]:
            row_signature = (
                row.Date.strftime('%Y-%m-%d') if row.Date else '',
                row.Source or '',
                row.Ref or '',
                row.Account or '',
                row.Type or '',
            )
            if row_signature != batch_signature:
                err_list.append('Selected payment rows must have the same date, source, reference, account, and type to reload together.')
                return holdvec, err_list
        customer = first.Source or ''
        pay_date = first.Date.strftime('%Y-%m-%d') if first.Date else datetime.date.today().strftime('%Y-%m-%d')
        pay_ref = first.Ref or ''
        deposit_account = first.Account or ''
        first_order = Orders.query.filter(Orders.Jo == first.Tcode).first()
        pay_method = first_order.PayMeth if first_order is not None and first_order.PayMeth else None
        if pay_method is None:
            pay_method = 'Direct Deposit' if first.Type == 'DD' else 'Check'
        if use_posted_values:
            pay_date = request.values.get('thisdate') or pay_date
            pay_ref = request.values.get('thisref') or pay_ref
            deposit_account = request.values.get('acctfordeposit') or deposit_account
            pay_method = request.values.get('paymethod') or pay_method

        allocation_rows = []
        if first.SourceTable == 'ReceivePaymentBatch' and first.Sid:
            allocation_rows = Gledger.query.filter(
                (Gledger.Sid == first.Sid) &
                (Gledger.Type == 'IC') &
                (Gledger.Com == cmpdata[10])
            ).order_by(Gledger.JournalSeq.asc(), Gledger.id.asc()).all()
            forwarding_allocations = [
                row for row in allocation_rows
                if Orders.query.filter(Orders.Jo == row.Tcode).first() is None
                and OverSeas.query.filter(OverSeas.Jo == row.Tcode).first() is not None
            ]
            if forwarding_allocations:
                err_list.append(
                    'Freight forwarding received-payment batches can be reviewed or included in a counter deposit, '
                    'but they cannot be reloaded in the trucking Receive By Account editor.'
                )
                return holdvec, err_list
        payment_detail_rows = allocation_rows or batch_rows

        lookbacktime, stopdate = receive_order_stopdate()
        receivable_statuses = [2, 3, 4, 6, 7]
        open_orders = Orders.query.filter(
            (Orders.Shipper == customer) &
            (Orders.Istat.in_(receivable_statuses)) &
            (Orders.Date > stopdate)
        ).order_by(Orders.Date.desc(), Orders.id.desc()).all()

        by_jo = {row.Tcode: row for row in payment_detail_rows if row.Tcode}
        current_ids = {order.id for order in open_orders}
        for row in payment_detail_rows:
            order = Orders.query.filter(Orders.Jo == row.Tcode).first()
            if order is not None and order.id not in current_ids:
                open_orders.append(order)
                current_ids.add(order.id)

        tjobs = Orders.query.filter(
            (Orders.Istat.in_(receivable_statuses)) &
            (Orders.Date > stopdate)
        ).all()
        comps = sorted({job.Shipper for job in tjobs if job.Shipper})
        if customer and customer not in comps:
            comps.append(customer)
            comps.sort()

        checks = [0] * len(open_orders)
        amts = ['0.00'] * len(open_orders)
        invts = ['0.00'] * len(open_orders)
        invotot = 0.0
        paytot = 0.0
        for index, order in enumerate(open_orders):
            invoice = Invoices.query.filter(Invoices.Jo == order.Jo).first()
            if invoice is not None:
                invts[index] = invoice.Total
            elif order.InvoTotal:
                invts[index] = order.InvoTotal

            payment_row = by_jo.get(order.Jo)
            if use_posted_values:
                posted_amount = request.values.get(f'amount{order.id}')
                if posted_amount is not None:
                    amts[index] = cents_to_money(money_to_cents(posted_amount))
                elif payment_row is not None:
                    amts[index] = cents_to_money((payment_row.Credit or 0) if payment_row.Type == 'IC' else (payment_row.Debit or 0))
                elif invoice is not None:
                    amts[index] = invoice.Total

                if request.values.get(f'oder{order.id}') is not None:
                    checks[index] = 1
                    try:
                        invotot += float(invts[index])
                    except:
                        pass
                    try:
                        paytot += float(amts[index])
                    except:
                        pass
            elif payment_row is not None:
                checks[index] = 1
                amts[index] = cents_to_money((payment_row.Credit or 0) if payment_row.Type == 'IC' else (payment_row.Debit or 0))
                try:
                    invotot += float(invts[index])
                except:
                    pass
                try:
                    paytot += float(amts[index])
                except:
                    pass
            elif invoice is not None:
                amts[index] = invoice.Total

        if pay_method in ['Cash', 'Check']:
            deposit_accounts = ['Undeposited Funds']
        elif pay_method == 'Credit Card':
            deposit_accounts = [account.Name for account in Accounts.query.filter(
                (Accounts.Type == 'Bank') & (Accounts.Description.contains('Merchant'))
            ).all()]
        else:
            deposit_accounts = [account.Name for account in Accounts.query.filter(
                (Accounts.Type == 'Bank') | (Accounts.Type == 'Exch')
            ).all()]
        if deposit_account and deposit_account not in deposit_accounts:
            deposit_accounts.append(deposit_account)

        holdvec[0] = customer
        holdvec[1] = open_orders
        holdvec[2] = comps
        holdvec[3] = checks
        holdvec[4] = d2s(invotot)
        holdvec[5] = d2s(paytot)
        holdvec[6] = amts
        holdvec[7] = pay_date
        holdvec[8] = pay_ref
        holdvec[9] = deposit_accounts
        holdvec[10] = deposit_account
        holdvec[11] = nonone(request.values.get('autodisbox'))
        holdvec[12] = invts
        holdvec[13] = 1 if paytot > 0 and deposit_account else 0
        holdvec[14] = pay_method
        holdvec[21] = ','.join(str(row_id) for row_id in row_ids)
        holdvec[20] = lookbacktime
        err_list.append(f'Reloaded payment batch for {customer} dated {pay_date}. Review amounts, then preview or record.')
        return holdvec, err_list

    batch_start_date, batch_end_date, batch_start, batch_end_exclusive = receive_batch_date_range()
    is_reloaded_batch = request.values.get('reloaded_batch_ids') is not None
    is_reloaded_record = is_reloaded_batch and (
        request.values.get('update_reloaded_batch') is not None or
        request.values.get('recordpayment') is not None
    )
    is_normal_record = request.values.get('recordpayment') is not None and not is_reloaded_batch

    if is_reloaded_record:
        completed = False
        err_list = []
        holdvec = [''] * 150
        err_list, updated_batch_ids = update_reloaded_receive_batch(err_list)
        holdvec, err_list = reload_receive_batch(
            holdvec,
            [updated_batch_ids or request.values.get('reloaded_batch_ids')],
            err_list,
            use_posted_values=True,
        )
    else:
        completed, err_list, holdvec = ReceiveByAccount_task([], [''] * 150, 0)
        if completed and is_normal_record:
            flash(receive_success_message(holdvec), 'success')
            return redirect(url_for('main.ReceiveByAccount'))

    if request.values.get('create_counter_deposit') is not None:
        err_list = create_counter_deposit(err_list)
    elif request.values.get('reload_batch') is not None:
        holdvec, err_list = reload_receive_batch(holdvec, request.values.getlist('payment_batch'), err_list)
    elif is_reloaded_batch and not is_reloaded_record:
        holdvec, err_list = reload_receive_batch(
            holdvec,
            [request.values.get('reloaded_batch_ids')],
            err_list,
            use_posted_values=True,
        )

    payment_groups = {}
    rows = Gledger.query.filter(
        (Gledger.Type.in_(['DD', 'ID'])) &
        (Gledger.Com == cmpdata[10]) &
        (Gledger.Date >= batch_start) &
        (Gledger.Date < batch_end_exclusive)
    ).order_by(Gledger.Date.desc(), Gledger.id.desc()).all()

    row_ids = [row.id for row in rows]
    counter_deposits = {}
    if row_ids:
        for deposit_row in Gledger.query.filter(
            (Gledger.SourceTable == 'CounterDepositItem') &
            (Gledger.SourceId.in_(row_ids)) &
            (Gledger.Type == 'XD')
        ).all():
            counter_deposits[deposit_row.SourceId] = deposit_row

    for row in rows:
        counter_deposit = counter_deposits.get(row.id)
        key = (
            row.Date.strftime('%Y-%m-%d') if row.Date else '',
            row.Source or '',
            row.Ref or '',
            row.Account or '',
            row.Type or '',
        )
        group = payment_groups.setdefault(key, {
            'date': row.Date,
            'source': row.Source or '',
            'ref': row.Ref or '',
            'account': row.Account or '',
            'type': row.Type or '',
            'ids': [],
            'amount': 0,
            'count': 0,
            'deposited': False,
            'deposit_ref': '',
            'deposit_account': '',
            'deposit_date': None,
        })
        group['ids'].append(row.id)
        group['amount'] += row.Debit or 0
        group['count'] += 1
        if counter_deposit is not None:
            group['deposited'] = True
            group['deposit_ref'] = counter_deposit.Ref or ''
            group['deposit_account'] = counter_deposit.Account or ''
            group['deposit_date'] = counter_deposit.Date

    recent_batches = sorted(
        payment_groups.values(),
        key=lambda item: item['date'] or datetime.datetime.min,
        reverse=True,
    )
    for batch in recent_batches:
        batch['id_value'] = ','.join(str(row_id) for row_id in batch['ids'])
        batch['amount_fmt'] = "${:,.2f}".format(batch['amount'] / 100)
        batch['deposit_date_fmt'] = batch['deposit_date'].strftime('%Y-%m-%d') if batch['deposit_date'] else ''

    counter_deposit_accounts = Accounts.query.filter(
        (Accounts.Type == 'Bank') &
        (Accounts.Co == cmpdata[10])
    ).order_by(Accounts.Name.asc()).all()

    return render_template(
        'receive_by_account.html',
        cmpdata=cmpdata,
        scac=scac,
        err='\n'.join(err_list) if isinstance(err_list, list) else err_list,
        holdvec=holdvec,
        completed=completed,
        recent_batches=recent_batches,
        batch_start_date=batch_start_date,
        batch_end_date=batch_end_date,
        counter_deposit_accounts=counter_deposit_accounts,
        today_value=datetime.date.today().strftime('%Y-%m-%d'),
    )


@main.route('/GeneralDeposits', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def GeneralDeposits():
    from webapp.class8_tasks_gledger import post_balanced_journal
    from webapp.viewfuncs import newjo

    def money_to_cents(value):
        try:
            clean = str(value).replace('$', '').replace(',', '').strip()
            return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))
        except:
            return 0

    company_code = cmpdata[10]
    today_value = datetime.date.today().strftime('%Y-%m-%d')
    err = []
    selected = {
        'deposit_date': request.values.get('deposit_date', today_value),
        'deposit_kind': request.values.get('deposit_kind', 'general'),
        'bank_account': request.values.get('bank_account', ''),
        'credit_account': request.values.get('credit_account', ''),
        'amount': request.values.get('amount', ''),
        'source': request.values.get('source', ''),
        'ref': request.values.get('ref', ''),
        'memo': request.values.get('memo', ''),
        'edit_journal': request.values.get('edit_journal', ''),
    }
    if selected['deposit_kind'] not in ['general', 'expense_reimbursement']:
        selected['deposit_kind'] = 'general'

    bank_accounts = Accounts.query.filter(
        (Accounts.Type == 'Bank') &
        (Accounts.Co == company_code)
    ).order_by(Accounts.Name).all()
    credit_accounts = Accounts.query.filter(
        (Accounts.Co == company_code) &
        (Accounts.Type.in_(['Income', 'Equity', 'Current Liability']))
    ).order_by(Accounts.Type, Accounts.Name).all()
    reimbursement_accounts = Accounts.query.filter(
        (Accounts.Co == company_code) &
        (Accounts.Type == 'Expense')
    ).order_by(Accounts.Category, Accounts.Subcategory, Accounts.Name).all()
    allowed_credit_types = ['Income', 'Equity', 'Current Liability']

    def deposit_kind_for_account(account):
        return 'expense_reimbursement' if account is not None and account.Type == 'Expense' else 'general'

    def account_for_ledger_line(line):
        if line is None:
            return None
        if line.Aid:
            account = Accounts.query.get(line.Aid)
            if account is not None:
                return account
        return Accounts.query.filter(
            (Accounts.Name == (line.Account or '')) &
            (Accounts.Co == company_code)
        ).first()

    def validate_credit_account(account):
        if account is None:
            err.append('Credit account is required')
        elif selected['deposit_kind'] == 'expense_reimbursement' and account.Type != 'Expense':
            err.append('Expense reimbursement deposits must credit an expense account')
        elif selected['deposit_kind'] != 'expense_reimbursement' and account.Type not in allowed_credit_types:
            err.append('General deposits can only credit income, equity, or current liability accounts')

    if not selected['bank_account'] and bank_accounts:
        selected['bank_account'] = bank_accounts[0].Name
    if selected['deposit_kind'] == 'expense_reimbursement' and not selected['credit_account'] and reimbursement_accounts:
        selected['credit_account'] = reimbursement_accounts[0].Name
    elif not selected['credit_account'] and credit_accounts:
        selected['credit_account'] = credit_accounts[0].Name

    def load_manual_deposit(journal_id):
        debit_line = Gledger.query.filter(
            (Gledger.JournalId == journal_id) &
            (Gledger.SourceTable == 'ManualDeposit') &
            (Gledger.Com == company_code) &
            (Gledger.Type == 'DD')
        ).first()
        credit_line = Gledger.query.filter(
            (Gledger.JournalId == journal_id) &
            (Gledger.SourceTable == 'ManualDeposit') &
            (Gledger.Com == company_code) &
            (Gledger.Credit > 0)
        ).first()
        if debit_line is None or credit_line is None:
            return None
        credit_account = account_for_ledger_line(credit_line)
        return {
            'deposit_date': debit_line.Date.strftime('%Y-%m-%d') if debit_line.Date else today_value,
            'deposit_kind': deposit_kind_for_account(credit_account),
            'bank_account': debit_line.Account or '',
            'credit_account': credit_line.Account or '',
            'amount': "{:,.2f}".format((debit_line.Debit or 0) / 100),
            'source': debit_line.Source or '',
            'ref': debit_line.Ref or '',
            'memo': debit_line.JournalMemo or '',
            'edit_journal': journal_id,
        }

    if request.method == 'POST' and request.values.get('load_deposit') is not None:
        selected_journals = request.values.getlist('deposit_journal')
        if len(selected_journals) != 1:
            err.append('Select exactly one general deposit to edit')
        else:
            loaded = load_manual_deposit(selected_journals[0])
            if loaded is None:
                err.append('Selected general deposit could not be loaded for editing')
            else:
                selected = loaded
                err.append(f'Editing general deposit {selected["edit_journal"]}')

    if request.method == 'POST' and request.values.get('delete_deposits') is not None:
        selected_journals = request.values.getlist('deposit_journal')
        if not selected_journals:
            err.append('No general deposits selected for deletion')
        else:
            deleted_rows = 0
            deleted_journals = 0
            for journal_id in selected_journals:
                lines = Gledger.query.filter(
                    (Gledger.JournalId == journal_id) &
                    (Gledger.SourceTable == 'ManualDeposit') &
                    (Gledger.Com == company_code)
                ).all()
                if lines:
                    deleted_journals += 1
                    deleted_rows += len(lines)
                    for line in lines:
                        db.session.delete(line)
            db.session.commit()
            err.append(f'Deleted {deleted_journals} general deposit journal(s), {deleted_rows} ledger row(s)')

    if request.method == 'POST' and request.values.get('save_deposit') is not None:
        deposit_date = None
        try:
            deposit_date = datetime.datetime.strptime(selected['deposit_date'], '%Y-%m-%d')
        except ValueError:
            err.append('Deposit date is invalid')

        amount = money_to_cents(selected['amount'])
        if amount <= 0:
            err.append('Deposit amount must be greater than zero')

        bank = Accounts.query.filter(
            (Accounts.Name == selected['bank_account']) &
            (Accounts.Co == company_code)
        ).first()
        credit = Accounts.query.filter(
            (Accounts.Name == selected['credit_account']) &
            (Accounts.Co == company_code)
        ).first()
        debit_line = Gledger.query.filter(
            (Gledger.JournalId == selected['edit_journal']) &
            (Gledger.SourceTable == 'ManualDeposit') &
            (Gledger.Com == company_code) &
            (Gledger.Type == 'DD')
        ).first()
        credit_line = Gledger.query.filter(
            (Gledger.JournalId == selected['edit_journal']) &
            (Gledger.SourceTable == 'ManualDeposit') &
            (Gledger.Com == company_code) &
            (Gledger.Credit > 0)
        ).first()
        if bank is None:
            err.append('Bank account is required')
        validate_credit_account(credit)
        if debit_line is None or credit_line is None:
            err.append('Manual deposit journal could not be found for update')
        if not selected['source'].strip():
            err.append('Deposit source is required')

        if not err:
            memo = selected['memo'].strip() or f"General deposit from {selected['source'].strip()}"
            recorded = datetime.datetime.now()
            ref = selected['ref'].strip() or None

            debit_line.Debit = amount
            debit_line.Credit = 0
            debit_line.Account = bank.Name
            debit_line.Aid = bank.id
            debit_line.Source = selected['source'].strip()
            debit_line.Sid = credit.id
            debit_line.Recorded = recorded
            debit_line.Date = deposit_date
            debit_line.Ref = ref
            debit_line.JournalMemo = memo
            debit_line.PostedAt = recorded

            credit_line.Debit = 0
            credit_line.Credit = amount
            credit_line.Account = credit.Name
            credit_line.Aid = credit.id
            credit_line.Source = bank.Name
            credit_line.Sid = bank.id
            credit_line.Recorded = recorded
            credit_line.Date = deposit_date
            credit_line.Ref = ref
            credit_line.JournalMemo = memo
            credit_line.PostedAt = recorded

            db.session.commit()
            err.append(f'General deposit {debit_line.Tcode} updated')
            selected = {
                'deposit_date': today_value,
                'deposit_kind': 'general',
                'bank_account': bank.Name,
                'credit_account': credit.Name,
                'amount': '',
                'source': '',
                'ref': '',
                'memo': '',
                'edit_journal': '',
            }

    if request.method == 'POST' and request.values.get('create_deposit') is not None:
        deposit_date = None
        try:
            deposit_date = datetime.datetime.strptime(selected['deposit_date'], '%Y-%m-%d')
        except ValueError:
            err.append('Deposit date is invalid')

        amount = money_to_cents(selected['amount'])
        if amount <= 0:
            err.append('Deposit amount must be greater than zero')

        bank = Accounts.query.filter(
            (Accounts.Name == selected['bank_account']) &
            (Accounts.Co == company_code)
        ).first()
        credit = Accounts.query.filter(
            (Accounts.Name == selected['credit_account']) &
            (Accounts.Co == company_code)
        ).first()
        if bank is None:
            err.append('Bank account is required')
        validate_credit_account(credit)
        if not selected['source'].strip():
            err.append('Deposit source is required')

        if not err:
            tcode = newjo(f'{company_code}D', selected['deposit_date'])
            memo = selected['memo'].strip() or f"General deposit from {selected['source'].strip()}"
            journal_id = f'GDEP-{tcode}'
            recorded = datetime.datetime.now()
            err = post_balanced_journal(
                [
                    {
                        'debit': amount,
                        'credit': 0,
                        'account': bank.Name,
                        'aid': bank.id,
                        'source': selected['source'].strip(),
                        'sid': credit.id,
                        'type': 'DD',
                        'tcode': tcode,
                        'com': company_code,
                        'recorded': recorded,
                        'date': deposit_date,
                        'ref': selected['ref'].strip() or None,
                    },
                    {
                        'debit': 0,
                        'credit': amount,
                        'account': credit.Name,
                        'aid': credit.id,
                        'source': bank.Name,
                        'sid': bank.id,
                        'type': 'GC',
                        'tcode': tcode,
                        'com': company_code,
                        'recorded': recorded,
                        'date': deposit_date,
                        'ref': selected['ref'].strip() or None,
                    },
                ],
                journal_id=journal_id,
                journal_memo=memo,
                posted_by='general_deposits',
                source_table='ManualDeposit',
                source_id=None,
            )
            if not err:
                err.append(f'General deposit {tcode} recorded for ${amount / 100:,.2f}')
                selected = {
                    'deposit_date': today_value,
                    'deposit_kind': 'general',
                    'bank_account': bank.Name,
                    'credit_account': credit.Name,
                    'amount': '',
                    'source': '',
                    'ref': '',
                    'memo': '',
                    'edit_journal': '',
                }

    recent_deposits = Gledger.query.filter(
        (Gledger.Type == 'DD') &
        (Gledger.Com == company_code) &
        (Gledger.SourceTable == 'ManualDeposit')
    ).order_by(Gledger.Date.desc(), Gledger.id.desc()).limit(50).all()
    for item in recent_deposits:
        item.amount_fmt = "${:,.2f}".format((item.Debit or 0) / 100)
        credit_line = Gledger.query.filter(
            (Gledger.JournalId == item.JournalId) &
            (Gledger.Credit > 0)
        ).first()
        item.deposit_type = ''
        if credit_line is not None:
            acct = Accounts.query.get(credit_line.Aid)
            if acct is not None:
                item.deposit_type = f'{credit_line.Account} ({acct.Type})'
            else:
                item.deposit_type = credit_line.Account

    return render_template(
        'general_deposits.html',
        cmpdata=cmpdata,
        scac=scac,
        err='\n'.join(err) if err else 'Ready',
        selected=selected,
        bank_accounts=bank_accounts,
        credit_accounts=credit_accounts,
        reimbursement_accounts=reimbursement_accounts,
        recent_deposits=recent_deposits,
    )


@main.route('/LoanInterestAdjustments', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def LoanInterestAdjustments():
    from webapp.viewfuncs import newjo

    def money_to_cents(value):
        try:
            clean = str(value).replace('$', '').replace(',', '').strip()
            return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))
        except:
            return 0

    company_code = cmpdata[10]
    today = datetime.date.today()
    today_value = today.strftime('%Y-%m-%d')
    err = []
    selected = {
        'adjustment_date': request.values.get('adjustment_date', today_value),
        'loan_account': request.values.get('loan_account', ''),
        'interest_account': request.values.get('interest_account', ''),
        'period_year': request.values.get('period_year', str(today.year)),
        'amount': request.values.get('amount', ''),
        'source': request.values.get('source', 'SBA'),
        'ref': request.values.get('ref', ''),
        'memo': request.values.get('memo', ''),
    }

    loan_accounts = Accounts.query.filter(
        (Accounts.Co == company_code) &
        (Accounts.Type.in_(['Current Liability', 'Long Term Liability']))
    ).order_by(Accounts.Type, Accounts.Name).all()
    interest_accounts = Accounts.query.filter(
        (Accounts.Co == company_code) &
        (Accounts.Type == 'Expense') &
        (
            Accounts.Name.ilike('%interest%') |
            Accounts.Category.ilike('%interest%') |
            Accounts.Subcategory.ilike('%interest%')
        )
    ).order_by(Accounts.Category, Accounts.Subcategory, Accounts.Name).all()
    if not interest_accounts:
        interest_accounts = Accounts.query.filter(
            (Accounts.Co == company_code) &
            (Accounts.Type == 'Expense')
        ).order_by(Accounts.Category, Accounts.Subcategory, Accounts.Name).all()

    if not selected['loan_account'] and loan_accounts:
        selected['loan_account'] = loan_accounts[0].Name
    if not selected['interest_account'] and interest_accounts:
        selected['interest_account'] = interest_accounts[0].Name

    if request.method == 'POST' and request.values.get('delete_adjustments') is not None:
        selected_journals = request.values.getlist('adjustment_journal')
        if not selected_journals:
            err.append('No loan interest adjustments selected for deletion')
        else:
            deleted_rows = 0
            deleted_journals = 0
            blocked_journals = []
            for journal_id in selected_journals:
                lines = Gledger.query.filter(
                    (Gledger.JournalId == journal_id) &
                    (Gledger.SourceTable == 'LoanInterestAdjustment') &
                    (Gledger.Com == company_code)
                ).all()
                if not lines:
                    continue
                if any(line.Reconciled not in [None, 0, 25] for line in lines):
                    blocked_journals.append(journal_id)
                    continue
                deleted_journals += 1
                deleted_rows += len(lines)
                for line in lines:
                    db.session.delete(line)
            db.session.commit()
            if deleted_journals:
                err.append(f'Deleted {deleted_journals} loan interest adjustment journal(s), {deleted_rows} ledger row(s)')
            if blocked_journals:
                err.append(f'Cannot delete reconciled adjustment(s): {", ".join(blocked_journals)}')

    if request.method == 'POST' and request.values.get('create_adjustment') is not None:
        adjustment_date = None
        try:
            adjustment_date = datetime.datetime.strptime(selected['adjustment_date'], '%Y-%m-%d')
        except ValueError:
            err.append('Adjustment date is invalid')

        amount = money_to_cents(selected['amount'])
        if amount <= 0:
            err.append('Interest amount must be greater than zero')

        loan_account = Accounts.query.filter(
            (Accounts.Name == selected['loan_account']) &
            (Accounts.Co == company_code)
        ).first()
        interest_account = Accounts.query.filter(
            (Accounts.Name == selected['interest_account']) &
            (Accounts.Co == company_code)
        ).first()
        if loan_account is None or loan_account.Type not in ['Current Liability', 'Long Term Liability']:
            err.append('Loan liability account is required')
        if interest_account is None or interest_account.Type != 'Expense':
            err.append('Interest expense account is required')

        if not err:
            tcode = newjo(f'{company_code}L', selected['adjustment_date'])
            period = selected['period_year'].strip() or str(adjustment_date.year)
            memo = selected['memo'].strip() or f'Loan interest adjustment for {selected["source"].strip() or "loan"} {period}'
            journal_id = f'LOANINT-{tcode}'
            recorded = datetime.datetime.now()
            err = post_balanced_journal(
                [
                    {
                        'debit': amount,
                        'credit': 0,
                        'account': interest_account.Name,
                        'aid': interest_account.id,
                        'source': selected['source'].strip() or loan_account.Name,
                        'sid': loan_account.id,
                        'type': 'LI',
                        'tcode': tcode,
                        'com': company_code,
                        'recorded': recorded,
                        'date': adjustment_date,
                        'ref': selected['ref'].strip() or period,
                    },
                    {
                        'debit': 0,
                        'credit': amount,
                        'account': loan_account.Name,
                        'aid': loan_account.id,
                        'source': interest_account.Name,
                        'sid': interest_account.id,
                        'type': 'LC',
                        'tcode': tcode,
                        'com': company_code,
                        'recorded': recorded,
                        'date': adjustment_date,
                        'ref': selected['ref'].strip() or period,
                    },
                ],
                journal_id=journal_id,
                journal_memo=memo,
                posted_by='loan_interest_adjustments',
                source_table='LoanInterestAdjustment',
                source_id=None,
            )
            if not err:
                err.append(f'Loan interest adjustment {tcode} recorded for ${amount / 100:,.2f}')
                selected = {
                    'adjustment_date': today_value,
                    'loan_account': loan_account.Name,
                    'interest_account': interest_account.Name,
                    'period_year': str(today.year),
                    'amount': '',
                    'source': selected['source'],
                    'ref': '',
                    'memo': '',
                }

    recent_adjustments = Gledger.query.filter(
        (Gledger.Type == 'LI') &
        (Gledger.Com == company_code) &
        (Gledger.SourceTable == 'LoanInterestAdjustment')
    ).order_by(Gledger.Date.desc(), Gledger.id.desc()).limit(50).all()
    for item in recent_adjustments:
        item.amount_fmt = "${:,.2f}".format((item.Debit or 0) / 100)
        credit_line = Gledger.query.filter(
            (Gledger.JournalId == item.JournalId) &
            (Gledger.Type == 'LC') &
            (Gledger.Com == company_code)
        ).first()
        item.loan_account = credit_line.Account if credit_line is not None else ''

    return render_template(
        'loan_interest_adjustments.html',
        cmpdata=cmpdata,
        scac=scac,
        err='\n'.join(err) if err else 'Ready',
        selected=selected,
        loan_accounts=loan_accounts,
        interest_accounts=interest_accounts,
        recent_adjustments=recent_adjustments,
    )


@main.route('/PayrollBatches', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def PayrollBatches():
    from webapp.class8_tasks_gledger import post_balanced_journal
    from webapp.viewfuncs import newjo

    def money_to_cents(value):
        try:
            clean = str(value).replace('$', '').replace(',', '').strip()
            return int((Decimal(clean) * Decimal('100')).quantize(Decimal('1')))
        except:
            return 0

    def money(cents):
        return "${:,.2f}".format((cents or 0) / 100)

    def default_account(accounts, names):
        lowered = [(account, (account.Name or '').lower()) for account in accounts]
        for name in names:
            needle = name.lower()
            for account, account_name in lowered:
                if needle in account_name:
                    return account.Name
        return accounts[0].Name if accounts else ''

    def account_by_name(name):
        return Accounts.query.filter(
            (Accounts.Name == name) &
            (Accounts.Co == company_code)
        ).first()

    def person_display(person):
        if person.Source:
            return person.Source
        parts = [person.First or '', person.Middle or '', person.Last or '']
        return ' '.join(part for part in parts if part).strip()

    def split_person_name(name):
        parts = name.strip().split()
        first = parts[0] if parts else ''
        last = parts[-1] if len(parts) > 1 else ''
        middle = ' '.join(parts[1:-1]) if len(parts) > 2 else ''
        return first, middle, last

    def get_or_create_payroll_vendor(name):
        vendor_name = name.strip() or 'Payroll'
        vendor = People.query.filter(
            (People.Ptype == 'Vendor') &
            (People.Company == vendor_name)
        ).first()
        if vendor is not None:
            return vendor
        first, middle, last = split_person_name(vendor_name)
        vendor = People(
            Ptype='Vendor',
            Company=vendor_name,
            First=first,
            Middle=middle,
            Last=last,
            Addr1=None,
            Addr2=None,
            Addr3=None,
            Idtype=None,
            Idnumber=None,
            Telephone=None,
            Email=None,
            Associate1=None,
            Associate2=None,
            Temp1='Payroll',
            Temp2=None,
            Date1=datetime.datetime.now(),
            Date2=None,
            Source=vendor_name,
            Accountid=None,
            Saljp=None,
            Saloa=None,
            Salap=None,
        )
        db.session.add(vendor)
        db.session.flush()
        return vendor

    def add_payroll_bill_row(jo, vendor, employee, account, amount_cents, pay_date, ref, source, bank_name, memo, journal_id, pay_method):
        if amount_cents <= 0:
            return
        existing = Bills.query.filter(Bills.Jo == jo).first()
        if existing is not None:
            return
        amount = f'{amount_cents / 100:.2f}'
        bill = Bills(
            Jo=jo,
            Pid=vendor.id,
            Company=vendor.Company,
            Memo=memo[:50],
            Description=f'{employee} payroll: {account.Name}'[:600],
            bAmount=amount,
            Status='Paid',
            Scache=0,
            Source=None,
            Ref=ref,
            Date=pay_date,
            pDate=pay_date,
            pAmount=amount,
            pMulti=None,
            pAccount=bank_name,
            bAccount=account.Name,
            bType=account.Type,
            bCat=account.Category,
            bSubcat=account.Subcategory,
            Link=None,
            User='payroll_batches',
            Co=company_code,
            Temp1='PayrollBatch',
            Temp2=journal_id,
            Recurring=0,
            dDate=pay_date,
            pAmount2='0.00',
            pDate2=None,
            Proof=None,
            Check=None,
            Ccache=0,
            QBi=0,
            iflag=0,
            PmtList=amount,
            PacctList=bank_name,
            RefList=ref,
            MemoList=memo[:200],
            PdateList=pay_date.strftime('%Y-%m-%d') if pay_date else None,
            CheckList=None,
            MethList=pay_method,
            Pcache=0,
            pMeth=pay_method,
        )
        db.session.add(bill)

    def create_payroll_bill_rows(tcode, journal_id, people_rows, vendor_name, bank_name, pay_date, ref, memo):
        vendor = get_or_create_payroll_vendor(vendor_name)
        pay_method = 'ACH'
        for index, row in enumerate(people_rows, start=1):
            row_tcode = f'{tcode}-{index}'
            add_payroll_bill_row(
                f'{row_tcode}-W',
                vendor,
                row['employee'].strip(),
                row['wage_account_obj'],
                row['gross_cents'],
                pay_date,
                ref,
                vendor_name,
                bank_name,
                memo,
                journal_id,
                pay_method,
            )
            add_payroll_bill_row(
                f'{row_tcode}-T',
                vendor,
                row['employee'].strip(),
                row['tax_account_obj'],
                row['employer_tax_cents'],
                pay_date,
                ref,
                vendor_name,
                bank_name,
                memo,
                journal_id,
                pay_method,
            )
        db.session.commit()

    company_code = cmpdata[10]
    today_value = datetime.date.today().strftime('%Y-%m-%d')

    def payroll_lines_for_journal(journal_id):
        if not journal_id:
            return []
        return Gledger.query.filter(
            (Gledger.JournalId == journal_id) &
            (Gledger.SourceTable == 'PayrollBatch') &
            (Gledger.Com == company_code)
        ).order_by(Gledger.JournalSeq, Gledger.id).all()

    def payroll_has_final_reconciliation(lines):
        return any(line.Reconciled not in [None, 0, 25] for line in lines)

    def payroll_bills_for_journal(journal_id):
        return Bills.query.filter(
            (Bills.Temp1 == 'PayrollBatch') &
            (Bills.Temp2 == journal_id) &
            (Bills.Co == company_code)
        ).all()

    def payroll_selection_from_lines(lines):
        if not lines:
            return None, []
        pc_line = next((line for line in lines if line.Type == 'PC'), None)
        tax_line = next((line for line in lines if line.Type == 'XC'), None)
        first_line = lines[0]
        line_date = first_line.Date
        provider = pc_line.Source if pc_line is not None else ''
        ref = first_line.Ref or ''
        bank_account = pc_line.Account if pc_line is not None else ''
        tax_withdrawal = tax_line.Credit if tax_line is not None else 0

        grouped_rows = {}
        for line in lines:
            if line.Type not in ['PW', 'PT']:
                continue
            row = grouped_rows.setdefault(line.Tcode, {
                'employee': line.Source or '',
                'wage_account': selected['wage_account'],
                'tax_account': selected['tax_account'],
                'gross_pay': '',
                'net_pay': '',
                'employer_tax': '',
                'gross_cents': 0,
                'net_cents': 0,
                'employer_tax_cents': 0,
            })
            if line.Type == 'PW':
                row['wage_account'] = line.Account
                row['gross_cents'] = line.Debit or 0
                row['gross_pay'] = money(line.Debit or 0).replace('$', '')
            elif line.Type == 'PT':
                row['tax_account'] = line.Account
                row['employer_tax_cents'] = line.Debit or 0
                row['employer_tax'] = money(line.Debit or 0).replace('$', '')

        selection = {
            'pay_date': line_date.strftime('%Y-%m-%d') if line_date else today_value,
            'provider': provider or '',
            'ref': ref,
            'bank_account': bank_account,
            'tax_withdrawal': money(tax_withdrawal).replace('$', ''),
            'memo': first_line.JournalMemo or '',
            'edit_payroll_journal_id': first_line.JournalId or '',
        }
        rows = list(grouped_rows.values()) or [
            {'employee': '', 'wage_account': selected['wage_account'], 'tax_account': selected['tax_account'], 'gross_pay': '', 'net_pay': '', 'employer_tax': '', 'gross_cents': 0, 'net_cents': 0, 'employer_tax_cents': 0}
        ]
        return selection, rows

    err = []
    selected = {
        'pay_date': request.values.get('pay_date', today_value),
        'provider': request.values.get('provider', ''),
        'ref': request.values.get('ref', ''),
        'bank_account': request.values.get('bank_account', ''),
        'wage_account': request.values.get('wage_account', ''),
        'tax_account': request.values.get('tax_account', ''),
        'tax_withdrawal': request.values.get('tax_withdrawal', ''),
        'memo': request.values.get('memo', ''),
        'edit_payroll_journal_id': request.values.get('edit_payroll_journal_id', ''),
    }

    bank_accounts = Accounts.query.filter(
        (Accounts.Type == 'Bank') &
        (Accounts.Co == company_code)
    ).order_by(Accounts.Name).all()
    expense_accounts = Accounts.query.filter(
        (Accounts.Type == 'Expense') &
        (Accounts.Co == company_code)
    ).order_by(Accounts.Name).all()

    if not selected['bank_account']:
        selected['bank_account'] = default_account(bank_accounts, ['bank of america', 'bank'])
    if not selected['wage_account']:
        selected['wage_account'] = default_account(expense_accounts, ['owner payroll expense', 'payroll expense', 'labor expense', 'wage'])
    if not selected['tax_account']:
        selected['tax_account'] = default_account(expense_accounts, ['payroll taxes expense', 'federal taxes expense', 'tax'])

    if request.method == 'POST' and request.values.get('add_payroll_person') is not None:
        person_name = request.values.get('new_person_name', '').strip()
        person_account_name = request.values.get('new_person_account', selected['wage_account'])
        person_tax_account_name = request.values.get('new_person_tax_account', selected['tax_account'])
        person_account = account_by_name(person_account_name)
        person_tax_account = account_by_name(person_tax_account_name)
        if not person_name:
            err.append('Payroll person name is required')
        if person_account is None:
            err.append('Payroll person gross pay expense account is required')
        if person_tax_account is None:
            err.append('Payroll person tax expense account is required')
        if not err:
            first, middle, last = split_person_name(person_name)
            person = People.query.filter(
                (People.Ptype == 'Payroll') &
                (People.Company == company_code) &
                (People.Source == person_name)
            ).first()
            if person is None:
                person = People(
                    Ptype='Payroll',
                    Company=company_code,
                    First=first,
                    Middle=middle,
                    Last=last,
                    Addr1=None,
                    Addr2=None,
                    Addr3=None,
                    Idtype=None,
                    Idnumber=None,
                    Telephone=None,
                    Email=None,
                    Associate1=None,
                    Associate2=None,
                    Temp1='Active',
                    Temp2=str(person_tax_account.id),
                    Date1=datetime.datetime.now(),
                    Date2=None,
                    Source=person_name,
                    Accountid=person_account.id,
                    Saljp=None,
                    Saloa=None,
                    Salap=None,
                )
                db.session.add(person)
                err.append(f'Added payroll person {person_name}')
            else:
                person.First = first
                person.Middle = middle
                person.Last = last
                person.Accountid = person_account.id
                person.Temp2 = str(person_tax_account.id)
                person.Temp1 = 'Active'
                err.append(f'Updated payroll person {person_name}')
            db.session.commit()

    if request.method == 'POST' and request.values.get('update_payroll_roster') is not None:
        roster_ids = request.values.getlist('roster_person_id[]')
        roster_names = request.values.getlist('roster_person_name[]')
        roster_accounts = request.values.getlist('roster_person_account[]')
        roster_tax_accounts = request.values.getlist('roster_person_tax_account[]')
        updated = 0
        for index, person_id in enumerate(roster_ids):
            person = People.query.filter(
                (People.id == person_id) &
                (People.Ptype == 'Payroll') &
                (People.Company == company_code)
            ).first()
            if person is None:
                continue
            person_name = roster_names[index].strip() if index < len(roster_names) else person_display(person)
            account_name = roster_accounts[index] if index < len(roster_accounts) else selected['wage_account']
            tax_account_name = roster_tax_accounts[index] if index < len(roster_tax_accounts) else selected['tax_account']
            account = account_by_name(account_name)
            tax_account_row = account_by_name(tax_account_name)
            if not person_name:
                err.append(f'Payroll roster row {index + 1} is missing person name')
                continue
            if account is None:
                err.append(f'Payroll roster row {index + 1} has invalid gross pay expense account')
                continue
            if tax_account_row is None:
                err.append(f'Payroll roster row {index + 1} has invalid tax expense account')
                continue
            first, middle, last = split_person_name(person_name)
            person.First = first
            person.Middle = middle
            person.Last = last
            person.Source = person_name
            person.Accountid = account.id
            person.Temp2 = str(tax_account_row.id)
            person.Temp1 = 'Active'
            updated += 1
        if updated:
            db.session.commit()
            err.append(f'Updated {updated} payroll roster person(s)')

    if request.method == 'POST' and request.values.get('deactivate_payroll_people') is not None:
        selected_people = request.values.getlist('roster_select')
        if not selected_people:
            err.append('Select at least one payroll person to deactivate')
        else:
            count = 0
            for person_id in selected_people:
                person = People.query.filter(
                    (People.id == person_id) &
                    (People.Ptype == 'Payroll') &
                    (People.Company == company_code)
                ).first()
                if person is not None:
                    person.Temp1 = 'Inactive'
                    count += 1
            db.session.commit()
            err.append(f'Deactivated {count} payroll roster person(s)')

    payroll_people = People.query.filter(
        (People.Ptype == 'Payroll') &
        (People.Company == company_code) &
        ((People.Temp1 != 'Inactive') | (People.Temp1.is_(None)))
    ).order_by(People.Source, People.Last, People.First).all()
    person_account_defaults = {}
    person_tax_account_defaults = {}
    for person in payroll_people:
        display = person_display(person)
        account = Accounts.query.get(person.Accountid) if person.Accountid else None
        try:
            tax_account = Accounts.query.get(int(person.Temp2)) if person.Temp2 else None
        except:
            tax_account = None
        if display:
            person.display_name = display
            person.default_account = account.Name if account is not None else selected['wage_account']
            person.default_tax_account = tax_account.Name if tax_account is not None else selected['tax_account']
            person_account_defaults[display] = person.default_account
            person_tax_account_defaults[display] = person.default_tax_account

    posted_people = request.values.getlist('employee[]')
    posted_wage_accounts = request.values.getlist('wage_account[]')
    posted_tax_accounts = request.values.getlist('tax_account[]')
    posted_gross = request.values.getlist('gross_pay[]')
    posted_net = request.values.getlist('net_pay[]')
    posted_employer_tax = request.values.getlist('employer_tax[]')
    people_rows = []
    row_count = max(len(posted_people), len(posted_wage_accounts), len(posted_tax_accounts), len(posted_gross), len(posted_net), len(posted_employer_tax), 3)
    for index in range(row_count):
        person = posted_people[index] if index < len(posted_people) else ''
        wage_account_name = posted_wage_accounts[index] if index < len(posted_wage_accounts) else ''
        tax_account_name = posted_tax_accounts[index] if index < len(posted_tax_accounts) else ''
        if not wage_account_name:
            wage_account_name = person_account_defaults.get(person, selected['wage_account'])
        if not tax_account_name:
            tax_account_name = person_tax_account_defaults.get(person, selected['tax_account'])
        gross = posted_gross[index] if index < len(posted_gross) else ''
        net = posted_net[index] if index < len(posted_net) else ''
        employer_tax_row = posted_employer_tax[index] if index < len(posted_employer_tax) else ''
        if request.method == 'POST' or index < 3:
            people_rows.append({
                'employee': person,
                'wage_account': wage_account_name,
                'tax_account': tax_account_name,
                'gross_pay': gross,
                'net_pay': net,
                'employer_tax': employer_tax_row,
                'gross_cents': money_to_cents(gross),
                'net_cents': money_to_cents(net),
                'employer_tax_cents': money_to_cents(employer_tax_row),
            })
    if request.method == 'POST':
        people_rows = [
            row for row in people_rows
            if row['employee'].strip() or row['gross_cents'] or row['net_cents'] or row['employer_tax_cents']
        ]
        if not people_rows:
            people_rows = [{'employee': '', 'wage_account': selected['wage_account'], 'tax_account': selected['tax_account'], 'gross_pay': '', 'net_pay': '', 'employer_tax': '', 'gross_cents': 0, 'net_cents': 0, 'employer_tax_cents': 0}]

    if request.method == 'POST' and request.values.get('load_payroll_batch') is not None:
        journal_id = request.values.get('selected_payroll_journal_id', '').strip()
        lines = payroll_lines_for_journal(journal_id)
        selection, loaded_rows = payroll_selection_from_lines(lines)
        if selection is None:
            err.append('Choose a payroll batch to edit.')
        else:
            selected.update(selection)
            people_rows = loaded_rows
            err.append(f'Loaded payroll batch {lines[0].Tcode} for editing.')

    if request.method == 'POST' and request.values.get('delete_payroll_batch') is not None:
        journal_id = request.values.get('selected_payroll_journal_id', '').strip()
        lines = payroll_lines_for_journal(journal_id)
        if not lines:
            err.append('Choose a payroll batch to delete.')
        elif payroll_has_final_reconciliation(lines):
            err.append('This payroll batch has been reconciled. Reopen the reconciliation statement before deleting it.')
        else:
            tcode = lines[0].Tcode
            for bill in payroll_bills_for_journal(journal_id):
                db.session.delete(bill)
            for line in lines:
                db.session.delete(line)
            db.session.commit()
            selected['edit_payroll_journal_id'] = ''
            err.append(f'Deleted payroll batch {tcode}.')

    if request.method == 'POST' and request.values.get('update_payroll_batch') is not None:
        journal_id = selected['edit_payroll_journal_id'].strip()
        lines = payroll_lines_for_journal(journal_id)
        try:
            pay_date = datetime.datetime.strptime(selected['pay_date'], '%Y-%m-%d')
        except:
            pay_date = None
            err.append('Pay date is invalid')
        bank = Accounts.query.filter(
            (Accounts.Name == selected['bank_account']) &
            (Accounts.Co == company_code)
        ).first()
        if not lines:
            err.append('The payroll batch being edited could not be found.')
        elif payroll_has_final_reconciliation(lines):
            err.append('This payroll batch has been reconciled. Reopen the reconciliation statement before editing it.')
        if bank is None:
            err.append('Bank account is required')

        if not err:
            source = selected['provider'].strip() or 'Payroll'
            memo = selected['memo'].strip() or f'Payroll batch {lines[0].Tcode}'
            ref = selected['ref'].strip() or None
            for line in lines:
                line.Date = pay_date
                line.Ref = ref
                line.JournalMemo = memo
                line.PostedAt = datetime.datetime.now()
                if line.Type in ['PC', 'XC']:
                    line.Account = bank.Name
                    line.Aid = bank.id
                    line.Source = source if line.Type == 'PC' else f'{source} payroll taxes'
                if line.Type in ['PW', 'PT']:
                    line.Sid = bank.id
            for bill in payroll_bills_for_journal(journal_id):
                bill.Ref = ref
                bill.Date = pay_date
                bill.pDate = pay_date
                bill.dDate = pay_date
                bill.pAccount = bank.Name
                bill.PacctList = bank.Name
                bill.Memo = memo[:50]
                bill.MemoList = memo[:200]
                bill.PdateList = pay_date.strftime('%Y-%m-%d') if pay_date else None
            db.session.commit()
            err.append(f'Updated payroll batch {lines[0].Tcode}.')
            selected['edit_payroll_journal_id'] = ''

    gross_pay = sum(row['gross_cents'] for row in people_rows)
    net_pay = sum(row['net_cents'] for row in people_rows)
    employer_tax = sum(row['employer_tax_cents'] for row in people_rows)
    tax_withdrawal = money_to_cents(selected['tax_withdrawal'])
    employee_withholding = gross_pay - net_pay
    debit_total = gross_pay + employer_tax
    credit_total = net_pay + tax_withdrawal
    balance_difference = debit_total - credit_total
    tax_expected = employee_withholding + employer_tax
    tax_difference = tax_withdrawal - tax_expected

    if request.method == 'POST' and request.values.get('record_payroll') is not None:
        try:
            pay_date = datetime.datetime.strptime(selected['pay_date'], '%Y-%m-%d')
        except:
            pay_date = None
            err.append('Pay date is invalid')

        bank = Accounts.query.filter(
            (Accounts.Name == selected['bank_account']) &
            (Accounts.Co == company_code)
        ).first()
        wage_account = account_by_name(selected['wage_account'])
        tax_account = Accounts.query.filter(
            (Accounts.Name == selected['tax_account']) &
            (Accounts.Co == company_code)
        ).first()
        if bank is None:
            err.append('Bank account is required')
        if wage_account is None:
            err.append('Gross wages expense account is required')
        if tax_account is None:
            err.append('Employer payroll tax expense account is required')
        if gross_pay <= 0:
            err.append('Gross pay must be greater than zero')
        if net_pay <= 0:
            err.append('Net pay withdrawal must be greater than zero')
        if tax_withdrawal < 0 or employer_tax < 0:
            err.append('Tax amounts cannot be negative')
        for index, row in enumerate(people_rows, start=1):
            if not row['employee'].strip():
                err.append(f'Payroll row {index} is missing employee/person')
            row_account = account_by_name(row['wage_account'])
            row_tax_account = account_by_name(row['tax_account'])
            row['wage_account_obj'] = row_account
            row['tax_account_obj'] = row_tax_account
            if row_account is None:
                err.append(f'Payroll row {index} expense account is required')
            if row_tax_account is None:
                err.append(f'Payroll row {index} tax expense account is required')
            if row['gross_cents'] <= 0:
                err.append(f'Payroll row {index} gross pay must be greater than zero')
            if row['net_cents'] < 0 or row['employer_tax_cents'] < 0:
                err.append(f'Payroll row {index} amounts cannot be negative')
            if row['net_cents'] > row['gross_cents']:
                err.append(f'Payroll row {index} net pay cannot exceed gross pay')
        if balance_difference != 0:
            err.append(f'Payroll batch is not balanced by {money(balance_difference)}')

        if not err:
            tcode = newjo(f'{company_code}P', selected['pay_date'])
            source = selected['provider'].strip() or 'Payroll'
            memo = selected['memo'].strip() or f"Payroll batch for {len(people_rows)} person(s)"
            journal_id = f'PAYROLL-{tcode}'
            recorded = datetime.datetime.now()
            lines = []
            for index, row in enumerate(people_rows, start=1):
                row_tcode = f'{tcode}-{index}'
                lines.extend([
                    {
                        'debit': row['gross_cents'],
                        'credit': 0,
                        'account': row['wage_account_obj'].Name,
                        'aid': row['wage_account_obj'].id,
                        'source': row['employee'].strip(),
                        'sid': bank.id,
                        'type': 'PW',
                        'tcode': row_tcode,
                        'com': company_code,
                        'recorded': recorded,
                        'date': pay_date,
                        'ref': selected['ref'].strip() or None,
                    },
                    {
                        'debit': row['employer_tax_cents'],
                        'credit': 0,
                        'account': row['tax_account_obj'].Name,
                        'aid': row['tax_account_obj'].id,
                        'source': row['employee'].strip(),
                        'sid': bank.id,
                        'type': 'PT',
                        'tcode': row_tcode,
                        'com': company_code,
                        'recorded': recorded,
                        'date': pay_date,
                        'ref': selected['ref'].strip() or None,
                    },
                ])
            lines.extend([
                {
                    'debit': 0,
                    'credit': net_pay,
                    'account': bank.Name,
                    'aid': bank.id,
                    'source': source,
                    'sid': wage_account.id,
                    'type': 'PC',
                    'tcode': tcode,
                    'com': company_code,
                    'recorded': recorded,
                    'date': pay_date,
                    'ref': selected['ref'].strip() or None,
                },
                {
                    'debit': 0,
                    'credit': tax_withdrawal,
                    'account': bank.Name,
                    'aid': bank.id,
                    'source': f'{source} payroll taxes',
                    'sid': tax_account.id,
                    'type': 'XC',
                    'tcode': tcode,
                    'com': company_code,
                    'recorded': recorded,
                    'date': pay_date,
                    'ref': selected['ref'].strip() or None,
                },
            ])
            err = post_balanced_journal(
                lines,
                journal_id=journal_id,
                journal_memo=memo,
                posted_by='payroll_batches',
                source_table='PayrollBatch',
                source_id=None,
            )
            if not err:
                create_payroll_bill_rows(
                    tcode=tcode,
                    journal_id=journal_id,
                    people_rows=people_rows,
                    vendor_name=source,
                    bank_name=bank.Name,
                    pay_date=pay_date,
                    ref=selected['ref'].strip() or None,
                    memo=memo,
                )
                err.append(f'Payroll batch {tcode} recorded. Net pay and tax withdrawals will appear in banking reconciliation.')
                selected.update({
                    'pay_date': today_value,
                    'provider': '',
                    'ref': '',
                    'tax_withdrawal': '',
                    'memo': '',
                })
                people_rows = [
                    {'employee': '', 'wage_account': selected['wage_account'], 'tax_account': selected['tax_account'], 'gross_pay': '', 'net_pay': '', 'employer_tax': '', 'gross_cents': 0, 'net_cents': 0, 'employer_tax_cents': 0},
                    {'employee': '', 'wage_account': selected['wage_account'], 'tax_account': selected['tax_account'], 'gross_pay': '', 'net_pay': '', 'employer_tax': '', 'gross_cents': 0, 'net_cents': 0, 'employer_tax_cents': 0},
                    {'employee': '', 'wage_account': selected['wage_account'], 'tax_account': selected['tax_account'], 'gross_pay': '', 'net_pay': '', 'employer_tax': '', 'gross_cents': 0, 'net_cents': 0, 'employer_tax_cents': 0},
                ]

    payroll_groups = {}
    recent_lines = Gledger.query.filter(
        (Gledger.SourceTable == 'PayrollBatch') &
        (Gledger.Com == company_code)
    ).order_by(Gledger.Date.desc(), Gledger.id.desc()).limit(400).all()
    for line in recent_lines:
        group = payroll_groups.setdefault(line.JournalId or line.Tcode, {
            'journal_id': line.JournalId or '',
            'tcode': line.Tcode or '',
            'date': line.Date,
            'memo': line.JournalMemo or '',
            'reconciled': False,
            'gross': 0,
            'employer_tax': 0,
            'net': 0,
            'tax_withdrawal': 0,
            'people': set(),
        })
        if line.Reconciled not in [None, 0, 25]:
            group['reconciled'] = True
        if line.Type == 'PW':
            group['gross'] += line.Debit or 0
            if line.Source:
                group['people'].add(line.Source)
        elif line.Type == 'PT':
            group['employer_tax'] += line.Debit or 0
            if line.Source:
                group['people'].add(line.Source)
        elif line.Type == 'PC':
            group['net'] += line.Credit or 0
        elif line.Type == 'XC':
            group['tax_withdrawal'] += line.Credit or 0
        if line.Date and (not group['date'] or line.Date > group['date']):
            group['date'] = line.Date

    recent_batches = sorted(
        payroll_groups.values(),
        key=lambda item: item['date'] or datetime.datetime.min,
        reverse=True,
    )
    for batch in recent_batches:
        batch['gross_fmt'] = money(batch['gross'])
        batch['employer_tax_fmt'] = money(batch['employer_tax'])
        batch['net_fmt'] = money(batch['net'])
        batch['tax_withdrawal_fmt'] = money(batch['tax_withdrawal'])
        batch['bank_total_fmt'] = money(batch['net'] + batch['tax_withdrawal'])
        batch['people_count'] = len(batch['people'])
        batch['people_text'] = ', '.join(sorted(batch['people']))

    summary = {
        'employee_withholding': money(employee_withholding),
        'employer_tax': money(employer_tax),
        'expected_tax_withdrawal': money(tax_expected),
        'tax_difference': money(tax_difference),
        'debit_total': money(debit_total),
        'credit_total': money(credit_total),
        'balance_difference': money(balance_difference),
        'balanced': balance_difference == 0 and debit_total > 0,
    }

    return render_template(
        'payroll_batches.html',
        cmpdata=cmpdata,
        scac=scac,
        err='\n'.join(err) if err else 'Ready',
        selected=selected,
        people_rows=people_rows,
        payroll_people=payroll_people,
        person_account_defaults=person_account_defaults,
        person_tax_account_defaults=person_tax_account_defaults,
        bank_accounts=bank_accounts,
        expense_accounts=expense_accounts,
        recent_batches=recent_batches,
        summary=summary,
    )



@main.route('/EasyStart', methods=['GET', 'POST'])
def EasyStart():
    calbut=request.values.get('calbut')
    if calbut is not None:
        return redirect(url_for('main.CalendarBig'))
    #print('Working the EasyStart route!!')
    srcpath = statpath('')
    return render_template('easystart.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac)


#@app.route('/hello', methods=['GET', 'POST'])
#def hello():

    # POST request
    #    if request.method == 'POST':
    #       print('Incoming..')
    #      print(request.get_json())  # parse as JSON
    #      return 'OK', 200

    # GET request
    #   else:
    #       message = {'greeting':'Hello from Flask!'}
#       return jsonify(message)  # serialize and use JSON headers

#@app.route('/CalendarTest', methods=['GET', 'POST'])
#def CalendarTest():
#    print('Working the Calendar route!!')
#    srcpath = statpath('')
#    return render_template('CalendarTest.html', srcpath=srcpath, cmpdata=cmpdata, scac=scac)

@main.route('/QuoteMaker', methods=['GET', 'POST'])
@login_required
def QuoteMaker():
    from webapp.iso_Q import isoQuote, step_timer
    with step_timer("isoQuote"):
        (bidname, costdata, biddata, expdata, timedata, distdata, emaildata, locto, locfrom, dirdata, qdata, bidthis, taskbox, thismuch, quot,
        qdat, tbox, ebodytxt, multibid, newmarkup, whouse, sboxes, htmltext, send_mode, reply_style, save_sent, equip) = isoQuote()
    if bidname == 'exitnow': return redirect(url_for('main.Class8Main',genre='Trucking'))
    else:
        with step_timer("render_template"):
            return render_template('Aquotemaker.html', cmpdata=cmpdata, scac=scac, costdata = costdata, biddata=biddata, expdata=expdata, timedata=timedata, whouse=whouse, sboxes = sboxes,
                           distdata=distdata, locto=locto, locfrom=locfrom, emaildata=emaildata, dirdata=dirdata, qdata = qdata, bidthis=bidthis, taskbox=taskbox, thismuch=thismuch, quot=quot, qdat=qdat,
                           bidname=bidname, tbox=tbox, ebodytxt=ebodytxt, multibid=multibid, newmarkup=newmarkup, htmltext=htmltext, send_mode=send_mode, reply_style=reply_style, save_sent=save_sent, equip=equip)

@main.route('/ARMaker', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def ARMaker():
    from iso_AR import isoAR
    status, ardata, arsent, this_shipper, odata, sdata, task, emaildata, boxes, sboxes, tboxes, invoname, packname, pdat, emailsend, ar_emails_cust, rview, lookbacktime= isoAR()
    if status == 'exitnow': return redirect(url_for('main.Class8Main',genre='Trucking'))
    else:
        return render_template('ARmaker.html', cmpdata=cmpdata, scac=scac, ardata=ardata, arsent=arsent, this_shipper=this_shipper, odata=odata, sdata=sdata, task=task, emaildata=emaildata, boxes=boxes, sboxes=sboxes, tboxes=tboxes, invoname=invoname, packname=packname, pdat=pdat, emailsend=emailsend, ar_emails_cust=ar_emails_cust, rview=rview, lookbacktime=lookbacktime)

@main.route('/ARPayments', methods=['GET', 'POST'])
@login_required
@financial_mfa_required
def ARPayments():
    from iso_Pay import isoPay
    status, this_id, odata, pdata, tot = isoPay()
    if status == 'exitnow': return redirect(url_for('main.Class8Main',genre='Trucking'))
    else:
        return render_template('ARpayments.html', cmpdata=cmpdata, scac=scac, this_id=this_id, odata=odata, pdata=pdata, tot=tot)

@main.route('/Reports', methods=['GET', 'POST'])
@login_required
def Reports():

    from iso_R import isoR
    idata1, idata2, idata3, idata4, hv, cache, err, leftscreen, docref, leftsize, today, now, doctxt, sdate, fdate, fyear, customerlist, thiscomp, clist = isoR()
    rightsize = 12-leftsize
    return render_template('Areports.html', cmpdata=cmpdata, scac=scac, clist=clist, thiscomp=thiscomp, customerlist=customerlist, fyear=fyear, cache=cache, sdate=sdate, fdate=fdate, err=err, doctxt=doctxt, leftscreen=leftscreen, docref=docref, leftsize=leftsize, rightsize=rightsize, idata1 = idata1, idata2=idata2, idata3=idata3, idata4=idata4, hv=hv)
